import os
import time
import subprocess
import uiautomator2 as u2
import sys
import datetime
import re
import xml.etree.ElementTree as ET
import openpyxl


from konfigurasi import (
    LDPLAYER_DNCONSOLE,
    LDPLAYER_ADB,
    EMULATOR_INDEX_1 as EMULATOR_INDEX,
    EMULATOR_PORTS_1 as EMULATOR_PORTS,
    SLEEP_SHORT,
    SLEEP_MEDIUM,
    SLEEP_LONG,
    CUSTOM_COLUMNS,
    FILTER_STATUS_TARGET,
)

# Variabel global untuk koneksi emulator
d = None


def hubungkan_emulator():
    """Menghubungkan uiautomator2 ke emulator LDPlayer"""
    global d
    print("[KONEKSI] Menghubungkan ke emulator...")
    for port in EMULATOR_PORTS:
        try:
            if os.path.exists(LDPLAYER_ADB):
                subprocess.run(
                    [LDPLAYER_ADB, "connect", f"127.0.0.1:{port}"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL
                )
            temp_d = u2.connect(f"127.0.0.1:{port}")
            _ = temp_d.info
            d = temp_d
            print(f"[KONEKSI] Berhasil terhubung ke emulator di port {port}!")
            return True
        except Exception:
            continue

    print("[ERROR] Gagal terhubung ke emulator. Pastikan LDPlayer sudah aktif.")
    return False


def swipe_aman(fx, fy, tx, ty, duration=0.1):
    """Melakukan swipe dengan uiautomator2 dan fallback aman ke ADB shell jika terjadi SecurityException / RPC error"""
    try:
        d.swipe(fx, fy, tx, ty, duration=duration)
    except Exception as e:
        duration_ms = int(duration * 1000)
        try:
            d.shell(f"input swipe {fx} {fy} {tx} {ty} {duration_ms}")
        except Exception as shell_err:
            print(f"[WARNING] Gagal swipe via ADB shell: {shell_err}")


def cek_dan_tangani_kebablasan():
    """
    Langkah 1: Cek jika kebablasan ke halaman 'Periode' atau 'Daftar Wilayah'
    dan kembali ke 'Daftar Assignment'.
    """
    print("\n[LANGKAH 1] Memeriksa apakah kebablasan ke halaman 'Periode' atau 'Daftar Wilayah'...")
    try:
        is_periode = d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Periode").exists()
        is_wilayah = d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Daftar Wilayah").exists()

        if is_periode or is_wilayah:
            current_page = "Periode" if is_periode else "Daftar Wilayah"
            print(f"[NAV] Terdeteksi kebablasan di halaman '{current_page}'.")

            submit_btn = d(text="Submit")
            if submit_btn.exists(timeout=2):
                print("[NAV] Mengetuk tombol 'Submit'...")
                submit_btn.click()
                time.sleep(SLEEP_LONG)

        # Cek jika berada di halaman SLS Wilayah (updateListingLayout)
        region_row = d(resourceId="id.go.bpsfasih:id/updateListingLayout")
        if region_row.exists():
            print("[NAV] Berada di halaman SLS Wilayah. Mengklik wilayah untuk masuk ke Daftar Assignment...")
            region_row.click()
            time.sleep(SLEEP_LONG)

        # Tunggu sampai halaman Daftar Assignment termuat
        title_el = d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Daftar Assignment")
        if title_el.wait(exists=True, timeout=10.0):
            print("[NAV] Berhasil berada di halaman 'Daftar Assignment'.")
            return True
        else:
            print("[WARNING] Halaman 'Daftar Assignment' tidak terdeteksi secara pasti, melanjutkan...")
            return True
    except Exception as e:
        print(f"[ERROR] Gagal saat memeriksa kebablasan halaman: {e}")
        return False


def ketuk_fab_dan_filter():
    """
    Langkah 2: Ketuk FAB & Ketuk Filter By Status
    """
    print("\n[LANGKAH 2] Mengetuk FAB dan memilih 'Filter By Status'...")
    try:
        # 1. Ketuk FAB (Expandable FAB)
        fab = d(resourceId="id.go.bpsfasih:id/expendable_fab")
        if not fab.exists():
            fab = d(descriptionContains="FAB")

        if fab.wait(exists=True, timeout=5.0):
            print("[FAB] Mengetuk tombol FAB (expendable_fab)...")
            fab.click()
            time.sleep(SLEEP_SHORT)
        else:
            print("[ERROR] Tombol FAB tidak ditemukan di layar.")
            return False

        # 2. Ketuk Filter By Status
        filter_btn = d(resourceId="id.go.bpsfasih:id/fab_filterAssignment")
        if not filter_btn.exists():
            filter_btn = d(text="Filter By Status")
        if not filter_btn.exists():
            filter_btn = d(textContains="Filter")

        if filter_btn.wait(exists=True, timeout=5.0):
            print("[FAB] Mengetuk opsi 'Filter By Status'...")
            filter_btn.click()
            time.sleep(SLEEP_MEDIUM)
            return True
        else:
            print("[ERROR] Opsi 'Filter By Status' tidak ditemukan setelah mengetuk FAB.")
            return False
    except Exception as e:
        print(f"[ERROR] Gagal saat mengetuk FAB / Filter By Status: {e}")
        return False


def terapkan_filter_reject():
    """
    Langkah 3: Mengatur filter status sesuai variabel FILTER_STATUS_TARGET
    (misal: Open, Pernah dibuka, Submit, Approve, Reject) lalu ketuk tombol 'TERAPKAN'.
    """
    print(f"\n[LANGKAH 3] Mengatur filter status ke {FILTER_STATUS_TARGET}...")
    try:
        # Cek apakah dialog filter terbuka, jika tidak, buka kembali via FAB
        filter_title = d(text="Filter Assignment By Status")
        if not filter_title.exists():
            print("[INFO] Dialog filter belum terbuka. Membuka dialog filter...")
            if not ketuk_fab_dan_filter():
                print("[ERROR] Gagal membuka dialog filter via FAB.")
                return False

        all_checkboxes = [
            ("Open", "id.go.bpsfasih:id/open_cb_bottomSheetFilterAssignment"),
            ("Pernah dibuka", "id.go.bpsfasih:id/pernahDibuka_cb_bottomSheetFilterAssignment"),
            ("Submit", "id.go.bpsfasih:id/submit_cb_bottomSheetFilterAssignment"),
            ("Approve", "id.go.bpsfasih:id/approve_cb_bottomSheetFilterAssignment"),
            ("Reject", "id.go.bpsfasih:id/reject_cb_bottomSheetFilterAssignment"),
        ]

        def cari_cb(label, res_id):
            cb = d(resourceId=res_id)
            if cb.exists():
                return cb
            cb = d(text=label)
            if cb.exists():
                return cb
            cb = d(textContains=label)
            if cb.exists():
                return cb
            return None

        # Atur setiap status checkbox sesuai target di FILTER_STATUS_TARGET
        for label, res_id in all_checkboxes:
            harus_centang = label in FILTER_STATUS_TARGET
            cb = cari_cb(label, res_id)

            if not cb:
                print(f"[WARNING] Checkbox '{label}' tidak ditemukan di dialog filter.")
                continue

            for attempt in range(1, 4):
                cb = cari_cb(label, res_id)
                if not cb:
                    break

                info = cb.info if hasattr(cb, 'info') else {}
                is_checked = info.get("checked", False)

                if harus_centang:
                    if not is_checked:
                        print(f"[FILTER] Mencentang checkbox '{label}' (Percobaan {attempt}/3)...")
                        cb.click()
                        time.sleep(0.3)
                    else:
                        print(f"[FILTER] [OK] Checkbox '{label}' sudah tercentang (checked=True).")
                        break
                else:
                    if is_checked:
                        print(f"[FILTER] Unchecking checkbox '{label}' (Percobaan {attempt}/3)...")
                        cb.click()
                        time.sleep(0.3)
                    else:
                        print(f"[FILTER] [OK] Checkbox '{label}' sudah bersih (checked=False).")
                        break

        # Ketuk tombol 'TERAPKAN'
        btn_terapkan = d(resourceId="id.go.bpsfasih:id/rButton_bottomDialogFilterAssignment")
        if not btn_terapkan.exists():
            btn_terapkan = d(text="TERAPKAN")

        if btn_terapkan.wait(exists=True, timeout=3.0):
            print("[FILTER] Mengetuk tombol 'TERAPKAN'...")
            btn_terapkan.click()
            time.sleep(SLEEP_LONG)
            print(f"[SUKSES] Filter status {FILTER_STATUS_TARGET} berhasil diterapkan!")
            return True
        else:
            print("[ERROR] Tombol 'TERAPKAN' tidak ditemukan.")
            return False

    except Exception as e:
        print(f"[ERROR] Gagal saat mengatur filter status: {e}")
        return False


def ekstrak_data_tabel():
    """
    Mengambil data baris tabel (No. Meter, Nama, ID Pelanggan) dari layar saat ini.
    Mendukung berbagai variasi hirarki XML (seperti dump.xml, dump copy 2.xml, dump copy 3.xml).
    Returns: list of tuple (no_meter, nama, id_pelanggan)
    """
    try:
        xml = d.dump_hierarchy()
        tree = ET.fromstring(xml)
    except Exception as e:
        print(f"[WARNING] Gagal membaca hierarki tabel: {e}")
        return []

    # 1. Cari container GridView / Tabel yang mengandung 'Showing ... entries'
    grid_container = None
    for elem in tree.iter('node'):
        hint = elem.attrib.get('hint', '')
        if 'Showing' in hint and 'entries' in hint:
            grid_container = elem
            break

    search_root = grid_container if grid_container is not None else tree

    # 2. Rekursif mengambil node teks (Abaikan sub-child jika parent sudah memiliki teks non-empty)
    nodes = []

    def collect_nodes(elem):
        txt = elem.attrib.get('text', '').strip()
        bounds = elem.attrib.get('bounds', '')

        if txt and bounds:
            # Filter elemen non-tabel dan header
            if not any(k in txt for k in ['Daftar Assignment', 'Search:', 'Show', 'entries', 'Previous', 'Next', 'Filter By', 'Showing', 'activate to sort']):
                if txt not in ['No. Meter', 'Nama', 'ID Pelanggan']:
                    match = re.search(r'\[(\d+),(\d+)\]\[(\d+),(\d+)\]', bounds)
                    if match:
                        x1, y1, x2, y2 = map(int, match.groups())
                        nodes.append({'text': txt, 'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2, 'yc': (y1 + y2) // 2})
                        return  # Hentikan rekursi ke child jika parent sudah memuat teks (cegah duplikat sub-child)

        for child in elem:
            collect_nodes(child)

    collect_nodes(search_root)

    # 3. Grouping berdasarkan Y-Center (yc) dengan toleransi 20px
    rows = {}
    for n in nodes:
        found_row = False
        for yc in rows:
            if abs(n['yc'] - yc) <= 20:
                rows[yc].append(n)
                found_row = True
                break
        if not found_row:
            rows[n['yc']] = [n]

    # 4. Parsing data per baris
    hasil = []
    for yc in sorted(rows.keys()):
        row_nodes = sorted(rows[yc], key=lambda item: item['x1'])
        texts = [rn['text'] for rn in row_nodes]

        no_meter = ""
        nama = ""
        idpel = ""

        for txt in texts:
            clean_digits = re.sub(r'[^\d]', '', txt)
            # ID Pelanggan: Tepat 12 digit angka
            if len(clean_digits) == 12:
                idpel = clean_digits
            # No. Meter: Digit angka <= 11 digit (dibersihkan dari '+', spasi, atau karakter lain)
            elif 1 <= len(clean_digits) <= 11 and (txt.startswith('+') or clean_digits == txt.replace('+', '').strip()):
                no_meter = clean_digits
            else:
                # Teks Nama (bukan angka murni & bukan kata kunci navigasi)
                if txt and not any(k in txt for k in ['Show', 'entries', 'Search:', 'Previous', 'Next']):
                    nama = txt

        # Baris dianggap valid jika IDPEL ditemukan (tepat 12 digit)
        if idpel:
            hasil.append((no_meter, nama, idpel))

    return hasil


def cek_apakah_halaman_terakhir():
    """
    Memeriksa apakah sudah mencapai halaman terakhir tabel berdasarkan teks 'Showing X to Y of Z entries'.
    Returns True jika Y >= Z (halaman terakhir) atau Z == 0, False jika masih ada halaman berikutnya.
    """
    try:
        xml = d.dump_hierarchy()
        match = re.search(r'Showing\s+[\d\,]+\s+to\s+([\d\,]+)\s+of\s+([\d\,]+)\s+entries', xml, re.IGNORECASE)
        if match:
            y = int(match.group(1).replace(',', ''))
            z = int(match.group(2).replace(',', ''))
            if z == 0 or y >= z:
                print(f"[PAGE CHECK] Telah mencapai data terakhir ({y}/{z}). Ini adalah HALAMAN TERAKHIR.")
                return True
            else:
                return False
    except Exception as e:
        print(f"[WARNING] Gagal memeriksa teks Showing entries: {e}")
    return False





def salin_reject_ke_excel(file_txt="reject.txt", file_excel="data_reject.xlsx"):
    """
    Menyalin data dari reject.txt ke file data_reject.xlsx.
    Menambahkan data di baris paling bawah.
    Kolom A = idpel
    Kolom B = no_meter
    """
    print(f"\n[EXCEL] Menyalin data dari '{file_txt}' ke '{file_excel}'...")
    if not os.path.exists(file_txt):
        print(f"[WARNING] File '{file_txt}' tidak ditemukan. Batal menyalin ke Excel.")
        return False

    # 1. Buka atau buat workbook Excel
    if os.path.exists(file_excel):
        try:
            wb = openpyxl.load_workbook(file_excel)
            ws = wb.active
            print(f"[EXCEL] File '{file_excel}' ditemukan. Menambahkan data di baris terakhir...")
        except Exception as e:
            print(f"[WARNING] Gagal membaca file '{file_excel}': {e}. Membuat file Excel baru.")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reject"
            ws.cell(row=1, column=1, value="IDPEL")
            ws.cell(row=1, column=2, value="NO_METER")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reject"
        ws.cell(row=1, column=1, value="IDPEL")
        ws.cell(row=1, column=2, value="NO_METER")
        print(f"[EXCEL] Membuat file Excel baru '{file_excel}'...")

    # 2. Baca isi file reject.txt dan simpan ke Excel
    baris_ditambahkan = 0
    try:
        with open(file_txt, "r", encoding="utf-8") as f:
            lines = f.readlines()

        for line in lines:
            line_str = line.strip()
            if not line_str:
                continue

            parts = line_str.split("\t")
            # Abaikan baris header jika ada
            if any(h in line_str for h in ["No. Meter", "ID Pelanggan", "IDPEL", "NO_METER"]):
                continue

            no_meter = ""
            idpel = ""

            for part in parts:
                p_clean = re.sub(r'[^\d]', '', part.strip())
                if len(p_clean) == 12:
                    idpel = p_clean
                elif 1 <= len(p_clean) <= 11:
                    no_meter = p_clean

            if idpel:
                # Kolom A = idpel, Kolom B = no_meter
                ws.append([idpel, no_meter])
                baris_ditambahkan += 1

        wb.save(file_excel)
        print(f"[SUKSES] Berhasil menyalin {baris_ditambahkan} baris data dari '{file_txt}' ke '{file_excel}'!")
        print(f"        Kolom A = IDPEL | Kolom B = NO_METER")
        return True

    except Exception as e:
        print(f"[ERROR] Gagal menyalin data ke Excel: {e}")
        return False


def proses_ekstraksi_dan_swipe():
    """
    Langkah 4: Mengambil data tabel sambil swipe dinamis, simpan ke reject.txt,
    navigasi tombol Next dan scroll dinamis super cepat kembali ke page paling atas.
    """
    print("\n[LANGKAH 4] Memulai ekstraksi data tabel dan swipe dinamis...")
    file_output = "reject.txt"
    set_terproses = set()

    # Helper simpan baris data berdasarkan CUSTOM_COLUMNS
    def simpan_data_baris(data_list):
        baris_baru = 0
        for no_meter, nama, idpel in data_list:
            if idpel not in set_terproses:
                set_terproses.add(idpel)
                baris_baru += 1
                data_map = {
                    "NO_METER": no_meter,
                    "NAMA": nama,
                    "IDPEL": idpel
                }
                row_values = [str(data_map.get(field, "")) for field, _ in CUSTOM_COLUMNS]
                row_str = "\t".join(row_values) + "\n"
                with open(file_output, "a", encoding="utf-8") as f:
                    f.write(row_str)
                log_detail = " | ".join([f"{header}: {val}" for (field, header), val in zip(CUSTOM_COLUMNS, row_values)])
                print(f"  [+ SIMPAN] {log_detail}")
            else:
                print(f"  [SKIP KEMBAR] IDPEL: {idpel} (sudah tersimpan sebelumnya)")
        return baris_baru

    # Cari index kolom IDPEL pada CUSTOM_COLUMNS untuk membaca file lama
    idpel_col_idx = None
    for idx, (field, _) in enumerate(CUSTOM_COLUMNS):
        if field == "IDPEL":
            idpel_col_idx = idx
            break

    # Cek jika file reject.txt sudah ada -> muat IDPEL lama ke set_terproses agar tidak duplikat
    if os.path.exists(file_output):
        try:
            with open(file_output, "r", encoding="utf-8") as f:
                lines = f.readlines()
                for line in lines[1:]:
                    parts = line.strip().split("\t")
                    if idpel_col_idx is not None and len(parts) > idpel_col_idx:
                        existing_idpel = parts[idpel_col_idx].strip()
                        if existing_idpel:
                            set_terproses.add(existing_idpel)
                    else:
                        for part in parts:
                            p = part.strip()
                            if p.isdigit() and len(p) >= 10:
                                set_terproses.add(p)
                                break
            print(f"[FILE] File '{file_output}' ditemukan. Memuat {len(set_terproses)} IDPEL lama (mencegah duplikat lintas eksekusi).")
        except Exception as e:
            print(f"[WARNING] Gagal membaca isi file '{file_output}' lama: {e}")
    else:
        # Tulis header dinamis jika file reject.txt belum ada
        header_text = "\t".join([header for _, header in CUSTOM_COLUMNS]) + "\n"
        with open(file_output, "w", encoding="utf-8") as f:
            f.write(header_text)
        print(f"[FILE] Membuat file baru '{file_output}' dengan header: '{header_text.strip()}'")

    width = d.info.get('displayWidth', 1080)
    height = d.info.get('displayHeight', 1920)
    cx = width // 2
    cy = height // 2
    target_y = cy - 800
    target_up_y = cy + 800

    page_count = 1

    while True:
        print(f"\n==================================================")
        print(f"        MEMPROSES HALAMAN TABEL {page_count}        ")
        print(f"==================================================")

        target_page_label = str(page_count + 1)

        while True:
            # 1. Ekstrak data tabel saat ini
            data_layar = ekstrak_data_tabel()
            simpan_data_baris(data_layar)

            # 2. Cek apakah tombol halaman target ('2', '3', dst) atau 'Next' sudah terlihat di layar
            target_btn_exists = False
            for label in [target_page_label, "Next"]:
                if d(text=label).exists() or d(description=label).exists():
                    target_btn_exists = True
                    break

            if target_btn_exists:
                print(f"[PAGE {page_count}] Tombol halaman '{target_page_label}' (atau 'Next') terdeteksi di layar. Melakukan ekstraksi akhir halaman...")
                data_bawah = ekstrak_data_tabel()
                simpan_data_baris(data_bawah)

                # Swipe statis 1x setelah menemukan tombol target
                print(f"[SWIPE] Swipe statis 1x setelah tombol terdeteksi ({cx}, {cy} -> {cx}, {target_y})...")
                swipe_aman(cx, cy, cx, target_y, duration=0.1)
                time.sleep(0.3)
                break

            # 3. Swipe dinamis dari (cx, cy) ke (cx, target_up_y)
            print(f"[SWIPE] Swipe dinamis ({cx}, {cy} -> {cx}, {target_y})...")
            swipe_aman(cx, cy, cx, target_y, duration=0.1)
            time.sleep(0.3)

        # Cek 1: Periksa apakah info pagination menunjukkan halaman terakhir (Y >= Z)
        if cek_apakah_halaman_terakhir():
            print(f"[INFO] Halaman terakhir terdeteksi ('Showing entries'). Ekstraksi Selesai!")
            break

        # 4. Helper & Cek tombol target apakah clickable="true"
        def is_target_clickable():
            for label in [target_page_label, "Next"]:
                desc_el = d(description=label)
                if desc_el.exists:
                    try:
                        if desc_el.info.get("clickable", False) and desc_el.info.get("enabled", True):
                            return True
                    except Exception:
                        pass

                text_el = d(text=label)
                if text_el.exists:
                    try:
                        if text_el.info.get("clickable", False) and text_el.info.get("enabled", True):
                            return True
                    except Exception:
                        pass

                try:
                    xp_parent = d.xpath(f"//*[@text='{label}' or @content-desc='{label}']/..")
                    if xp_parent.exists:
                        if xp_parent.info.get("clickable", False) and xp_parent.info.get("enabled", True):
                            return True
                except Exception:
                    pass

            return False

        is_clickable = is_target_clickable()

        # Jika masih clickable=False, lakukan swipe y = cy - 350 lagi max 3x loop sampai clickable=True
        if not is_clickable:
            print(f"[PAGE {page_count}] Tombol '{target_page_label}' / 'Next' terdeteksi clickable=False. Memulai retry swipe down (max 3x loop)...")
            for retry in range(1, 4):
                print(f"[RETRY {retry}/3] Swipe dinamis ({cx}, {cy} -> {cx}, {target_y}) untuk memunculkan tombol...")
                swipe_aman(cx, cy, cx, target_y, duration=0.1)
                time.sleep(0.3)

                # Ekstrak data jika ada baris baru yang muncul saat retry
                data_retry = ekstrak_data_tabel()
                simpan_data_baris(data_retry)

                if is_target_clickable():
                    is_clickable = True
                    print(f"[RETRY {retry}/3] [OK] Tombol '{target_page_label}' / 'Next' sekarang CLICKABLE (clickable=True)!")
                    break

        # Re-check halaman terakhir setelah retry swipe jika ada
        if cek_apakah_halaman_terakhir():
            print(f"[INFO] Halaman terakhir terdeteksi setelah retry ('Showing entries'). Ekstraksi Selesai!")
            break

        print(f"[PAGE {page_count}] Status akhir tombol '{target_page_label}': clickable={is_clickable}")

        if not is_clickable:
            print(f"[INFO] Tombol '{target_page_label}' tetap clickable=False setelah 3x retry swipe (Halaman Terakhir). Ekstraksi Selesai!")
            break

        # 5. Ketuk tombol target halaman (misal '2', '3', dst) atau fallback 'Next'
        print(f"[PAGE {page_count}] Mengetuk tombol '{target_page_label}' (atau 'Next') untuk pindah ke Halaman {page_count + 1}...")
        clicked = False
        for label in [target_page_label, "Next"]:
            if clicked:
                break
            try:
                if d(description=label).exists():
                    d(description=label).click()
                    clicked = True
            except Exception:
                pass

            if not clicked:
                try:
                    if d(text=label).exists():
                        d(text=label).click()
                        clicked = True
                except Exception:
                    pass

            if not clicked:
                try:
                    d.xpath(f"//*[@text='{label}' or @content-desc='{label}']").click()
                    clicked = True
                except Exception:
                    pass

        # Beri jeda 1.0s agar data halaman baru selesai termuat di WebView
        time.sleep(1.0)
        page_count += 1

        # 6. Swipe statis cepat dari (cx, cy) ke (cx, cy + 800) sampai mentok ke atas
        target_up_y = cy + 800
        print(f"[SWIPE UP] Swipe statis cepat dari ({cx}, {cy}) ke ({cx}, {target_up_y}) sampai mentok ke atas...")
        for swipe_up_idx in range(1, 10):
            if d(textContains="Filter").exists() or d(textContains="Search").exists():
                print(f"[SWIPE UP] Sudah kembali ke header paling atas halaman {page_count}.")
                break
            swipe_aman(cx, cy, cx, target_up_y, duration=0.05)
            time.sleep(0.05)

        time.sleep(0.4)

    # 7. Salin data dari reject.txt ke data_reject.xlsx di baris paling bawah (Kolom A = idpel, Kolom B = no_meter)
    salin_reject_ke_excel()


def reset_posisi_entri_tabel():
    """
    Mereset posisi entri tabel ke Halaman 1 (Showing 1 to Y of Z entries) 
    dan scroll dinamis cepat dari (cx, cy) ke (cx, cy + 800) sampai kembali ke paling atas 'Filter By Reject'.
    """
    try:
        width = d.info.get('displayWidth', 1080)
        height = d.info.get('displayHeight', 1920)
        cx = width // 2
        cy = height // 2
        target_y = cy - 800
        target_up_y = cy + 800

        # Pre-check: jika posisi saat ini berada di paling atas (Filter By Reject / Search), swipe down sampai ketemu 'Previous'
        is_top_page = (
            d(textContains="Filter By Reject").exists() or
            d(textContains="Filter By").exists() or
            d(textContains="Search").exists()
        )

        if is_top_page:
            print("[RESET PRE-CHECK] Posisi terdeteksi di paling atas ('Filter By Reject'). Swiping down sampai ketemu tombol 'Previous'...")
            for idx in range(1, 15):
                has_prev = d(text="Previous").exists() or d(description="Previous").exists()
                if has_prev:
                    print(f"[RESET PRE-CHECK] Tombol 'Previous' ditemukan di layar (percobaan ke-{idx}).")
                    break
                swipe_aman(cx, cy, cx, target_y, duration=0.1)
                time.sleep(0.3)

        print("\n[RESET] Memeriksa & mereset posisi entri tabel ke Halaman 1...")

        # Loop menekan tombol Previous hingga berada di Halaman 1 (Showing 1 to Y of Z entries)
        for attempt in range(1, 20):
            xml = d.dump_hierarchy()
            match_page1 = re.search(r'Showing\s+1\s+to\s+[\d\,]+\s+of\s+[\d\,]+\s+entries', xml, re.IGNORECASE)
            if match_page1:
                print(f"[RESET] Berhasil berada di Halaman 1 ({match_page1.group(0)}).")
                break

            print(f"[RESET] Belum di Halaman 1 (Percobaan {attempt}). Mencari & mengetuk tombol 'Previous'...")

            prev_btn = d(text="Previous")
            if not prev_btn.exists():
                prev_btn = d(description="Previous")

            if not prev_btn.exists():
                print(f"[RESET] Tombol 'Previous' belum terlihat. Swipe down ({cx}, {cy} -> {cx}, {target_y})...")
                swipe_aman(cx, cy, cx, target_y, duration=0.1)
                time.sleep(0.3)
                prev_btn = d(text="Previous")
                if not prev_btn.exists():
                    prev_btn = d(description="Previous")

            if prev_btn.exists():
                is_clickable = False
                try:
                    is_clickable = prev_btn.info.get("clickable", False)
                except Exception:
                    pass

                if is_clickable or d(text="Previous").exists():
                    print("[RESET] Mengetuk tombol 'Previous'...")
                    try:
                        prev_btn.click()
                    except Exception:
                        try:
                            d.xpath("//*[@text='Previous' or @content-desc='Previous']").click()
                        except Exception:
                            pass
                    time.sleep(0.5)
                else:
                    print("[RESET] Tombol 'Previous' sudah clickable=False. Berada di Halaman 1.")
                    break
            else:
                print("[WARNING] Tombol 'Previous' tidak ditemukan. Berhenti loop Previous.")
                break

        # Swipe dinamis cepat dari (cx, cy) ke (cx, cy + 800) sampai kembali ke paling atas
        print(f"[SWIPE UP] Swipe dinamis cepat dari ({cx}, {cy}) ke ({cx}, {target_up_y}) sampai kembali ke paling atas...")
        for swipe_up_idx in range(1, 20):
            if d(textContains="Filter").exists() or d(textContains="Search").exists():
                print(f"[SCROLL] Teks header terdeteksi di paling atas (percobaan ke-{swipe_up_idx}).")
                break
            swipe_aman(cx, cy, cx, target_up_y, duration=0.05)
            time.sleep(0.05)

        time.sleep(0.3)
        return True
    except Exception as e:
        print(f"[ERROR] Gagal saat mereset posisi entri tabel: {e}")
        return False


def main():
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

    print("==================================================")
    print("       BOT EMULATOR REJECT - FASIH SCRAPPER       ")
    print("==================================================")

    # Hubungkan ke emulator
    if not hubungkan_emulator():
        return

    # Cek apakah dialog filter sudah terbuka di layar saat ini
    is_filter_open = (
        d(text="Filter Assignment By Status").exists() or
        d(resourceId="id.go.bpsfasih:id/lButton_bottomDialogFilterAssignment").exists() or
        d(resourceId="id.go.bpsfasih:id/reject_cb_bottomSheetFilterAssignment").exists()
    )

    if is_filter_open:
        print("[INFO] Dialog 'Filter Assignment By Status' sudah terbuka di layar. Langsung menuju Langkah 3...")
    else:
        # 1. Cek jika kebablasan ke halaman 'Periode' atau 'Daftar Wilayah'
        if not cek_dan_tangani_kebablasan():
            print("[HALT] Proses terhenti di Langkah 1.")
            return

        # 2. Ketuk FAB & Ketuk Filter By Status
        if not ketuk_fab_dan_filter():
            print("[HALT] Proses terhenti di Langkah 2.")
            return

    # 3. Ketuk tombol 'BERSIHKAN' -> checklist Reject -> ketuk tombol 'TERAPKAN'
    if not terapkan_filter_reject():
        print("[HALT] Proses terhenti di Langkah 3.")
        return

    # 4. Reset posisi entri tabel ke Halaman 1 & scroll ke paling atas
    reset_posisi_entri_tabel()

    # 5. Mengambil data tabel sambil swipe dinamis, simpan ke reject.txt, dan navigasi Next
    proses_ekstraksi_dan_swipe()

    print("\n==================================================")
    print("      PROSES BOT REJECT SELESAI DENGAN SUKSES     ")
    print("==================================================")


if __name__ == "__main__":
    main()
