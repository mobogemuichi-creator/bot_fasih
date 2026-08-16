import pdfplumber
import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import urllib.request
import urllib.error
import socket
import sys
import time
from openpyxl import load_workbook          # <--- TAMBAHKAN INI
from openpyxl.styles import Alignment       # <--- TAMBAHKAN INI

# ==========================================
# ⚙️ PENGATURAN FITUR (TOGGLE)
# True  = Gabung semua PDF terpilih jadi 1 file Excel (Muncul pop-up Save As)
# False = Convert masing-masing PDF jadi Excel terpisah dengan nama yang sama
# ==========================================
MODE_GABUNG = True 

def cek_status_online():
    print("Memeriksa status lisensi online...")
    
    # GANTI LINK DI BAWAH INI DENGAN LINK GIST MILIKMU!
    base_url = "https://gist.githubusercontent.com/mobogemuichi-creator/99ccd20c30911356dfce972a6d3c022f/raw/status_pdf_to_excel.txt"
    url_saklar = f"{base_url}?t={int(time.time())}"
    
    try:
        import ssl
        # Buat context SSL unverified untuk menghindari issue sertifikat SSL di Windows
        context = ssl._create_unverified_context()
        
        # Gunakan header User-Agent agar tidak diblokir oleh GitHub
        req = urllib.request.Request(url_saklar, headers={'User-Agent': 'Mozilla/5.0'})
        
        with urllib.request.urlopen(req, context=context, timeout=5) as response:
            status = response.read().decode('utf-8').strip()
        
        if status == "ACTIVE":
            print("✅ Lisensi Valid. Program diizinkan berjalan.\n")
            return True
        else:
            warning_msg = f"⛔ AKSES DITOLAK: Aplikasi ini telah diblokir. (Status: {status})"
            print(warning_msg)
            
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            messagebox.showerror("Akses Diblokir", warning_msg)
            return False
            
    except Exception as e:
        # Tentukan pesan error secara rinci berdasarkan tipe exception
        if isinstance(e, urllib.error.HTTPError):
            error_msg = f"🌐 ERROR SERVER (HTTP {e.code}): Gagal mengakses database lisensi.\nDetail: {e.reason}"
        elif isinstance(e, urllib.error.URLError):
            if isinstance(e.reason, socket.timeout) or "timed out" in str(e.reason).lower():
                error_msg = "⏳ TIMEOUT KONEKSI: Waktu tunggu habis. Koneksi internet Anda kemungkinan terlalu lambat atau tidak stabil."
            elif "getaddrinfo failed" in str(e.reason) or "[Errno 11001]" in str(e.reason):
                error_msg = "⚠️ TIDAK ADA INTERNET: Komputer tidak terhubung ke internet, atau DNS gagal menemukan server GitHub."
            else:
                error_msg = f"🔌 GAGAL KONEKSI: Terjadi masalah jaringan saat menghubungkan ke database.\nDetail: {e.reason}"
        else:
            error_msg = f"❌ ERROR SISTEM: Terjadi kesalahan tak terduga saat memverifikasi lisensi.\nDetail: {e}"

        print(error_msg)
        print(f"[DEBUG INFO] Detail Error: {e}")
        
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        messagebox.showwarning("Koneksi Gagal", error_msg)
        return False

# FUNGSI INTI UNTUK EKSTRAKSI (Bisa 1 atau Banyak PDF)
def process_pdfs(pdf_paths, excel_path):
    all_data = []
    header_saved = False

    for pdf_path in pdf_paths:
        print(f"\nMembuka file PDF:\n-> {os.path.basename(pdf_path)}")
        
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages):
                print(f"Memproses halaman {i + 1}...")
                
                # --- Filter Watermark ---
                def filter_watermark(obj):
                    if obj.get("object_type") == "char":
                        if obj.get("size") > 14.0:
                            return False
                    return True

                clean_page = page.filter(filter_watermark)
                # ------------------------
                
                table = clean_page.extract_table(table_settings={
                    "vertical_strategy": "lines",   
                    "horizontal_strategy": "lines", 
                    "snap_tolerance": 3,            
                    "join_tolerance": 3
                })

                if not table or len(table) == 0:
                    print(f"-> Melewati halaman {i + 1} (Tidak ada tabel terdeteksi)")
                    continue

                # Logika Penyaringan Header (Lintas Dokumen)
                if not header_saved:
                    all_data.extend(table)
                    header_saved = True
                    print(f"-> Header tabel dikunci dari {os.path.basename(pdf_path)} (Halaman {i + 1}).")
                else:
                    all_data.extend(table[1:])

    if not all_data:
        print("❌ Gagal: Tidak ada tabel dengan border yang terdeteksi.")
        return False

    print("\nMenyaring dan mengonversi data ke Excel...")
    
    headers = all_data[0]
    raw_data = all_data[1:]

    # --- Filter Baris 'TOTAL:' ---
    data = []
    baris_dihapus = 0
    
    for row in raw_data:
        abaikan_baris = False
        for cell in row:
            if cell and isinstance(cell, str) and "TOTAL:" in cell:
                abaikan_baris = True
                baris_dihapus += 1
                break
        
        if not abaikan_baris:
            data.append(row)
            
    if baris_dihapus > 0:
        print(f"-> Info: Ditemukan dan dihapus {baris_dihapus} baris rekapitulasi (TOTAL: ...)")
    # -----------------------------

    df = pd.DataFrame(data, columns=headers)
    df.dropna(how='all', inplace=True)

    # --- 1. FITUR PEMBERSIHAN ENTER / NEWLINE ---
    # Mengubah karakter enter tersembunyi dari PDF menjadi spasi biasa
    df = df.replace('\n', ' ', regex=True)
    
    # Membersihkan nama kolom (header) dari enter
    df.columns = [str(col).replace('\n', ' ') for col in df.columns]
    # ---------------------------------------------

    try:
        # Save data ke Excel
        df.to_excel(excel_path, index=False)
        
        # --- 2. FITUR MENGHILANGKAN WRAP TEXT EXCEL ---
        print("Merapikan format tabel (Menghilangkan Wrap Text)...")
        
        # Buka kembali file Excel yang baru saja disave untuk diformat
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Buat gaya format sel: matikan wrap text, posisikan teks di tengah secara vertikal
        format_rapi = Alignment(wrap_text=False, vertical='center')
        
        # Terapkan format ke seluruh sel yang ada isinya
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = format_rapi
                
        # Simpan kembali perubahannya
        wb.save(excel_path)
        # ----------------------------------------------

        print(f"✅ Selesai! File Excel bersih dan rapi berhasil disimpan di:\n-> {excel_path}")
        return True
    except PermissionError:
        warning_msg = (
            "⚠️ AKSES DITOLAK: File Excel tujuan sedang terbuka!\n\n"
            f"File '{os.path.basename(excel_path)}' sedang dibuka oleh program lain.\n"
            "Tutup file tersebut, lalu jalankan ulang."
        )
        print(warning_msg)
        messagebox.showerror("Error: File Sedang Terbuka", warning_msg)
        return False

# --- FUNGSI PILIH FILE DAN PENGATUR ALUR ---
def select_and_process():
    root = tk.Tk()
    root.withdraw() 
    root.attributes('-topmost', True) 
    
    # Dialog buka banyak file
    input_pdfs = filedialog.askopenfilenames(
        title="Pilih File PDF (Bisa blok/pilih lebih dari 1 file)",
        filetypes=[("PDF Files", "*.pdf")]
    )
    
    if not input_pdfs:
        print("Proses dibatalkan oleh pengguna.")
        return

    # LOGIKA TOGGLE
    if MODE_GABUNG:
        print(f"\n[MODE GABUNG] {len(input_pdfs)} file PDF akan digabung jadi 1 Excel.")
        
        # Ambil jalur direktori dari file PDF pertama yang baru saja dipilih
        folder_terakhir = os.path.dirname(input_pdfs[0])
        
        # Munculkan popup Save As di folder yang sama dengan PDF
        output_excel = filedialog.asksaveasfilename(
            title="Simpan File Excel Gabungan",
            initialdir=folder_terakhir,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile="Data_Gabungan_dari_pdf.xlsx"
        )
        
        if not output_excel:
            print("Proses dibatalkan (Penyimpanan file dibatalkan).")
            return
            
        success = process_pdfs(input_pdfs, output_excel)
        if success:
            messagebox.showinfo("Berhasil", f"Semua {len(input_pdfs)} file berhasil digabung dan disimpan.")
            
    else:
        print(f"\n[MODE PISAH] Mengonversi {len(input_pdfs)} file PDF secara terpisah.")
        sukses_count = 0
        
        # Loop tiap file dan proses satu per satu dengan nama masing-masing
        for pdf in input_pdfs:
            base_name = os.path.splitext(pdf)[0]
            output_excel = f"{base_name}.xlsx"
            
            success = process_pdfs([pdf], output_excel) # Kirim sebagai list isi 1
            if success:
                sukses_count += 1
                
        if sukses_count > 0:
            messagebox.showinfo("Berhasil", f"{sukses_count} dari {len(input_pdfs)} file berhasil diconvert secara terpisah.")

if __name__ == "__main__":
    if os.name == 'nt':
        os.system("title PDF to XLSX by baliaga")
    
    print("=============================================")
    print("           PDF to XLSX Converter             ")
    print("           Developed by: baliaga             ")
    print("=============================================\n")

    if not cek_status_online():
        input("\nTekan Enter untuk keluar...")
        sys.exit()

    try:
        select_and_process()
    except Exception as e:
        print(f"\n❌ Terjadi kesalahan sistem:\n{e}")
    
    input("\nTekan Enter untuk keluar...")