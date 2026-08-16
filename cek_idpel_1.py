"""
Script Otomatisasi Cek Status IDPEL di FASIH (cek_idpel_1.py)
-------------------------------------------------------------
Membaca IDPEL dari kolom A pada file Excel 'data_cek.xlsx',
memeriksa status keberadaan IDPEL di aplikasi FASIH,
dan menyimpan hasilnya ke file 'cek_nik.txt' dengan format 'idpel|status_exist'.
Jika IDPEL sudah pernah tercatat di 'cek_nik.txt', script akan otomatis di-skip.
"""
import os
import time
import subprocess
import openpyxl
import uiautomator2 as u2
import sys
import datetime

from konfigurasi import (
    LDPLAYER_ADB,
    EMULATOR_PORTS_1 as EMULATOR_PORTS,
    SLEEP_SHORT,
    SLEEP_MEDIUM,
    SLEEP_LONG,
)

d = None

def hubungkan_emulator():
    """Menghubungkan ke emulator via uiautomator2"""
    global d
    print("[KONEKSI] Mencoba menghubungkan ke emulator LDPlayer...")
    
    for port in EMULATOR_PORTS:
        try:
            if os.path.exists(LDPLAYER_ADB):
                subprocess.run([LDPLAYER_ADB, "connect", f"127.0.0.1:{port}"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=2)
            temp_d = u2.connect(f"127.0.0.1:{port}")
            info = temp_d.info
            d = temp_d
            print(f"[KONEKSI] Berhasil terhubung ke emulator di port {port}!")
            print(f"[EMULATOR] Layar: {info.get('displayWidth')}x{info.get('displayHeight')} ({info.get('productName', 'Android')})")
            return True
        except Exception:
            continue

    try:
        temp_d = u2.connect()
        info = temp_d.info
        d = temp_d
        print("[KONEKSI] Berhasil terhubung ke emulator via default connection!")
        return True
    except Exception:
        pass
            
    print("[ERROR] Gagal terhubung ke emulator. Pastikan LDPlayer sudah berjalan.")
    return False


def baca_idpel_terproses(txt_path="cek_nik.txt"):
    """
    Membaca file txt 'cek_nik.txt' jika sudah ada,
    dan mengembalikan set dari IDPEL yang sudah diproses.
    Format file txt per baris: idpel|status_exist
    """
    terproses = set()
    if os.path.exists(txt_path):
        try:
            with open(txt_path, "r", encoding="utf-8") as f:
                for line in f:
                    line_str = line.strip()
                    if line_str and "|" in line_str:
                        parts = line_str.split("|")
                        idpel_key = parts[0].strip()
                        if idpel_key:
                            terproses.add(idpel_key)
            print(f"[TXT] Ditemukan {len(terproses)} IDPEL yang sudah tercatat di '{txt_path}'.")
        except Exception as e:
            print(f"[WARNING] Gagal membaca '{txt_path}': {e}")
    else:
        print(f"[TXT] File '{txt_path}' belum ada, akan dibuat secara otomatis saat menyimpan hasil.")
    return terproses


def simpan_ke_txt(idpel, status_exist, txt_path="cek_nik.txt"):
    """Menyimpan status IDPEL ke file 'cek_nik.txt' dengan format 'idpel|status_exist'"""
    try:
        with open(txt_path, "a", encoding="utf-8") as f:
            f.write(f"{idpel}|{status_exist}\n")
        print(f"[TXT] Berhasil menyimpan '{idpel}|{status_exist}' ke '{txt_path}'.")
        return True
    except Exception as e:
        print(f"[ERROR TXT] Gagal menyimpan ke '{txt_path}': {e}")
        return False


def proses_cek_idpel(excel_path="data_cek.xlsx", txt_path="cek_nik.txt"):
    """Fungsi utama membaca data_cek.xlsx kolom A dan melakukan pengecekan IDPEL di Fasih"""
    if not os.path.exists(excel_path):
        print(f"[ERROR] File Excel '{excel_path}' tidak ditemukan!")
        return

    # 1. Baca IDPEL yang sudah diproses sebelumnya dari cek_nik.txt
    sudah_dibaca = baca_idpel_terproses(txt_path)

    # 2. Baca data dari data_cek.xlsx (Kolom A)
    print(f"[EXCEL] Membaca data IDPEL dari kolom A pada '{excel_path}'...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
    except Exception as e:
        print(f"[ERROR] Gagal membaca file Excel '{excel_path}': {e}")
        return

    max_row = sheet.max_row
    print(f"[EXCEL] Total baris di Sheet: {max_row}")

    for row in range(2, max_row + 1):
        idpel_raw = sheet.cell(row=row, column=1).value  # Kolom A
        if idpel_raw is None:
            continue

        # Clean string idpel
        if isinstance(idpel_raw, float):
            idpel = str(int(idpel_raw)).strip()
        else:
            idpel = str(idpel_raw).strip()

        if not idpel or idpel.lower() in ["none", "nan"]:
            continue

        # 3. Cek apakah IDPEL sudah dibaca di cek_nik.txt
        if idpel in sudah_dibaca:
            print(f"[SKIP] Baris {row} | IDPEL {idpel} sudah tercatat di '{txt_path}'. Dilewati.")
            continue

        print(f"\n--------------------------------------------------")
        print(f"Memproses Baris {row} | IDPEL: {idpel}")
        print(f"--------------------------------------------------")

        try:
            # ==========================================
            # NAVIGASI DASHBOARD (Referensi L740-L850)
            # ==========================================
            is_on_form = d(textContains="101a. ID pelanggan").exists() or d(text="Cek ID Pelanggan").exists()

            if not is_on_form:
                # Transisi ke Daftar Assignment (Self-Healing)
                region_row = d(resourceId="id.go.bpsfasih:id/updateListingLayout")
                if region_row.exists():
                    print("[DASHBOARD] Berada di halaman SLS Wilayah. Mengklik wilayah untuk masuk...")
                    region_row.click()
                    time.sleep(SLEEP_LONG)

                # Klik FAB & Tambah Assignment
                ya_btn_found = False
                for attempt in range(1, 4):
                    print(f"[DASHBOARD] Mengklik FAB (Percobaan {attempt}/3)...")
                    fab = d(resourceId="id.go.bpsfasih:id/expendable_fab")
                    if fab.wait(exists=True, timeout=5):
                        fab.click()
                        time.sleep(SLEEP_SHORT)
                    else:
                        print(f"[WARNING] Tombol FAB tidak ditemukan pada percobaan {attempt}.")
                        time.sleep(1.0)
                        continue

                    print(f"[DASHBOARD] Mengklik 'Tambah Assignment' (Percobaan {attempt}/3)...")
                    tambah_btn = d(resourceId="id.go.bpsfasih:id/fab_addAssignment")
                    if not tambah_btn.exists():
                        tambah_btn = d(text="Tambah Assignment")
                    if tambah_btn.exists(timeout=5):
                        tambah_btn.click()
                        time.sleep(SLEEP_MEDIUM)
                    else:
                        print(f"[WARNING] Tombol 'Tambah Assignment' tidak ditemukan pada percobaan {attempt}.")
                        time.sleep(1.0)
                        continue

                    print("[DASHBOARD] Menyetujui konfirmasi dialog 'YA'...")
                    ya_btn = d(resourceId="id.go.bpsfasih:id/rButton_bottomDialog")
                    if not ya_btn.exists(): ya_btn = d(text="YA")
                    if not ya_btn.exists(): ya_btn = d(text="Ya")
                    if ya_btn.exists(timeout=5):
                        ya_btn.click()
                        print("Menunggu form termuat (3 detik)...")
                        time.sleep(3.0)
                        ya_btn_found = True
                        break
                    else:
                        print(f"[WARNING] Tombol konfirmasi 'YA' tidak muncul (Percobaan {attempt}/3). Mengulangi...")
                        time.sleep(1.0)

                if not ya_btn_found:
                    raise Exception("Tombol konfirmasi 'YA' tidak ditemukan setelah 3 kali percobaan (FAB -> Tambah Assignment).")

                # Scan & Ambil Waktu
                found_scan_waktu = False
                for scan_attempt in range(1, 11):
                    print(f"[BLOK I] Scanning teks 'Ambil Waktu' (Percobaan {scan_attempt}/10)...")
                    if d(text="Ambil Waktu").wait(exists=True, timeout=1.0) or d(textContains="Ambil Waktu").exists():
                        found_scan_waktu = True
                        print(f"[BLOK I] Teks 'Ambil Waktu' ditemukan pada percobaan {scan_attempt}.")
                        break

                if not found_scan_waktu:
                    print(f"[SKIP] Baris {row} | IDPEL {idpel} dilewati karena teks 'Ambil Waktu' tidak ditemukan.")
                    continue

                print("[BLOK I] Mengetuk tombol 'Ambil Waktu'...")
                waktu_btn = d(text="Ambil Waktu")
                if waktu_btn.exists():
                    waktu_sukses = False
                    for attempt_waktu in range(1, 4):
                        print(f"[BLOK I] Mengetuk tombol 'Ambil Waktu' (Percobaan {attempt_waktu}/3)...")
                        waktu_btn.click()
                        time.sleep(SLEEP_SHORT)
                        
                        dialog_waktu = d(textContains="Apakah Anda yakin ingin mengambil waktu saat ini")
                        if not dialog_waktu.exists():
                            dialog_waktu = d(textContains="mengambil waktu saat ini")
                        
                        if dialog_waktu.wait(exists=True, timeout=2.0) or d(text="ya").exists() or d(text="Ya").exists() or d(text="YA").exists():
                            print("[BLOK I] Teks konfirmasi waktu terdeteksi. Mengetuk tombol 'ya'...")
                            ya_waktu = d(text="ya")
                            if not ya_waktu.exists(): ya_waktu = d(text="Ya")
                            if not ya_waktu.exists(): ya_waktu = d(text="YA")
                            if ya_waktu.exists():
                                ya_waktu.click()
                                time.sleep(SLEEP_MEDIUM)
                            waktu_sukses = True
                            break
                        else:
                            print(f"[WARNING] Dialog konfirmasi waktu tidak muncul. Mengulangi...")
                            time.sleep(1.0)
                    if not waktu_sukses:
                        print("[WARNING] Dialog konfirmasi waktu tidak muncul setelah 3 kali percobaan.")
                else:
                    print("[INFO] Waktu wawancara sudah terisi. Melewati langkah Ambil Waktu.")

                # Scroll ke tombol Cek ID Pelanggan
                print("[BLOK I] Men-scroll secara dinamis ke tombol Cek ID Pelanggan...")
                try:
                    d(scrollable=True).scroll.to(text="Cek ID Pelanggan")
                except Exception:
                    pass
                time.sleep(SLEEP_SHORT)
            else:
                print("[INFO] Sudah berada di form pengisian ID Pelanggan. Melewati langkah awal.")

            # Input ID Pelanggan
            print("[BLOK I] Mencari kolom input ID Pelanggan...")
            id_label = d(textContains="101a. ID pelanggan")
            if not id_label.exists():
                try:
                    d(scrollable=True).scroll.to(textContains="101a. ID pelanggan")
                except Exception:
                    pass

            if id_label.wait(exists=True, timeout=5):
                id_input = id_label.down(className="android.widget.EditText")
                if id_input.exists():
                    print(f"[BLOK I] Menulis ID Pelanggan via set_text: {idpel}...")
                    id_input.set_text(str(idpel))
                    time.sleep(SLEEP_SHORT)
                else:
                    raise Exception("Kolom input EditText ID Pelanggan tidak ditemukan.")
            else:
                raise Exception("Label '101a. ID pelanggan PLN' tidak ditemukan.")

            # Kirim ENTER
            d.shell("input keyevent KEYCODE_ENTER")
            time.sleep(SLEEP_SHORT)

            # Ketuk tombol "Cek ID Pelanggan"
            print("[BLOK I] Mengetuk tombol 'Cek ID Pelanggan'...")
            cek_id_btn = d(text="Cek ID Pelanggan")
            if not cek_id_btn.exists():
                try:
                    d(scrollable=True).scroll.to(text="Cek ID Pelanggan")
                except Exception:
                    pass

            if cek_id_btn.exists(timeout=5):
                t_start = time.time()
                cek_id_btn.click()
                print(f"[{datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Mengetuk tombol 'Cek ID Pelanggan'...")
                
                # Menunggu progress bar loading selesai
                print(f"[{datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Menunggu progress bar loading selesai...")
                time.sleep(0.1)
                try:
                    d(resourceId="id.go.bpsfasih:id/card_progress").wait_gone(timeout=30)
                except Exception:
                    pass
                t_loading = time.time()
                print(f"[{datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Progress bar loading selesai (+{t_loading - t_start:.2f}s dari awal)")
                
                # Cek apakah limit API check-idpln tercapai
                limit_msg = "Permintaan API check-idpln sudah terlampaui (limit)."
                if d(textContains=limit_msg).exists():
                    print(f"[{datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]}] PROSES BERHENTI KARENA CEK ID MELEWATI LIMIT API!")
                    input("Tekan ENTER untuk keluar...")
                    sys.exit(1)
                
                # Cek status hasil
                target_status = "DITEMUKAN DAN BELUM TERCATAT PADA SISTEM FASIH"
                already_recorded = "DITEMUKAN DAN SUDAH TERCATAT PADA SISTEM FASIH"
                
                time.sleep(0.1)
                status_found = d(textContains=target_status).exists()
                
                is_already_registered = (
                    d(textContains=already_recorded).exists() or
                    d(textContains="Id Pelanggan sudah terdaftar di FASIH").exists() or
                    d(textContains="id pelanggan sudah terdaftar di FASIH").exists() or
                    d(textContains="sudah terdaftar di FASIH").exists()
                )

                is_not_found = (
                    d(textContains="id pelanggan tidak ditemukan").exists() or
                    d(textContains="ID pelanggan tidak ditemukan").exists() or
                    d(textContains="Id pelanggan tidak ditemukan").exists() or
                    d(textContains="tidak ditemukan").exists() or
                    d(text="STATUS").exists()
                )

                status_exist = "TIDAK DITEMUKAN / UNKNOWN"
                if status_found:
                    status_exist = "DITEMUKAN DAN BELUM TERCATAT PADA SISTEM FASIH"
                    print(f"[STATUS] IDPEL {idpel} -> {status_exist}")
                elif is_already_registered:
                    status_exist = "SUDAH TERCATAT PADA SISTEM FASIH"
                    print(f"[STATUS] IDPEL {idpel} -> {status_exist}")
                elif is_not_found:
                    status_exist = "id pelanggan tidak ditemukan"
                    print(f"[STATUS] IDPEL {idpel} -> {status_exist}")
                else:
                    print(f"[WARNING] Status respon khusus tidak terdeteksi secara eksplisit untuk IDPEL {idpel}.")
                    status_exist = "BELUM TERDAFTAR FASIH"

                # Simpan ke cek_nik.txt & tambahkan ke memori
                simpan_ke_txt(idpel, status_exist, txt_path)
                sudah_dibaca.add(idpel)

            else:
                raise Exception("Tombol 'Cek ID Pelanggan' tidak ditemukan di layar.")

        except Exception as err:
            print(f"[ERROR] Gagal memproses baris {row} (IDPEL {idpel}): {err}")
            err_msg = f"Error : {err}"
            simpan_ke_txt(idpel, err_msg, txt_path)
            sudah_dibaca.add(idpel)

def main():
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

    print("==================================================")
    print("      BOT CEK STATUS IDPEL FASIH (cek_idpel_1.py) ")
    print("==================================================")

    if not hubungkan_emulator():
        return

    proses_cek_idpel()

    print("\n==================================================")
    print("     PROSES CEK STATUS IDPEL SELESAI DENGAN SUKSES ")
    print("==================================================")

if __name__ == "__main__":
    main()
