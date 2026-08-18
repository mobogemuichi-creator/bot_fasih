import os
import time
import subprocess
import openpyxl
import uiautomator2 as u2
import re
import sys
import datetime


from konfigurasi import (
    LDPLAYER_DNCONSOLE,
    LDPLAYER_ADB,
    EMULATOR_INDEX_2 as EMULATOR_INDEX,
    EMULATOR_PORTS_2 as EMULATOR_PORTS,
    EXCEL_FILE_2 as EXCEL_FILE,
    FOTO_DIRECTORY,
    SLEEP_SHORT,
    SLEEP_MEDIUM,
    SLEEP_LONG,
)

# Variabel global untuk koneksi emulator
d = None

def set_lokasi_gps(lat, lng):
    """Mengubah koordinat GPS LDPlayer menggunakan dnconsole.exe"""
    if not os.path.exists(LDPLAYER_DNCONSOLE):
        print(f"[WARNING] dnconsole.exe tidak ditemukan di '{LDPLAYER_DNCONSOLE}'")
        return False
    try:
        subprocess.run([
            LDPLAYER_DNCONSOLE, "locate", 
            "--index", EMULATOR_INDEX, 
            "--LLI", f"{lng},{lat}"
        ], check=True)
        print(f"[GPS] Lokasi emulator diubah ke: {lat}, {lng}")
        time.sleep(SLEEP_MEDIUM)
        return True
    except Exception as e:
        print(f"[ERROR] Gagal mengubah lokasi GPS: {e}")
        return False

def push_dan_scan_foto(local_photo_path, dest_in_android="/sdcard/Pictures/temp_upload.jpg"):
    """Mengirim foto dari PC ke emulator dan memicu scan galeri"""
    if not local_photo_path or not os.path.exists(local_photo_path):
        print(f"[UPLOAD] Foto tidak ditemukan di lokal: '{local_photo_path}'")
        return False
    try:
        print(f"[UPLOAD] Mengirim {local_photo_path} ke emulator...")
        d.push(local_photo_path, dest_in_android)
        d.shell(f"am broadcast -a android.intent.action.MEDIA_SCANNER_SCAN_FILE -d file://{dest_in_android}")
        time.sleep(SLEEP_MEDIUM)
        print("[UPLOAD] Transfer foto sukses dan terdaftar di galeri.")
        return True
    except Exception as e:
        print(f"[ERROR] Gagal mengirim/mendaftarkan foto: {e}")
        return False

def validasi_nik(nik):
    """
    Validasi NIK berdasarkan aturan Kemendagri terbaru:
    1. Harus tepat 16 digit angka.
    2. Kode wilayah (2 digit pertama) terdaftar dalam 38 provinsi resmi Indonesia:
       11-19, 21, 31-36, 51-53, 61-65, 71-76, 81-82, 91-96.
    3. Tanggal lahir (digit 7-12) format DDMMYY:
       - Perempuan: digit 7-8 dikurangi 40 jika > 40.
       - Validasi tanggal nyata di kalender (termasuk bulan 1-12 dan tahun kabisat).
    """
    if nik is None:
        return False
    
    nik_str = str(nik).strip()

    # 1. Panjang Karakter: Harus tepat 16 digit dan semuanya angka
    if len(nik_str) != 16 or not nik_str.isdigit():
        return False

    # 2. Kode Wilayah (2 Digit Pertama): 38 Provinsi di Indonesia
    try:
        prov_code = int(nik_str[:2])
    except ValueError:
        return False

    valid_provs = {
        11, 12, 13, 14, 15, 16, 17, 18, 19,
        21,
        31, 32, 33, 34, 35, 36,
        51, 52, 53,
        61, 62, 63, 64, 65,
        71, 72, 73, 74, 75, 76,
        81, 82,
        91, 92, 93, 94, 95, 96
    }
    if prov_code not in valid_provs:
        return False

    # 3. Tanggal Lahir (Digit 7 s.d. 12: DDMMYY)
    try:
        day_val = int(nik_str[6:8])
        month = int(nik_str[8:10])
        year_suffix = int(nik_str[10:12])
    except ValueError:
        return False

    # Deteksi Gender (Perempuan: Tanggal + 40)
    if day_val > 40:
        day = day_val - 40
    else:
        day = day_val

    # Cek Bulan (1 s.d. 12)
    if not (1 <= month <= 12):
        return False

    # Cek keabsahan Tanggal di Kalender (19YY atau 20YY)
    date_valid = False
    for century in [1900, 2000]:
        try:
            datetime.date(century + year_suffix, month, day)
            date_valid = True
            break
        except ValueError:
            pass

    if not date_valid:
        return False

    return True


def swipe_ke_bawah_cepat(times=3):
    """Men-scroll layar ke bawah dengan cepat menggunakan swipe"""
    print(f"[SCROLL] Men-scroll ke bawah cepat ({times}x)...")
    for _ in range(times):
        try:
            d.swipe(540, 1300, 540, 200, duration=0.05)
            time.sleep(0.05)
        except Exception as e:
            print(f"[WARNING] Gagal swipe ke bawah: {e}")

def swipe_ke_atas_cepat(times=3):
    """Men-scroll layar ke atas dengan cepat menggunakan swipe"""
    print(f"[SCROLL] Men-scroll ke atas cepat ({times}x)...")
    for _ in range(times):
        try:
            d.swipe(540, 200, 540, 1300, duration=0.05)
            time.sleep(0.05)
        except Exception as e:
            print(f"[WARNING] Gagal swipe ke atas: {e}")


def press_back_and_check_periode():
    """Menekan BACK dan langsung memeriksa apakah kebablasan ke halaman Periode untuk memicu Submit"""
    try:
        # Ketuk tombol OK statis (bounds 528, 1691) jika ada modal OK yang menghalangi sebelum BACK
        try:
            btn_close = d(resourceId="id.go.bpsfasih:id/btn_submit_progress_close")
            if btn_close.exists:
                btn_close.click()
            elif d(text="OK").exists:
                d(text="OK").click()
            else:
                d.click(528, 1691)
        except Exception:
            d.click(528, 1691)
        time.sleep(0.5)

        d.press("back")
        time.sleep(SLEEP_SHORT)
        
        # Cek jika kebablasan ke halaman 'Periode' atau 'Daftar Wilayah'
        kebablasan = False
        while True:
            is_periode = d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Periode").exists()
            is_wilayah = d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Daftar Wilayah").exists()
            
            if is_periode or is_wilayah:
                kebablasan = True
                current_page = "Periode" if is_periode else "Daftar Wilayah"
                print(f"[BACK] Kebablasan ke halaman '{current_page}', mengetuk 'Submit'...")
                submit_btn = d(text="Submit")
                if submit_btn.exists(timeout=2):
                    submit_btn.click()
                    time.sleep(SLEEP_LONG)
                else:
                    break
            else:
                break
                
        if kebablasan:
            # Tunggu sampai halaman Daftar Assignment termuat
            title_el = d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Daftar Assignment")
            print("Menunggu halaman 'Daftar Assignment' termuat...")
            if title_el.wait(exists=True, timeout=15.0):
                print("[BACK] Halaman 'Daftar Assignment' terdeteksi, melanjutkan...")
                time.sleep(SLEEP_MEDIUM)
                return True
            else:
                print("[WARNING] Halaman 'Daftar Assignment' tidak terdeteksi setelah klik Submit.")
    except Exception as e:
        print(f"[WARNING] Gagal saat press back / check periode: {e}")
    return False

def reset_ke_dashboard_dari_form():
    """Membersihkan form yang sedang terbuka dan kembali ke dashboard secara aman"""
    print("[RESET] Membersihkan form dan kembali ke Dashboard...")
    for _ in range(3):
        try:
            if press_back_and_check_periode():
                break
            time.sleep(SLEEP_SHORT)
            
            # Cek konfirmasi keluar (IYA / YA)
            iya_btn = d(text="IYA")
            if iya_btn.exists():
                iya_btn.click()
                time.sleep(SLEEP_MEDIUM)
                break
                
            ya_btn = d(text="Ya")
            if not ya_btn.exists():
                ya_btn = d(text="YA")
            if ya_btn.exists():
                ya_btn.click()
                time.sleep(SLEEP_MEDIUM)
                break
        except Exception as e:
            print(f"[RESET] Gagal pada langkah back/konfirmasi: {e}")
def ketuk(target_text, exact=False, sleep_after=SLEEP_SHORT):
    """
    Fungsi universal & reusable untuk mencari dan mengetuk elemen teks/description/XPath.
    Memprioritaskan elemen berjenis android.widget.Button atau yang bernilai clickable=True.
    """
    print(f"[KLIK] Mencari dan mengetuk '{target_text}'...")
    target_btn = None

    # 1. Prioritas 1: Target spesifik android.widget.Button
    try:
        btn_el = d(className="android.widget.Button", text=target_text)
        if not btn_el.exists():
            btn_el = d(className="android.widget.Button", textContains=target_text)
        if btn_el.exists():
            target_btn = btn_el
            print(f"[KLIK] Ditemukan via android.widget.Button: '{target_text}'")
    except Exception:
        pass

    # 2. Prioritas 2: Target elemen dengan clickable=True
    if not target_btn:
        try:
            if exact:
                click_el = d(text=target_text, clickable=True)
                if not click_el.exists():
                    click_el = d(description=target_text, clickable=True)
            else:
                click_el = d(textContains=target_text, clickable=True)
                if not click_el.exists():
                    click_el = d(descriptionContains=target_text, clickable=True)

            if click_el.exists():
                target_btn = click_el
                print(f"[KLIK] Ditemukan via clickable=True: '{target_text}'")
        except Exception:
            pass

    # 3. Prioritas 3: Target via XPath Button / clickable=True
    if not target_btn:
        try:
            if exact:
                xp = f"//android.widget.Button[@text='{target_text}' or @content-desc='{target_text}'] | //*[@clickable='true' and (@text='{target_text}' or @content-desc='{target_text}')]"
            else:
                xp = f"//android.widget.Button[contains(@text, '{target_text}') or contains(@content-desc, '{target_text}')] | //*[@clickable='true' and (contains(@text, '{target_text}') or contains(@content-desc, '{target_text}'))]"

            if d.xpath(xp).exists:
                target_btn = d.xpath(xp)
                print(f"[KLIK] Ditemukan via XPath clickable: '{xp}'")
        except Exception:
            pass

    # 4. Fallback selector umum jika di atas tidak ketemu
    if not target_btn:
        if exact:
            if d(text=target_text).exists():
                target_btn = d(text=target_text)
            elif d(description=target_text).exists():
                target_btn = d(description=target_text)
        else:
            if d(textContains=target_text).exists():
                target_btn = d(textContains=target_text)
                try:
                    print(f"[KLIK] Ditemukan via textContains: '{target_btn.info.get('text', target_text)}'")
                except Exception:
                    pass
            elif d(descriptionContains=target_text).exists():
                target_btn = d(descriptionContains=target_text)

    if not target_btn:
        try:
            if exact:
                xp = f"//*[@text='{target_text}' or @content-desc='{target_text}']"
            else:
                xp = f"//*[contains(@text, '{target_text}') or contains(@content-desc, '{target_text}')]"

            if d.xpath(xp).exists:
                target_btn = d.xpath(xp)
                print(f"[KLIK] Ditemukan via XPath fallback: '{xp}'")
        except Exception:
            pass

    # 5. Eksekusi Klik
    success = False
    if target_btn:
        try:
            target_btn.click()
            print(f"[KLIK] Berhasil mengetuk '{target_text}'")
            success = True
        except Exception:
            try:
                if exact:
                    d.xpath(f"//android.widget.Button[@text='{target_text}'] | //*[@text='{target_text}']").click()
                else:
                    d.xpath(f"//android.widget.Button[contains(@text, '{target_text}')] | //*[contains(@text, '{target_text}')]").click()
                print(f"[KLIK] Berhasil mengetuk '{target_text}' via XPath fallback click")
                success = True
            except Exception as err:
                print(f"[WARNING] Gagal mengetuk '{target_text}': {err}")

    # 6. Fallback khusus untuk tombol 'Aksi', 'Buka' / 'BUKA', dan 'sidebar-toggle' jika belum berhasil diklik
    if not success and target_text.upper() in ["BUKA", "AKSI", "SIDEBAR-TOGGLE", "OK"]:
        if target_text.upper() == "BUKA":
            print("[KLIK] Mengetuk tombol 'BUKA' via resourceId 'id.go.bpsfasih:id/openAssignment_b'...")
            try:
                if d(resourceId="id.go.bpsfasih:id/openAssignment_b").exists():
                    d(resourceId="id.go.bpsfasih:id/openAssignment_b").click()
                    success = True
            except Exception:
                pass
            if not success:
                print("[KLIK] Mengetuk tombol 'BUKA' via koordinat bounds [63,905][1017,1014] -> (540, 959)...")
                try:
                    d.click(540, 959)
                    success = True
                except Exception as e:
                    print(f"[ERROR] Gagal klik koordinat BUKA: {e}")
        elif target_text == "Aksi":
            print("[KLIK] Mengetuk tombol 'Aksi' via koordinat bounds [306,1296][468,1401] -> (387, 1348)...")
            try:
                d.click(387, 1348)
                success = True
            except Exception as e:
                print(f"[ERROR] Gagal klik koordinat Aksi: {e}")
        elif target_text == "sidebar-toggle":
            print("[KLIK] Mengetuk tombol 'sidebar-toggle' via koordinat bounds [24,114][120,210] -> (72, 162)...")
            try:
                d.click(72, 162)
                success = True
            except Exception as e:
                print(f"[ERROR] Gagal klik koordinat sidebar-toggle: {e}")
        elif target_text.upper() == "BLOK II":
            print("[KLIK] Mengetuk 'BLOK II' via koordinat sidebar bounds [570, 792] -> (350, 681)...")
            try:
                d.click(350, 681)
                success = True
            except Exception as e:
                print(f"[ERROR] Gagal klik koordinat BLOK II: {e}")
        elif target_text.upper() == "BATAL":
            print("[KLIK] Mengetuk tombol 'Batal' via koordinat bounds [564,1629][729,1725] -> (646, 1677)...")
            try:
                d.click(646, 1677)
                success = True
            except Exception as e:
                print(f"[ERROR] Gagal klik koordinat Batal: {e}")

    if sleep_after > 0:
        time.sleep(sleep_after)
    return success


def normalisasi_nama_desa(val):
    """
    Normalisasi nama Desa/Kelurahan khusus variasi Padang Sambian:
    - 'PADANG SAMBIAN KAJA' -> 'PADANGSAMBIAN KAJA'
    - 'PADANG SAMBIAN KELOD' / 'PADANGSAMBIAN KELOD' -> 'PADANGSAMBIAN KLOD'
    - 'PADANG SAMBIAN' -> 'PADANGSAMBIAN'
    """
    if not val:
        return val
    replacements = [
        ("PADANG SAMBIAN KELOD", "PADANGSAMBIAN KLOD"),
        ("PADANGSAMBIAN KELOD", "PADANGSAMBIAN KLOD"),
        ("Padang Sambian Kelod", "PADANGSAMBIAN KLOD"),
        ("Padangsambian Kelod", "PADANGSAMBIAN KLOD"),
        ("PADANG SAMBIAN KAJA", "PADANGSAMBIAN KAJA"),
        ("Padang Sambian Kaja", "PADANGSAMBIAN KAJA"),
        ("PADANG SAMBIAN", "PADANGSAMBIAN"),
        ("Padang Sambian", "PADANGSAMBIAN"),
    ]
    for old_str, new_str in replacements:
        if old_str in val:
            val = val.replace(old_str, new_str)
    return val


def baca_temp_alamat():
    # Default fallbacks jika file tidak ditemukan
    data = {
        "Provinsi": "[51] BALI",
        "Kabupaten": "[5103] KAB. BADUNG",
        "Kecamatan": "[510306] KUTA UTARA",
        "Desa/Kelurahan": "[5103061002] KEROBOKAN",
        "Alamat": "JL. GN. MANDALA"
    }
    
    file_path = "temp_alamat.txt"
    if os.path.exists(file_path):
        print(f"[EXCEL/TXT] Membaca data alamat dari {file_path}...")
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                for line in f:
                    if ":" in line:
                        key, val = line.split(":", 1)
                        key = key.strip()
                        val = val.strip()
                        # Normalisasi key agar cocok
                        if "Desa" in key or "Kelurahan" in key:
                            data["Desa/Kelurahan"] = normalisasi_nama_desa(val)
                        elif "Kabupaten" in key or "Kota" in key:
                            data["Kabupaten"] = val
                        elif key in data:
                            data[key] = val
            print("[SUKSES] Berhasil membaca data alamat:")
            for k, v in data.items():
                print(f"  - {k}: '{v}'")
        except Exception as e:
            print(f"[WARNING] Gagal membaca {file_path}: {e}. Menggunakan default.")
    else:
        print(f"[WARNING] File {file_path} tidak ditemukan. Menggunakan nilai default.")
        
    return data

def ekstrak_nama_saja(teks):
    # Hilangkan bagian "[...]" beserta kurung sikunya jika ada
    clean = re.sub(r'\[.*?\]', '', teks)
    return clean.strip()

def klik_opsi_dropdown(d, teks_pilihan):
    nama_saja = ekstrak_nama_saja(teks_pilihan)
    print(f"[DROPDOWN] Mencari opsi untuk clean name: '{nama_saja}' (dari '{teks_pilihan}')...")
    
    # Lakukan polling tunggu opsi muncul (maksimal 4 detik)
    for attempt in range(8):
        try:
            # Utamakan pencarian di dalam ListView atau Dialog (popup overlay dropdown)
            xpath_list = [
                # 1. Cari di dalam android.widget.ListView (opsi dropdown asli)
                f"//android.widget.ListView//*[contains(@text, '{teks_pilihan}')]",
                f"//android.widget.ListView//*[contains(@text, '{nama_saja}') and contains(@text, '[')]",
                f"//android.widget.ListView//*[contains(@text, '{nama_saja}')]",
                # 2. Backup: search di dalam android.app.Dialog
                f"//android.app.Dialog//*[contains(@text, '{teks_pilihan}')]",
                f"//android.app.Dialog//*[contains(@text, '{nama_saja}') and contains(@text, '[')]",
                f"//android.app.Dialog//*[contains(@text, '{nama_saja}')]",
            ]
            
            for xp in xpath_list:
                xpath_el = d.xpath(xp)
                if xpath_el.exists:
                    matching_elements = xpath_el.all()
                    target_el = None
                    
                    # Prioritaskan elemen yang teksnya paling cocok (misal: ending with "] NAMA" atau exact match)
                    for el in matching_elements:
                        txt = el.text or ""
                        # Abaikan jika elemen adalah EditText
                        if "EditText" in (el.info.get("className", "") or ""):
                            continue
                        if txt == teks_pilihan or txt.endswith(f"] {nama_saja}") or txt.endswith(f" {nama_saja}") or txt == nama_saja:
                            target_el = el
                            break
                    
                    if not target_el:
                        # Ambil elemen pertama yang bukan EditText
                        for el in matching_elements:
                            if "EditText" not in (el.info.get("className", "") or ""):
                                target_el = el
                                break
                    
                    if target_el:
                        print(f"[DROPDOWN] Menemukan opsi via XPath ('{xp}'): '{target_el.text}' -> Mengklik...")
                        target_el.click()
                        time.sleep(0.4)
                        return True
        except Exception as xpath_err:
            print(f"[DROPDOWN] [WARNING] Gagal mengevaluasi XPath: {xpath_err}")
            
        time.sleep(0.5)
            
    print(f"[DROPDOWN] [WARNING] Opsi '{teks_pilihan}' tidak ditemukan di layar. Menutup overlay.")
    d.click(540, 300)  # Klik area netral untuk menutup overlay
    time.sleep(0.5)
    return False

def ketuk_ok_submit_diproses(max_attempts=5):
    """
    Memeriksa dan mengetuk tombol 'OK' pada modal 'Submit diproses'.
    Jika setelah diketuk modal masih tetap muncul ('Submit diproses' masih ada di layar),
    akan mengulang pengetukan hingga `max_attempts` kali (default 5).
    Jika setelah 5x masih tidak bisa, menggunakan fallback pengetukan statis pada koordinat bounds (528, 1691).
    """
    print("\n[SUBMIT DIPROSES] Memeriksa modal 'Submit diproses'...")
    
    # Koordinat tengah tombol OK dari bounds [137,1639][919,1744] (id: btn_submit_progress_close)
    ok_bounds_x, ok_bounds_y = 528, 1691
    
    for attempt in range(1, max_attempts + 1):
        is_submit_modal = (
            d(textContains="Submit diproses").exists or
            d(resourceId="id.go.bpsfasih:id/tv_submit_progress_title").exists or
            d(resourceId="id.go.bpsfasih:id/btn_submit_progress_close").exists or
            d(text="OK").exists
        )
        
        if not is_submit_modal:
            print("[SUBMIT DIPROSES] Modal 'Submit diproses' tidak terdeteksi (sudah tertutup).")
            return True

        print(f"[SUBMIT DIPROSES] Terdeteksi modal 'Submit diproses'. Percobaan ketuk 'OK' ke-{attempt}/{max_attempts}...")
        
        clicked = False
        try:
            btn_close = d(resourceId="id.go.bpsfasih:id/btn_submit_progress_close")
            if btn_close.exists:
                btn_close.click()
                clicked = True
            elif d(text="OK").exists:
                d(text="OK").click()
                clicked = True
            else:
                clicked = ketuk("OK", sleep_after=SLEEP_SHORT)
        except Exception as e:
            print(f"[WARNING] Klik tombol OK via elemen gagal: {e}")

        if not clicked:
            print(f"[SUBMIT DIPROSES] Klik via elemen gagal. Mengetuk koordinat statis bounds ({ok_bounds_x}, {ok_bounds_y})...")
            d.click(ok_bounds_x, ok_bounds_y)

        time.sleep(2)

        still_exists = (
            d(textContains="Submit diproses").exists or
            d(resourceId="id.go.bpsfasih:id/tv_submit_progress_title").exists
        )
        if not still_exists:
            print(f"[SUBMIT DIPROSES] [SUKSES] Modal 'Submit diproses' berhasil ditutup pada percobaan ke-{attempt}.")
            return True
        else:
            print(f"[SUBMIT DIPROSES] [RETRY] Modal 'Submit diproses' masih ada di layar setelah percobaan ke-{attempt}.")

    # Fallback jika 5x percobaan elemen tidak berhasil menutup modal
    print(f"[SUBMIT DIPROSES] [FALLBACK STATIS] Sudah {max_attempts}x percobaan dan modal masih ada. Mengetuk koordinat statis bounds OK ({ok_bounds_x}, {ok_bounds_y})...")
    d.click(ok_bounds_x, ok_bounds_y)
    time.sleep(2)
    return True

def tampilkan_ringkasan_akhir(sheet):
    """Menampilkan rekapitulasi ringkasan data di akhir proses"""
    if sheet is None:
        return

    total_idpel = 0
    tanpa_foto_count = 0
    sukses_tercatat_count = 0
    error_count = 0

    for r in range(2, sheet.max_row + 1):
        val_idpel = sheet.cell(row=r, column=1).value
        if val_idpel is None or str(val_idpel).strip() == "":
            continue

        total_idpel += 1
        val_status = str(sheet.cell(row=r, column=9).value or "").strip()

        # Check keberadaan foto
        has_photo = False
        for ext in [".jpg", ".png", ".jpeg", ".JPG", ".PNG", ".JPEG"]:
            if os.path.exists(os.path.join(FOTO_DIRECTORY, f"{val_idpel}{ext}")):
                has_photo = True
                break

        if not has_photo:
            tanpa_foto_count += 1

        # Check status
        if val_status in ["SUKSES", "Error : SUDAH TERCATAT PADA SISTEM FASIH"]:
            sukses_tercatat_count += 1
        elif val_status != "":
            error_count += 1

    print("\n==================================================")
    print("      PEMBERITAHUAN / RINGKASAN AKHIR PROSES      ")
    print("==================================================")
    print(f" Total IDPEL                     : {total_idpel}")
    print(f" Jumlah IDPEL yang belum ada foto : {tanpa_foto_count}")
    print(f" Jumlah IDPEL SUKSES + TERCATAT   : {sukses_tercatat_count}")
    print(f" Jumlah IDPEL yang Error         : {error_count}")
    if total_idpel > 0 and total_idpel == tanpa_foto_count:
        print("\n [PERINGATAN] SEMUA FOTO TIDAK DITEMUKAN!")
        print(" Mohon periksa variable direktori foto (FOTO_DIRECTORY) apa sudah benar atau belum.")
    print("==================================================")

def main():
    import sys
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

    global d
    print("Menghubungkan ke emulator instance 2...")
    for port in EMULATOR_PORTS:
        try:
            if os.path.exists(LDPLAYER_ADB):
                subprocess.run([LDPLAYER_ADB, "connect", f"127.0.0.1:{port}"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            temp_d = u2.connect(f"127.0.0.1:{port}")
            _ = temp_d.info
            d = temp_d
            print(f"[KONEKSI] Berhasil terhubung ke emulator instance 2 di port {port}!")
            break
        except Exception:
            continue

    if not d:
        print("[ERROR] Gagal terhubung ke emulator. Pastikan LDPlayer sudah aktif.")
        return

    # Membuka file Excel
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] File Excel '{EXCEL_FILE}' tidak ditemukan!")
        return

    # Pengecekan apakah file Excel sedang dibuka di program lain (seperti Microsoft Excel)
    while True:
        try:
            with open(EXCEL_FILE, "r+"):
                pass
            break
        except (PermissionError, IOError):
            print(f"\n==================================================")
            print(f"[PERINGATAN] File '{EXCEL_FILE}' sedang DIBUKA di Excel/program lain!")
            print(f"Mohon CLOSE / TUTUP terlebih dahulu file '{EXCEL_FILE}'.")
            print(f"==================================================")
            input("Tekan ENTER di sini setelah Anda menutup file Excel tersebut untuk melanjutkan...")

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
    except Exception as e:
        print(f"[ERROR] Gagal membaca file Excel: {e}")
        return

    print(f"\n[START] Memulai pemrosesan data (Total baris: {sheet.max_row - 1})")

    for row in range(2, sheet.max_row + 1):
        idpel        = sheet.cell(row=row, column=1).value  # Column A: IDPEL
        no_meter     = sheet.cell(row=row, column=2).value  # Column B: NOMETER_BARU
        nama         = sheet.cell(row=row, column=3).value  # Column C: NAMA
        lat          = sheet.cell(row=row, column=5).value  # Column E: KOORDINAT_X
        lng          = sheet.cell(row=row, column=6).value  # Column F: KOORDINAT_Y
        nik          = sheet.cell(row=row, column=7).value  # Column G: NIK
        telp         = sheet.cell(row=row, column=8).value  # Column H: TELEPON
        status_exist = sheet.cell(row=row, column=9).value  # Column I: JENISLAYANAN (Status)

        # Bersihkan NAMA
        if nama is not None:
            nama = str(nama).strip()
        else:
            nama = ""

        # Bersihkan NIK jika berupa float/int
        if nik is not None:
            if isinstance(nik, float):
                nik = str(int(nik)).strip()
            else:
                nik = str(nik).strip()

        # Bersihkan nomor telepon jika berupa float/int
        if telp is not None:
            if isinstance(telp, float):
                telp = str(int(telp)).strip()
            else:
                telp = str(telp).strip()
        if not telp or telp == '0' or telp == "":
            telp = "-"

        # Berhenti jika baris IDPEL kosong
        if not idpel:
            break

        # Cek no_meter: jika tidak ada isinya (kosong), skip / lewati
        if no_meter is None or str(no_meter).strip() in ["", "None", "nan"]:
            print(f"[SKIP] Baris {row} | IDPEL {idpel} dilewati karena no_meter kosong.")
            continue

        # Lewati jika baris ini sudah berstatus SUKSES atau error tertentu
        if status_exist in ["SUKSES", "Error : SUDAH TERCATAT PADA SISTEM FASIH", "Error : Nik tidak valid", "Error : data alamat server kosong", "Error : Nomor meter tidak ditemukan", "Error : nama tidak murni alphabeth"]:
            print(f"[SKIP] Baris {row} | IDPEL {idpel} sudah diproses sebelumnya ({status_exist}).")
            continue

        # Validasi NAMA: Harus alphabet saja (a-z A-Z dan spasi). Skip jika mengandung karakter selain alphabet.
        if not nama or not all(c.isalpha() or c.isspace() for c in nama):
            print(f"[SKIP] Baris {row} | IDPEL {idpel} dilewati karena NAMA '{nama}' mengandung karakter selain alphabet a-z A-Z.")
            sheet.cell(row=row, column=9).value = "Error : nama tidak murni alphabeth"
            sheet.cell(row=row, column=10).value = time.strftime("%Y-%m-%d %H:%M:%S")
            try:
                wb.save(EXCEL_FILE)
            except Exception:
                pass
            continue

        # Cek foto: jika tidak ditemukan di folder, skip / lewati
        local_photo = None
        for ext in [".jpg", ".png", ".jpeg", ".JPG", ".PNG", ".JPEG"]:
            path_test = os.path.join(FOTO_DIRECTORY, f"{idpel}{ext}")
            if os.path.exists(path_test):
                local_photo = path_test
                break

        if not local_photo:
            print(f"[SKIP] Baris {row} | IDPEL {idpel} dilewati karena foto tidak ditemukan di '{FOTO_DIRECTORY}'.")
            continue

        # Validasi NIK berdasarkan aturan Kemendagri
        if not validasi_nik(nik):
            print(f"[SKIP] Baris {row} | IDPEL {idpel} NIK '{nik}' tidak valid berdasarkan aturan Kemendagri.")
            sheet.cell(row=row, column=9).value = "Error : Nik tidak valid"
            sheet.cell(row=row, column=10).value = time.strftime("%Y-%m-%d %H:%M:%S")
            try:
                wb.save(EXCEL_FILE)
            except Exception:
                pass
            continue

        print(f"\n--------------------------------------------------")
        print(f"Memproses Baris {row} | IDPEL: {idpel}")
        print(f"--------------------------------------------------")

        try:
            # ==========================================
            # PERSIAPAN DATA & INITIALIZATION
            # ==========================================
            # === LANGKAH 1: GPS SPOOFING ===
            if lat is not None and lng is not None:
                set_lokasi_gps(lat, lng)
            else:
                print("[WARNING] Koordinat GPS kosong. Menggunakan lokasi saat ini.")

            # === LANGKAH 2: CARI & UPLOAD FOTO TERLEBIH DAHULU ===
            ext = os.path.splitext(local_photo)[1]
            android_dest = f"/sdcard/Pictures/{idpel}{ext}"
            if not push_dan_scan_foto(local_photo, android_dest):
                print(f"[SKIP] Baris {row} | IDPEL {idpel} dilewati karena gagal upload foto ke emulator.")
                continue

            # ==========================================
            # NAVIGASI DASHBOARD
            # ==========================================
            # === CEK APAKAH SUDAH BERADA DI FORM INPUT NOMOR METER ===
            is_on_form = d(textContains="101b. Nomor meter").exists() or d(text="Cek Nomor Meter").exists()

            if not is_on_form:
                # === LANGKAH 3: TRANSISI KE DAFTAR ASSIGNMENT (SELF-HEALING) ===
                region_row = d(resourceId="id.go.bpsfasih:id/updateListingLayout")
                if region_row.exists():
                    print("[DASHBOARD] Berada di halaman SLS Wilayah. Mengklik wilayah untuk masuk...")
                    region_row.click()
                    time.sleep(SLEEP_LONG)

                # === LANGKAH 4: KLIK FAB & TAMBAH ASSIGNMENT ===
                ya_btn_found = False
                for attempt in range(1, 4):
                    print(f"[DASHBOARD] Mengklik FAB (Percobaan {attempt}/3)...")
                    fab = d(resourceId="id.go.bpsfasih:id/expendable_fab")
                    if fab.wait(exists=True, timeout=5):
                        fab.click()
                        time.sleep(SLEEP_SHORT)
                    else:
                        print(f"[WARNING] Tombol FAB tidak ditemukan pada percobaan {attempt}.")
                        time.sleep(1.0)
                        continue

                    print(f"[DASHBOARD] Mengklik 'Tambah Assignment' (Percobaan {attempt}/3)...")
                    tambah_btn = d(resourceId="id.go.bpsfasih:id/fab_addAssignment")
                    if not tambah_btn.exists():
                        tambah_btn = d(text="Tambah Assignment")
                    if tambah_btn.exists(timeout=5):
                        tambah_btn.click()
                        time.sleep(SLEEP_MEDIUM)
                    else:
                        print(f"[WARNING] Tombol 'Tambah Assignment' tidak ditemukan pada percobaan {attempt}.")
                        time.sleep(1.0)
                        continue

                    print("[DASHBOARD] Menyetujui konfirmasi dialog 'YA'...")
                    ya_btn = d(resourceId="id.go.bpsfasih:id/rButton_bottomDialog")
                    if not ya_btn.exists(): ya_btn = d(text="YA")
                    if not ya_btn.exists(): ya_btn = d(text="Ya")
                    if ya_btn.exists(timeout=5):
                        ya_btn.click()
                        print("Menunggu form termuat (5 detik)...")
                        time.sleep(5.0)
                        ya_btn_found = True
                        break
                    else:
                        print(f"[WARNING] Tombol konfirmasi 'YA' tidak muncul (Percobaan {attempt}/3). Mengulangi klik FAB -> Tambah Assignment...")
                        time.sleep(1.0)

                if not ya_btn_found:
                    raise Exception("Tombol konfirmasi 'YA' tidak ditemukan setelah 3 kali percobaan (FAB -> Tambah Assignment).")

                # ==========================================
                # PENGISIAN BLOK I (FORM UTAMA)
                # ==========================================

                # === SCAN TEKS 'Ambil Waktu' SEBELUM LANGKAH 5 ===
                found_scan_waktu = False
                for scan_attempt in range(1, 11):
                    print(f"[BLOK I] Scanning teks 'Ambil Waktu' (Percobaan {scan_attempt}/10)...")
                    if d(text="Ambil Waktu").wait(exists=True, timeout=1.0) or d(textContains="Ambil Waktu").exists():
                        found_scan_waktu = True
                        print(f"[BLOK I] Teks 'Ambil Waktu' ditemukan pada percobaan {scan_attempt}.")
                        break

                if not found_scan_waktu:
                    print(f"[SKIP] Baris {row} | IDPEL {idpel} dilewati karena teks 'Ambil Waktu' tidak ditemukan setelah 10 kali scan.")
                    continue

                # === LANGKAH 5: AMBIL WAKTU WAWANCARA ===
                print("[BLOK I] Mengetuk tombol 'Ambil Waktu'...")
                waktu_btn = d(text="Ambil Waktu")
                if waktu_btn.exists():
                    waktu_sukses = False
                    for attempt_waktu in range(1, 4):
                        print(f"[BLOK I] Mengetuk tombol 'Ambil Waktu' (Percobaan {attempt_waktu}/3)...")
                        waktu_btn.click()
                        time.sleep(SLEEP_SHORT)
                        
                        # Cek keberadaan dialog 'Apakah Anda yakin ingin mengambil waktu saat ini?'
                        dialog_waktu = d(textContains="Apakah Anda yakin ingin mengambil waktu saat ini")
                        if not dialog_waktu.exists():
                            dialog_waktu = d(textContains="mengambil waktu saat ini")
                        
                        if dialog_waktu.wait(exists=True, timeout=2.0) or d(text="ya").exists() or d(text="Ya").exists() or d(text="YA").exists():
                            print("[BLOK I] Teks konfirmasi waktu terdeteksi. Mengetuk tombol 'ya'...")
                            ya_waktu = d(text="ya")
                            if not ya_waktu.exists(): ya_waktu = d(text="Ya")
                            if not ya_waktu.exists(): ya_waktu = d(text="YA")
                            if ya_waktu.exists():
                                ya_waktu.click()
                                time.sleep(SLEEP_MEDIUM)
                            waktu_sukses = True
                            break
                        else:
                            print(f"[WARNING] Teks 'Apakah Anda yakin ingin mengambil waktu saat ini?' tidak muncul (Percobaan {attempt_waktu}/3). Mengulangi...")
                            time.sleep(1.0)
                    if not waktu_sukses:
                        print("[WARNING] Dialog konfirmasi waktu tidak muncul setelah 3 kali percobaan.")
                else:
                    print("[INFO] Waktu wawancara sudah terisi. Melewati langkah Ambil Waktu.")

                # === LANGKAH 7: SCROLL CEPAT KE BAWAH SAMPAI MENTOK ===
                print("[BLOK I] Men-scroll cepat ke bawah sampai mentok...")
                swipe_ke_bawah_cepat(times=5)
                time.sleep(SLEEP_SHORT)
            else:
                print("[INFO] Sudah berada di form pengisian Nomor Meter. Melewati Langkah 3, 4, 5 & 7.")

            # === LANGKAH 6: INPUT METER PELANGGAN ===
            print("[BLOK I] Mencari kolom input Nomor Meter...")
            id_label = d(textContains="101b. Nomor meter")
            if not id_label.exists():
                try:
                    d(scrollable=True).scroll.to(textContains="101b. Nomor meter")
                except Exception:
                    pass

            if id_label.wait(exists=True, timeout=5):
                id_input = id_label.down(className="android.widget.EditText")
                if id_input.exists():
                    print(f"[BLOK I] Menulis Nomor Meter via set_text: {no_meter}...")
                    id_input.set_text(str(no_meter))
                    time.sleep(SLEEP_SHORT)
                else:
                    raise Exception("Kolom input EditText Nomor Meter tidak ditemukan.")
            else:
                raise Exception("Label '101b. Nomor meter' tidak ditemukan.")

            # Kirim ENTER
            d.shell("input keyevent KEYCODE_ENTER")
            time.sleep(SLEEP_SHORT)

            print("[BLOK I] Mengetuk tombol 'Cek Nomor Meter'...")
            cek_id_btn = d(text="Cek Nomor Meter")
            if not cek_id_btn.exists():
                try:
                    d(scrollable=True).scroll.to(text="Cek Nomor Meter")
                except Exception:
                    pass

            if cek_id_btn.exists(timeout=5):
                cek_id_btn.click()
                
                # Menunggu respon API selesai (memantau progress bar)
                print("Menunggu progress bar loading selesai...")
                time.sleep(0.5) # Beri jeda agar progress bar sempat muncul
                d(resourceId="id.go.bpsfasih:id/card_progress").wait_gone(timeout=30)
                
                # Cek apakah limit API check-nometerpln tercapai (langsung via .exists())
                limit_msg = "Permintaan API check-nometerpln sudah terlampaui (limit)."
                if d(textContains=limit_msg).exists():
                    print("PROSES BOT BERHENTI KARENA BATAS CEK NO METER MELEWATI LIMIT")
                    tampilkan_ringkasan_akhir(sheet)
                    input("Press any key to exit")
                    sys.exit(1)
                
                # Cek respon status langsung (tanpa timeout 5 detik)
                target_status = "DITEMUKAN DAN BELUM TERCATAT PADA SISTEM FASIH"
                already_recorded = "DITEMUKAN DAN SUDAH TERCATAT PADA SISTEM FASIH"
                
                time.sleep(0.1)
                status_found = d(textContains=target_status).exists()
                
                if not status_found:
                    # Pengecekan kondisi jika Nomor Meter sudah terdaftar
                    is_already_registered = (
                        d(textContains=already_recorded).exists() or
                        d(textContains="Nomor Meter sudah terdaftar di FASIH").exists() or
                        d(textContains="Nomor meter sudah terdaftar di FASIH").exists() or
                        d(textContains="sudah terdaftar di FASIH").exists()
                    )

                    # Pengecekan kondisi jika Nomor Meter tidak ditemukan / status invalid
                    is_not_found = (
                        d(textContains="Nomor meter tidak ditemukan").exists() or
                        d(textContains="Nomor Meter tidak ditemukan").exists() or
                        d(textContains="nomor meter tidak ditemukan").exists() or
                        d(textContains="tidak ditemukan").exists() or
                        d(text="STATUS").exists()
                    )

                    if is_already_registered:
                        print(f"[BLOK I] Baris {row} | IDPEL {idpel}: Nomor Meter SUDAH TERDAFTAR di FASIH. Melewati recovery, lanjut ke IDPEL berikutnya...")
                        sheet.cell(row=row, column=9).value = "Error : SUDAH TERCATAT PADA SISTEM FASIH"
                        sheet.cell(row=row, column=10).value = time.strftime("%Y-%m-%d %H:%M:%S")
                        continue

                    if is_not_found:
                        print(f"[BLOK I] Baris {row} | IDPEL {idpel}: Nomor Meter TIDAK DITEMUKAN di FASIH. Melewati recovery, lanjut ke IDPEL berikutnya...")
                        sheet.cell(row=row, column=9).value = "Error : Nomor meter tidak ditemukan"
                        sheet.cell(row=row, column=10).value = time.strftime("%Y-%m-%d %H:%M:%S")
                        continue

                    # Jika tidak ada kondisi khusus terdeteksi tapi status target tidak ditemukan
                    raise Exception("Respon status target tidak ditemukan.")
            else:
                raise Exception("Tombol 'Cek Nomor Meter' tidak ditemukan.")

             # === LANGKAH 8: VALIDASI STATUS DITEMUKAN ===
            if status_found:
                print(f"[OK] Status valid: '{target_status}' ditemukan!")
                
                # Jeda agar UI selesai me-render data alamat PLN dari API
                print("[BLOK I] Menunggu UI memuat data alamat (0.8 detik)...")
                time.sleep(1)
                
                # Scroll ke bawah secara dinamis hingga elemen alamat terlihat di layar
                print("[BLOK I] Men-scroll ke bawah secara dinamis hingga data alamat terlihat...")
                max_swipes = 10
                for swipe_idx in range(1, max_swipes + 1):
                    if d(textContains="e. Alamat").exists():
                        print(f"[BLOK I] Teks alamat ditemukan di layar (pemeriksaan ke-{swipe_idx}).")
                        break
                    
                    print(f"[BLOK I] Performing dynamic swipe {swipe_idx} (540, 1650 -> 540, 500)...")
                    try:
                        d.swipe(540, 800, 540, 155, duration=0.1)
                        time.sleep(0.2)
                    except Exception as scroll_err:
                        print(f"[WARNING] Gagal swipe ke bawah pada percobaan {swipe_idx}: {scroll_err}")
                        break
                time.sleep(0.2)

                # Sistem pencarian & verifikasi label "a. Provinsi"
                def cek_label_provinsi():
                    patterns = ["a. Provinsi", "a.  Provinsi", "a.Provinsi", "Provinsi", "provinsi"]
                    for p in patterns:
                        if d(textContains=p).exists():
                            return True
                    try:
                        if d.xpath("//*[contains(@text, 'Provinsi') or contains(@text, 'provinsi')]").exists:
                            return True
                    except Exception:
                        pass
                    return False

                if not cek_label_provinsi():
                    print("[BLOK I] Label 'a. Provinsi' tidak terdeteksi. Men-scroll ke atas (max 10x)...")
                    for swipe_up_idx in range(1, 11):
                        if cek_label_provinsi():
                            print(f"[BLOK I] Label 'a. Provinsi' ditemukan di layar (pemeriksaan ke-{swipe_up_idx}).")
                            break
                        try:
                            d.swipe(540, 155, 540, 300, duration=0.1)
                            time.sleep(0.2)
                        except Exception as scroll_up_err:
                            print(f"[WARNING] Gagal swipe ke atas pada percobaan {swipe_up_idx}: {scroll_up_err}")
                            break
                    time.sleep(0.2)
                
                # === AMBIL DATA ALAMAT & SIMPAN KE temp_alamat.txt ===
                print("[BLOK I] Mengambil data alamat dari screen...")
                info_alamat = {
                    "Provinsi": "",
                    "Kabupaten": "",
                    "Kecamatan": "",
                    "Desa/Kelurahan": "",
                    "Alamat": ""
                }
                
                for key in info_alamat.keys():
                    # 1. Coba cari elemen dengan pencarian textContains
                    el = d(textContains=key)
                    if el.exists():
                        txt = el.info.get('text', '').strip()
                        # Jika teks sudah mengandung nilai langsung (ada ':' dan ada nilai setelahnya)
                        if ":" in txt and len(txt) > len(key) + 2:
                            info_alamat[key] = txt
                        else:
                            # Jika hanya label, coba cari EditText sibling (bulatan input) dulu, lalu TextView sibling
                            sibling = el.sibling(className="android.widget.EditText")
                            if not sibling.exists():
                                sibling = el.sibling(className="android.widget.TextView")
                                
                            if sibling.exists() and sibling.info.get('text', '').strip():
                                info_alamat[key] = f"{key}: {sibling.info.get('text', '').strip()}"
                            else:
                                # Coba cari EditText di bawahnya (down) dulu, lalu TextView di bawahnya
                                down_el = el.down(className="android.widget.EditText")
                                if not down_el.exists():
                                    down_el = el.down(className="android.widget.TextView")
                                    
                                if down_el.exists() and down_el.info.get('text', '').strip():
                                    info_alamat[key] = f"{key}: {down_el.info.get('text', '').strip()}"
                                else:
                                    info_alamat[key] = ""
                    
                    # 2. Jika belum ditemukan, coba cari menggunakan XPath contains text
                    if not info_alamat[key]:
                        try:
                            # Gunakan pola penelusuran XPath dinamis yang mencakup spasi non-breaking space
                            xpath_el = d.xpath(f"//*[contains(@text, '{key}')]")
                            if xpath_el.exists:
                                txt = xpath_el.all()[0].text.strip()
                                if ":" in txt and len(txt) > len(key) + 2:
                                    info_alamat[key] = txt
                                else:
                                    # Cari sibling EditText/TextView menggunakan XPath parent-child
                                    siblings = d.xpath(f"//*[contains(@text, '{key}')]/../*").all()
                                    if len(siblings) > 1:
                                        for sib in siblings:
                                            # Pastikan bukan label key itu sendiri dan memiliki teks yang bukan nama label
                                            sib_text = sib.text.strip()
                                            if key not in sib_text and sib_text and not (sib_text.startswith("a. ") or sib_text.startswith("b. ") or sib_text.startswith("c. ") or sib_text.startswith("d. ") or sib_text.startswith("e. ")):
                                                info_alamat[key] = f"{key}: {sib_text}"
                                                break
                        except Exception as e:
                            print(f"[WARNING] Gagal mencari via XPath untuk key '{key}': {e}")
                            
                    # Fallback jika benar-benar tidak ditemukan di screen
                    if not info_alamat[key]:
                        info_alamat[key] = f"{key}: (tidak ditemukan)"

                # Normalisasi nama Desa/Kelurahan jika ada variasi Padang Sambian
                if info_alamat.get("Desa/Kelurahan"):
                    info_alamat["Desa/Kelurahan"] = normalisasi_nama_desa(info_alamat["Desa/Kelurahan"])
                
                # Validasi jika data alamat server kosong / hanya kurung siku "[]"
                alamat_kosong = False
                if d(text="[]").exists() or d(text="[ ]").exists():
                    alamat_kosong = True
                else:
                    nilai_bersih = []
                    for k, v in info_alamat.items():
                        clean_v = v.split(":", 1)[1].strip() if ":" in v else v.strip()
                        nilai_bersih.append(clean_v)
                    if any(v in ["[]", "[ ]"] for v in nilai_bersih) or all(v in ["", "[]", "[ ]", "(tidak ditemukan)"] for v in nilai_bersih):
                        alamat_kosong = True

                if alamat_kosong:
                    print(f"[BLOK I] Data alamat server kosong (berisi '[]') untuk IDPEL {idpel}. Men-skip ke baris berikutnya...")
                    raise Exception("Error : data alamat server kosong")
                
                # Simpan ke file temp_alamat.txt
                try:
                    with open("temp_alamat.txt", "w", encoding="utf-8") as f_temp:
                        for key in ["Provinsi", "Kabupaten", "Kecamatan", "Desa/Kelurahan", "Alamat"]:
                            val = info_alamat[key]
                            # Pastikan formatnya selalu "Kunci: Nilai"
                            if not val.startswith(key):
                                val = f"{key}: {val}"
                            f_temp.write(val + "\n")
                    print("[BLOK I] Data alamat berhasil disimpan ke temp_alamat.txt:")
                    for key, val in info_alamat.items():
                        print(f"  - {val}")
                    
                    # while True:
                    #     input_val = input("[PAUSE] Ketik 'y' untuk melanjutkan atau 'n' untuk menghentikan script: ").strip().lower()
                    #     if input_val == 'y':
                    #         break
                    #     elif input_val == 'n':
                    #         print("[PAUSE] Eksekusi script dihentikan oleh pengguna secara aman.")
                    #         sys.exit(0)
                except Exception as write_err:
                    print(f"[WARNING] Gagal menulis ke temp_alamat.txt: {write_err}")
            else:
                # Cek jika ada error khusus dari API
                err_text = "Status tidak cocok atau API Gagal"
                for error_msg in ["Terjadi kesalahan", "tidak ditemukan", "Gagal", "client error"]:
                    err_el = d(textContains=error_msg)
                    if err_el.exists():
                        err_text = err_el.info['text']
                        break
                raise Exception(f"Validasi BPS Gagal: {err_text}")

            # === LANGKAH 9: PILIH HASIL PENDATAAN ===
            print("[BLOK I] Men-scroll ke opsi Hasil Pendataan...")
            # Karena sebelumnya kita scroll ke paling atas, kita perlu swipe ke bawah lebih banyak
            swipe_ke_bawah_cepat(times=5)
            time.sleep(SLEEP_SHORT)

            clicked = False
            
            # Coba menggunakan kalkulasi koordinat dinamis dari test_click.py
            try:
                label_el = None
                # Cari elemen teks Berhasil didata yang lebar (bukan checkbox kecil)
                for pattern in ["1. Berhasil didata", "1.  Berhasil didata", "Berhasil didata"]:
                    elements = d.xpath(f"//*[contains(@text, '{pattern}')]").all()
                    for el in elements:
                        bounds_str = el.attrib.get('bounds')
                        import re
                        pts = [int(x) for x in re.findall(r'\d+', bounds_str)]
                        if len(pts) == 4:
                            x1, y1, x2, y2 = pts
                            if (x2 - x1) > 50:
                                label_el = el
                                break
                    if label_el:
                        break

                if label_el:
                    label_bounds_str = label_el.attrib.get('bounds')
                    import re
                    l_pts = [int(x) for x in re.findall(r'\d+', label_bounds_str)]
                    label_left, label_top, label_right, label_bottom = l_pts
                    
                    label_text = label_el.text
                    parent = d.xpath(f"//*[contains(@text, '{label_text}')]/..")
                    
                    if parent.exists:
                        parent_bounds = parent.all()[0].attrib.get('bounds')
                        p_pts = [int(x) for x in re.findall(r'\d+', parent_bounds)]
                        parent_left, parent_top, parent_right, parent_bottom = p_pts
                        
                        # Hitung koordinat lingkaran radio button di sebelah kiri label secara dinamis
                        click_x = parent_left + (label_left - parent_left) // 2
                        click_y = label_top + (label_bottom - label_top) // 2
                        
                        print(f"[KLIK] Parent Left: {parent_left}, Label Left: {label_left}")
                        print(f"[KLIK] Mengklik koordinat dinamis radio button: ({click_x}, {click_y})")
                        d.click(click_x, click_y)
                        time.sleep(SLEEP_SHORT)
                        clicked = True
            except Exception as click_err:
                print(f"[WARNING] Percobaan klik koordinat dinamis gagal: {click_err}")

            # Fallback jika klik koordinat dinamis gagal
            if not clicked:
                try:
                    matched_elements = d.xpath("//*[contains(@text, 'Berhasil didata') or contains(@text, 'Berhasil Didata') or contains(@text, 'berhasil didata')]").all()
                    if matched_elements:
                        for el in matched_elements:
                            print(f"[FALLBACK] Mengklik elemen Berhasil didata: '{el.text}'")
                            el.click()
                            time.sleep(SLEEP_SHORT)
                        clicked = True
                except Exception as xpath_err:
                    print(f"[WARNING] Fallback xpath gagal: {xpath_err}")

            if not clicked:
                # Fallback ke selector standar jika yang lainnya gagal
                for pattern in ["1. Berhasil didata", "1.  Berhasil didata", "Berhasil didata"]:
                    radio_btn = d(textContains=pattern)
                    if radio_btn.exists():
                        radio_btn.click()
                        clicked = True
                        time.sleep(SLEEP_SHORT)
                        break

            if not clicked:
                raise Exception("Opsi Hasil Pendataan 'Berhasil didata' tidak ditemukan.")

            # === SCROLL MENTOK KE BAWAH UNTUK LANGKAH BERIKUTNYA ===
            print("[BLOK I] Men-scroll ke bawah sampai mentok...")
            swipe_ke_bawah_cepat(times=4)
            time.sleep(SLEEP_LONG)

            # === LANGKAH 10: PILIH FOTO DARI GALERI ===
            foto_berhasil_diunggah = False
            for attempt in range(1, 3):
                print(f"\n[BLOK I] === SELEKSI FOTO (Percobaan {attempt}/2) ===")
                
                pilih_foto_clicked = False
                for click_try in range(3):
                    try:
                        pilih_foto_btn = d(text="Pilih", className="android.widget.Button")
                        if not pilih_foto_btn.exists() and click_try == 0:
                            print("[BLOK I] Tombol 'Pilih' tidak terlihat. Kemungkinan stuck di Galeri. Menekan tombol BACK...")
                            d.press("back")
                            time.sleep(SLEEP_LONG)
                            pilih_foto_btn = d(text="Pilih", className="android.widget.Button")
                        
                        if not pilih_foto_btn.exists():
                            print(f"[BLOK I] Tombol 'Pilih' tidak ditemukan (percobaan {click_try+1}). Melakukan scroll ke atas dan ke bawah...")
                            swipe_ke_atas_cepat(times=2)
                            time.sleep(SLEEP_SHORT)
                            swipe_ke_bawah_cepat(times=2)
                            time.sleep(SLEEP_MEDIUM)
                            pilih_foto_btn = d(text="Pilih", className="android.widget.Button")

                        if pilih_foto_btn.exists(timeout=3):
                            print(f"[BLOK I] Mengetuk tombol 'Pilih' Foto (percobaan {click_try+1})...")
                            pilih_foto_btn.click()
                            pilih_foto_clicked = True
                            time.sleep(SLEEP_LONG)
                            break
                    except Exception as err_pilih:
                        print(f"[WARNING] Gagal mengetuk tombol 'Pilih' ({err_pilih}). Mencoba ulang...")
                        time.sleep(SLEEP_LONG)

                if not pilih_foto_clicked:
                    print("[WARNING] Tombol 'Pilih' Foto tidak dapat diklik. Menekan BACK 2 kali, men-skip upload foto, dan lanjut ke langkah berikutnya...")
                    for _ in range(2):
                        d.press("back")
                        time.sleep(SLEEP_LONG)
                    foto_berhasil_diunggah = True
                    break

                print("[BLOK I] Mengetuk opsi 'GALERI'...")
                galeri_btn = d(text="GALERI")
                if not galeri_btn.exists(): galeri_btn = d(text="Galeri")
                if not galeri_btn.exists(): galeri_btn = d(text="galeri")
                if galeri_btn.exists(timeout=5):
                    galeri_btn.click()
                    time.sleep(SLEEP_MEDIUM)
                else:
                    raise Exception("Pilihan 'GALERI' tidak ditemukan.")

                print("[BLOK I] Mengetuk ikon burger di kiri atas...")
                burger_btn = None
                for desc in ["Show roots", "Tampilkan laci", "Tampilkan root", "Show navigation drawer", "Open navigation drawer", "Laci navigasi", "Menu"]:
                    if d(descriptionContains=desc).exists():
                        burger_btn = d(descriptionContains=desc)
                        break
                if not burger_btn:
                    for res_id in ["android:id/home", "com.android.documentsui:id/toolbar"]:
                        if d(resourceId=res_id).exists():
                            burger_btn = d(resourceId=res_id)
                            break
                if not burger_btn:
                    image_buttons = d(className="android.widget.ImageButton")
                    if image_buttons.exists():
                        burger_btn = image_buttons[0]

                if burger_btn and burger_btn.exists(timeout=5):
                    burger_btn.click()
                    time.sleep(SLEEP_SHORT)
                else:
                    print("[WARNING] Ikon burger tidak ditemukan, mencoba melanjutkan...")

                print("[BLOK I] Mengetuk opsi 'Images' atau 'Gambar'...")
                images_btn = None
                for text_val in ["Images", "Gambar", "images", "gambar"]:
                    if d(text=text_val, resourceId="android:id/title").exists():
                        images_btn = d(text=text_val, resourceId="android:id/title")
                        break
                if not images_btn:
                    for text_val in ["Images", "Gambar", "images", "gambar"]:
                        if d(text=text_val).exists():
                            images_btn = d(text=text_val)
                            break
                            
                if images_btn and images_btn.exists(timeout=5):
                    images_btn.click()
                    time.sleep(SLEEP_MEDIUM)
                else:
                    print("[WARNING] Opsi 'Images'/'Gambar' tidak ditemukan, mencoba melanjutkan...")

                # --- PENCARIAN FOTO BERDASARKAN IDPEL ---
                print("[BLOK I] Mencari tombol search/cari di Galeri...")
                search_btn = None
                for search_id in ["com.android.documentsui:id/option_menu_search", "com.google.android.documentsui:id/option_menu_search"]:
                    if d(resourceId=search_id).exists():
                        search_btn = d(resourceId=search_id)
                        break
                        
                if not search_btn:
                    for desc in ["Search", "Cari", "Search query", "Temukan"]:
                        if d(descriptionContains=desc).exists():
                            search_btn = d(descriptionContains=desc)
                            break

                if search_btn:
                    print("[BLOK I] Tombol search ditemukan. Mengklik...")
                    search_btn.click()
                    time.sleep(SLEEP_SHORT)
                    
                    # Cari input field pencarian
                    search_input = None
                    for input_id in ["com.android.documentsui:id/search_src_text", "com.google.android.documentsui:id/search_src_text"]:
                        if d(resourceId=input_id).exists():
                            search_input = d(resourceId=input_id)
                            break
                    if not search_input:
                        search_input = d(className="android.widget.EditText")
                        
                    if search_input and search_input.exists():
                        print(f"[BLOK I] Memasukkan teks pencarian IDPEL: '{idpel}'...")
                        search_input.set_text(str(idpel))
                        time.sleep(SLEEP_SHORT)
                        d.shell("input keyevent KEYCODE_ENTER")
                        time.sleep(SLEEP_MEDIUM)
                        
                        # Cek apakah hasil pencarian kosong (No matches in Images)
                        no_matches = d(text="No matches in Images")
                        if not no_matches.exists():
                            no_matches = d(resourceId="com.android.documentsui:id/message", text="No matches in Images")
                        if not no_matches.exists():
                            no_matches = d(textContains="No matches")
                            
                        if no_matches.exists():
                            print(f"[BLOK I] File foto tidak ditemukan di Android (No matches in Images) untuk IDPEL {idpel}!")
                            raise Exception("No matches in Images")
                    else:
                        print("[WARNING] Input pencarian tidak ditemukan. Menekan BACK sampai menemukan label 'Foto rumah tampak depan'...")
                        for _ in range(10):
                            if d(textContains="Foto rumah tampak depan").exists():
                                print("[BLOK I] Kembali ke form utama. Label 'Foto rumah tampak depan' ditemukan.")
                                break
                            d.press("back")
                            time.sleep(1.0)
                        foto_berhasil_diunggah = True
                        break
                else:
                    print("[WARNING] Tombol search tidak ditemukan di layar. Melanjutkan pencarian langsung...")

                print(f"[BLOK I] Memilih item foto dengan nama IDPEL: {idpel}...")
                foto_terpilih = False
                
                # Tentukan nama file yang dicari (harus ada ekstensi .jpg/.png)
                photo_filenames = [f"{idpel}.jpg", f"{idpel}.png", f"{idpel}.JPEG", f"{idpel}.PNG"]
                if local_photo:
                    photo_filenames = [os.path.basename(local_photo)] + photo_filenames

                # Metode 1: Pencarian teks persis (exact text match)
                for name in photo_filenames:
                    el_text = d(text=name)
                    if el_text.exists():
                        print(f"[BLOK I] Menemukan elemen foto (Metode 1: Exact Text)='{name}'")
                        bounds = el_text.info.get('bounds')
                        click_x = (bounds['left'] + bounds['right']) // 2
                        click_y = (bounds['top'] + bounds['bottom']) // 2
                        print(f"[BLOK I] Mengklik koordinat teks: ({click_x}, {click_y})")
                        d.click(click_x, click_y)
                        foto_terpilih = True
                        break
                
                # Metode 2: Pencarian textContains (harus ada ekstensi .jpg/.png)
                if not foto_terpilih:
                    for ext in [".jpg", ".png", ".JPEG", ".PNG"]:
                        target_name = f"{idpel}{ext}"
                        el_contains = d(textContains=target_name)
                        if el_contains.exists():
                            print(f"[BLOK I] Menemukan elemen foto (Metode 2: Contains)='{target_name}'")
                            bounds = el_contains.info.get('bounds')
                            click_x = (bounds['left'] + bounds['right']) // 2
                            click_y = (bounds['top'] + bounds['bottom']) // 2
                            print(f"[BLOK I] Mengklik koordinat teks: ({click_x}, {click_y})")
                            d.click(click_x, click_y)
                            foto_terpilih = True
                            break
                
                # Metode 3: Pencarian XPath contains text (harus ada ekstensi .jpg/.png)
                if not foto_terpilih:
                    for ext in [".jpg", ".png", ".JPEG", ".PNG"]:
                        target_name = f"{idpel}{ext}"
                        xpath_elements = d.xpath(f"//*[contains(@text, '{target_name}')]").all()
                        if xpath_elements:
                            print(f"[BLOK I] Menemukan elemen foto (Metode 3: XPath Contains)='{target_name}'")
                            import re
                            bounds_str = xpath_elements[0].attrib.get('bounds', '')
                            pts = [int(x) for x in re.findall(r'\d+', bounds_str)]
                            if len(pts) == 4:
                                x1, y1, x2, y2 = pts
                                click_x = (x1 + x2) // 2
                                click_y = (y1 + y2) // 2
                                print(f"[BLOK I] Mengklik koordinat teks XPath: ({click_x}, {click_y})")
                                d.click(click_x, click_y)
                                foto_terpilih = True
                                break
                            else:
                                xpath_elements[0].click()
                                foto_terpilih = True
                                break

                # Fallback jika pencarian nama file gagal/tidak ada (Hindari klik icon_thumb agar tidak multi-select)
                if not foto_terpilih:
                    print(f"[WARNING] Foto dengan nama IDPEL tidak ditemukan, mencoba memilih judul/teks item pertama...")
                    first_title = None
                    for title_id in ["android:id/title", "com.android.documentsui:id/title", "com.google.android.documentsui:id/title"]:
                        if d(resourceId=title_id).exists():
                            first_title = d(resourceId=title_id)
                            break
                    
                    if first_title and first_title.exists():
                        bounds = first_title.info.get('bounds')
                        click_x = (bounds['left'] + bounds['right']) // 2
                        click_y = (bounds['top'] + bounds['bottom']) // 2
                        print(f"[BLOK I] Mengklik koordinat judul item pertama: ({click_x}, {click_y})")
                        d.click(click_x, click_y)
                        foto_terpilih = True
                    else:
                        # Klik koordinat cadangan di tengah-kanan baris pertama
                        print("[BLOK I] Mengklik koordinat cadangan baris pertama (500, 500)...")
                        d.click(500, 500)
                        foto_terpilih = True
                time.sleep(SLEEP_MEDIUM)

                # SAFETY NET: Cek jika masuk ke mode multi-select (muncul tombol OPEN di kanan atas)
                open_btn = d(text="OPEN")
                if not open_btn.exists(): open_btn = d(text="Open")
                if not open_btn.exists(): open_btn = d(text="OPENS")
                if not open_btn.exists(): open_btn = d(text="BUKA")
                if not open_btn.exists(): open_btn = d(text="Buka")
                if open_btn.exists():
                    print("[BLOK I] Mendeteksi mode multi-select aktif. Mengklik tombol 'OPEN'...")
                    open_btn.click()
                    time.sleep(SLEEP_MEDIUM)

                # Cek text label "Dimuat dari local" jika ada maka ketuk "Unggah"
                print("[BLOK I] Memeriksa label 'Dimuat dari local'...")
                dimuat_local = d(textContains="Dimuat dari local")
                if dimuat_local.wait(exists=True, timeout=10):
                    sudah_terunggah = d(textContains="Sudah Terunggah")
                    for try_unggah in range(1, 4):
                        print(f"[BLOK I] Mengetuk tombol 'Unggah' (Pengetukan {try_unggah}/3)...")
                        unggah_btn = d(text="Unggah")
                        if not unggah_btn.exists():
                            unggah_btn = d(text="UNGGAH")
                        if not unggah_btn.exists():
                            unggah_btn = d(textContains="Unggah")
                        
                        if unggah_btn.exists(timeout=10):
                            unggah_btn.click()
                            time.sleep(SLEEP_MEDIUM)
                            
                            # Ketuk "Ya" / "YA" konfirmasi setelah klik Unggah
                            ya_btn = d(text="Ya")
                            if not ya_btn.exists(): ya_btn = d(text="YA")
                            if not ya_btn.exists(): ya_btn = d(text="ya")
                            if ya_btn.exists(timeout=5):
                                print("[BLOK I] Menyetujui konfirmasi unggah...")
                                ya_btn.click()
                                time.sleep(SLEEP_MEDIUM)
                        else:
                            print(f"[WARNING] Tombol 'Unggah' tidak ditemukan pada pengetukan {try_unggah}/3.")
                        
                        print(f"[BLOK I] Menunggu status 'Sudah Terunggah' (Pengetukan {try_unggah}/3)...")
                        if sudah_terunggah.wait(exists=True, timeout=10):
                            print(f"[BLOK I] Sukses: Foto berhasil diunggah ('Sudah Terunggah' terkonfirmasi pada pengetukan {try_unggah}).")
                            foto_berhasil_diunggah = True
                            break
                        else:
                            print(f"[WARNING] Pengetukan {try_unggah}/3 belum mengubah status ke 'Sudah Terunggah'.")

                    if foto_berhasil_diunggah:
                        break
                else:
                    print("[BLOK I] Label 'Dimuat dari local' tidak muncul, memeriksa apakah sudah terunggah...")
                    sudah_terunggah = d(textContains="Sudah Terunggah")
                    if sudah_terunggah.exists():
                        print("[BLOK I] Label 'Sudah Terunggah' sudah aktif.")
                        foto_berhasil_diunggah = True
                        break
                    else:
                        print("[WARNING] Baik 'Dimuat dari local' maupun 'Sudah Terunggah' tidak terdeteksi pada percobaan ini.")

            if not foto_berhasil_diunggah:
                raise Exception("Gagal mengunggah foto setelah 2 kali percobaan.")

            # === LANGKAH 11: AMBIL LOKASI GPS ===
            ambil_lokasi_btn = d(text="Ambil Lokasi")
            if not ambil_lokasi_btn.exists():
                print("[BLOK I] Tombol 'Ambil Lokasi' tidak terdeteksi. Men-scroll ke bawah (max 10x)...")
                for swipe_down_idx in range(1, 11):
                    if d(text="Ambil Lokasi").exists():
                        print(f"[BLOK I] Tombol 'Ambil Lokasi' ditemukan di layar (pemeriksaan ke-{swipe_down_idx}).")
                        break
                    try:
                        d.swipe(540, 800, 540, 500, duration=0.1)
                        time.sleep(0.2)
                    except Exception as scroll_down_err:
                        print(f"[WARNING] Gagal swipe ke bawah pada percobaan {swipe_down_idx}: {scroll_down_err}")
                        break
                time.sleep(0.2)

            ambil_lokasi_btn = d(text="Ambil Lokasi")
            if ambil_lokasi_btn.wait(exists=True, timeout=5):
                print("[BLOK I] Mengetuk tombol 'Ambil Lokasi'...")
                ambil_lokasi_btn.click()
                time.sleep(SLEEP_SHORT)
            else:
                raise Exception("Tombol 'Ambil Lokasi' tidak ditemukan.")

            print("[BLOK I] Mengetuk opsi 'AMBIL LANGSUNG'...")
            opsi_lokasi = d(resourceId="id.go.bpsfasih:id/lButton_bottomDialog")
            if not opsi_lokasi.exists():
                opsi_lokasi = d(text="AMBIL LANGSUNG")
            if not opsi_lokasi.exists():
                opsi_lokasi = d(textContains="LANGSUNG")
                
            if opsi_lokasi.exists(timeout=5):
                opsi_lokasi.click()
                time.sleep(SLEEP_SHORT)
            else:
                raise Exception("Opsi 'AMBIL LANGSUNG' tidak ditemukan.")

            print("[BLOK I] Menyetujui konfirmasi lokasi...")
            dialog_keluar = d(text="Apakah Anda yakin akan keluar dari halaman ini ?")
            if not dialog_keluar.exists():
                dialog_keluar = d(textContains="Apakah Anda yakin akan keluar dari halaman ini")
            if dialog_keluar.exists(timeout=3):
                print("[WARNING] Muncul dialog konfirmasi keluar halaman. Menekan tombol 'tidak'...")
                tidak_btn = d(text="tidak")
                if not tidak_btn.exists(): tidak_btn = d(text="Tidak")
                if not tidak_btn.exists(): tidak_btn = d(text="TIDAK")
                if tidak_btn.exists():
                    tidak_btn.click()
                    time.sleep(SLEEP_SHORT)

            ya_lokasi = d(text="ya")
            if not ya_lokasi.exists(): ya_lokasi = d(text="Ya")
            if not ya_lokasi.exists(): ya_lokasi = d(text="YA")
            if ya_lokasi.exists(timeout=5):
                ya_lokasi.click()
                time.sleep(SLEEP_MEDIUM)
            else:
                raise Exception("Dialog konfirmasi lokasi tidak muncul.")

            # Transisi ke BLOK II menggunakan tombol BERIKUTNYA secara dinamis
            print("[BLOK I] Mengetuk tombol 'BERIKUTNYA'...")
            transisi_sukses = False
            
            # Cari tombol BERIKUTNYA BLOK II di layar
            berikutnya_btn = d(textContains="BERIKUTNYA BLOK II")
            if not berikutnya_btn.exists(): berikutnya_btn = d(textContains="BERIKUTNYA")
            
            if berikutnya_btn.exists(timeout=5):
                # Klik menggunakan koordinat tengah secara dinamis
                bounds = berikutnya_btn.info.get('bounds')
                click_x = (bounds['left'] + bounds['right']) // 2
                click_y = (bounds['top'] + bounds['bottom']) // 2
                print(f"[BLOK I] Mengetuk koordinat tombol 'BERIKUTNYA': ({click_x}, {click_y})")
                d.click(click_x, click_y)
                
                # Verifikasi transisi halaman dengan mempolling field '201. Nama penghuni'
                print("[BLOK I] Menunggu halaman BLOK II termuat...")
                for poll_attempt in range(30):
                    time.sleep(0.2)
                    try:
                        # Paksa refresh cache uiautomator2 agar melihat perubahan WebView
                        d.dump_hierarchy()
                    except Exception:
                        pass
                    if d(textContains="201. Nama penghuni").exists():
                        print("[BLOK I] Sukses berpindah ke halaman 'BLOK II' via tombol.")
                        transisi_sukses = True
                        time.sleep(0.2) # Jeda tambahan agar UI benar-benar stabil sebelum proses selanjutnya
                        break
            else:
                # Jika tombol tidak ada tapi text 201 sudah ada, artinya sudah di halaman tujuan
                if d(textContains="201. Nama penghuni").exists():
                    transisi_sukses = True
                    
            if not transisi_sukses:
                print("[WARNING] Gagal melakukan konfirmasi transisi halaman ke BLOK II via tombol.")
                raise Exception("Gagal melakukan konfirmasi transisi halaman ke BLOK II via tombol.")



            # ==========================================
            # PENGISIAN BLOK II (KETERANGAN PERUMAHAN)
            # ==========================================
            print("\n[BLOK II] Memulai pemrosesan BLOK II...")
            
            # Ambil nama penghuni dari kolom C (kolom 3) untuk baris saat ini
            nama_penghuni = sheet.cell(row=row, column=3).value
            print(f"[BLOK II] Nama Penghuni (Cell C{row}): '{nama_penghuni}'")

            # 1. Cek text "BLOK II" (Verifikasi sederhana tanpa scroll)
            blok2_header = d(textContains="BLOK II")
            if blok2_header.exists(timeout=5):
                print("[BLOK II] Berhasil berada di bagian 'BLOK II'.")
            else:
                print("[WARNING] Header 'BLOK II' tidak terdeteksi, melanjutkan pengisian...")

            # 2. Mengisi "201. Nama penghuni" dan verifikasi pengisian (max 3x percobaan)
            print("[BLOK II] Mengisi '201. Nama penghuni'...")
            nama_terisi = False
            for try_nama in range(1, 4):
                try:
                    # Pastikan label 201 terlihat di layar
                    label_201 = d(textContains="201. Nama penghuni")
                    if not label_201.exists():
                        label_201 = d(textContains="201.")
                    if not label_201.exists():
                        try:
                            d(scrollable=True).scroll.to(textContains="201.")
                        except Exception:
                            pass

                    # Target spesifik EditText 201 berdasarkan hirarki dump.xml
                    target_input = None
                    xp_input = d.xpath("//*[contains(@text, '201.')]/following::android.widget.EditText[1]")
                    if xp_input.exists:
                        target_input = xp_input
                    elif label_201.exists():
                        target_input = label_201.down(className="android.widget.EditText")
                    else:
                        target_input = d(className="android.widget.EditText")

                    if target_input:
                        try:
                            if hasattr(target_input, 'click'):
                                target_input.click()
                            time.sleep(SLEEP_SHORT)
                        except Exception:
                            pass

                        target_input.set_text(str(nama_penghuni))
                        time.sleep(SLEEP_SHORT)
                        
                        # Scan / Verifikasi apakah teks Nama Penghuni sudah berhasil terisi
                        if d(textContains=str(nama_penghuni)).exists():
                            print(f"[BLOK II] Berhasil mengisi & memverifikasi Nama Penghuni: '{nama_penghuni}' (Percobaan {try_nama}/3)")
                            nama_terisi = True
                            break
                        else:
                            val_check = ""
                            try:
                                val_check = d(className="android.widget.EditText").get_text() or ""
                            except Exception:
                                pass
                            if str(nama_penghuni).strip() in val_check:
                                print(f"[BLOK II] Berhasil mengisi & memverifikasi Nama Penghuni: '{nama_penghuni}' (Percobaan {try_nama}/3)")
                                nama_terisi = True
                                break
                            else:
                                print(f"[WARNING] Teks Nama Penghuni belum terinput pada percobaan {try_nama}/3. Mengulang pengisian...")
                    else:
                        print(f"[WARNING] Input text box '201. Nama penghuni' tidak ditemukan pada percobaan {try_nama}/3.")
                except Exception as err_input:
                    print(f"[WARNING] Gagal input Nama Penghuni pada percobaan {try_nama}/3: {err_input}")
                time.sleep(SLEEP_SHORT)

            if not nama_terisi:
                raise Exception(f"Gagal mengisi Nama Penghuni '{nama_penghuni}' setelah 3 kali percobaan.")

            
            # # Input manual untuk jeda / konfirmasi pengguna setelah proses baris selesai
            # while True:
            #     input_val = input("[PAUSE] Ketik 'y' untuk melanjutkan ke baris berikutnya atau 'n' untuk menghentikan script: ").strip().lower()
            #     if input_val == 'y':
            #         break
            #     elif input_val == 'n':
            #         print("[PAUSE] Eksekusi script dihentikan oleh pengguna secara aman.")
            #         sys.exit(0)
            
            # 3. Mengisi "202. NIK penghuni"
            print("[BLOK II] Mengisi '202. NIK penghuni'...")
            label_202 = d(textContains="202. NIK penghuni")
            if not label_202.exists():
                d(scrollable=True).scroll.to(textContains="202. NIK penghuni")
                
            if label_202.exists(timeout=2):
                input_202 = label_202.down(className="android.widget.EditText")
                if input_202.exists():
                    input_202.set_text(str(nik))
                    print(f"[BLOK II] Berhasil mengisi NIK Penghuni: '{nik}'")
                    time.sleep(SLEEP_SHORT)
                else:
                    raise Exception("Input text box untuk '202. NIK penghuni' tidak ditemukan.")
            else:
                raise Exception("Label '202. NIK penghuni' tidak ditemukan.")

            # 4. Ketuk "Cek NIK"
            print("[BLOK II] Mengetuk tombol 'Cek NIK'...")
            cek_nik_btn = d(text="Cek NIK")
            if not cek_nik_btn.exists():
                d(scrollable=True).scroll.to(text="Cek NIK")
                
            if cek_nik_btn.exists(timeout=5):
                # Klik koordinat tengah agar aman
                bounds = cek_nik_btn.info.get('bounds')
                click_x = (bounds['left'] + bounds['right']) // 2
                click_y = (bounds['top'] + bounds['bottom']) // 2
                d.click(click_x, click_y)
                print("[BLOK II] Tombol 'Cek NIK' diketuk. Menunggu respon Cek NIK selesai...")
                time.sleep(SLEEP_LONG)
                d(resourceId="id.go.bpsfasih:id/card_progress").wait_gone(timeout=20)
                time.sleep(SLEEP_MEDIUM)

                # Verifikasi hasil Cek NIK
                limit_nik_msg = "Permintaan API check-nikpln sudah terlampaui (limit)."
                if d(textContains=limit_nik_msg).exists() or d(textContains="check-nikpln sudah terlampaui").exists():
                    print("[BLOK II] Permintaan API check-nikpln sudah terlampaui (limit). Melanjutkan ke langkah 5...")
                elif d(textContains="TIDAK DITEMUKAN").exists():
                    print("[BLOK II] Hasil Cek NIK berstatus 'TIDAK DITEMUKAN'. Melanjutkan ke langkah 5...")
                elif not d(textContains="DITEMUKAN").exists():
                    print("[BLOK II] Hasil Cek NIK tidak berstatus 'DITEMUKAN'.")
                    raise Exception("Error : NIK salah")
                else:
                    print("[BLOK II] Hasil Cek NIK terdeteksi 'DITEMUKAN'. Lanjut ke langkah selanjutnya...")
            else:
                raise Exception("Tombol 'Cek NIK' tidak ditemukan.")


            # 5. Mengisi "203. Nomor telepon/HP penghuni"
            print("[BLOK II] Mengisi '203. Nomor telepon/HP penghuni'...")
            
            # Scroll statis sedikit agar posisi menu 203 bergeser naik
            try:
                d.swipe(540, 1000, 540, 750, duration=0.15)
                time.sleep(SLEEP_SHORT)
            except Exception as swipe_err:
                print(f"[WARNING] Gagal melakukan scroll statis step 5: {swipe_err}")
                
            label_203 = d(textContains="203. Nomor telepon")
            if not label_203.exists():
                d(scrollable=True).scroll.to(textContains="203. Nomor telepon")
                
            if label_203.exists(timeout=5):
                input_203 = label_203.down(className="android.widget.EditText")
                if input_203.exists():
                    input_203.set_text(str(telp))
                    print(f"[BLOK II] Berhasil mengisi Nomor telepon: '{telp}'")
                    time.sleep(SLEEP_SHORT)
                else:
                    raise Exception("Input text box untuk '203. Nomor telepon/HP penghuni' tidak ditemukan.")
            else:
                raise Exception("Label '203. Nomor telepon/HP penghuni' tidak ditemukan.")

            # Lakukan scroll sekali setelah step 5
            print("[BLOK II] Melakukan scroll ke bawah sekali setelah Step 5...")
            try:
                d(scrollable=True).scroll.forward()
                time.sleep(SLEEP_SHORT)
            except Exception as scroll_err:
                print(f"[WARNING] Gagal scroll setelah step 5: {scroll_err}")

            # 6. Mengisi "204. Status kepemilikan bangunan tempat tinggal" (Centang "1. Milik sendiri")
            print("[BLOK II] Mengisi '204. Status kepemilikan bangunan tempat tinggal'...")
            label_204 = d(textContains="204. Status kepemilikan")
            if not label_204.exists():
                d(scrollable=True).scroll.to(textContains="204. Status kepemilikan")
                
            clicked_204 = False
            try:
                label_el = None
                for pattern in ["1. Milik sendiri", "1.  Milik sendiri", "Milik sendiri"]:
                    elements = d.xpath(f"//*[contains(@text, '{pattern}')]").all()
                    for el in elements:
                        bounds_str = el.attrib.get('bounds')
                        import re
                        pts = [int(x) for x in re.findall(r'\d+', bounds_str)]
                        if len(pts) == 4:
                            x1, y1, x2, y2 = pts
                            if (x2 - x1) > 50:
                                label_el = el
                                break
                    if label_el:
                        break

                if label_el:
                    label_bounds_str = label_el.attrib.get('bounds')
                    import re
                    l_pts = [int(x) for x in re.findall(r'\d+', label_bounds_str)]
                    label_left, label_top, label_right, label_bottom = l_pts
                    
                    label_text = label_el.text
                    parent = d.xpath(f"//*[contains(@text, '{label_text}')]/..")
                    
                    if parent.exists:
                        parent_bounds = parent.all()[0].attrib.get('bounds')
                        p_pts = [int(x) for x in re.findall(r'\d+', parent_bounds)]
                        parent_left, parent_top, parent_right, parent_bottom = p_pts
                        
                        click_x = parent_left + (label_left - parent_left) // 2
                        click_y = label_top + (label_bottom - label_top) // 2
                        
                        print(f"[KLIK] Parent Left: {parent_left}, Label Left: {label_left}")
                        print(f"[KLIK] Mengklik koordinat dinamis radio button 204: ({click_x}, {click_y})")
                        d.click(click_x, click_y)
                        time.sleep(SLEEP_SHORT)
                        clicked_204 = True
            except Exception as click_err:
                print(f"[WARNING] Percobaan klik koordinat dinamis 204 gagal: {click_err}")

            if not clicked_204:
                try:
                    matched_elements = d.xpath("//*[contains(@text, 'Milik sendiri') or contains(@text, 'milik sendiri') or contains(@text, 'Milik Sendiri') or contains(@text, 'Milik')]").all()
                    if matched_elements:
                        for el in matched_elements:
                            print(f"[FALLBACK] Mengklik elemen Milik sendiri: '{el.text}'")
                            el.click()
                            time.sleep(SLEEP_SHORT)
                        clicked_204 = True
                except Exception as xpath_err:
                    print(f"[WARNING] Fallback xpath 204 gagal: {xpath_err}")
                    
            if not clicked_204:
                raise Exception("Opsi radio button '1. Milik sendiri' tidak ditemukan.")


                
            # Transisi ke BLOK III menggunakan tombol BERIKUTNYA secara dinamis
            # Transisi ke BLOK III menggunakan tombol BERIKUTNYA secara dinamis
            print("[BLOK II] Mengetuk tombol 'BERIKUTNYA'...")
            transisi_sukses = False
            
            # Cari tombol BERIKUTNYA BLOK III di layar
            berikutnya_btn = d(textContains="BERIKUTNYA BLOK III")
            if not berikutnya_btn.exists(): berikutnya_btn = d(textContains="BERIKUTNYA")
            
            if berikutnya_btn.exists(timeout=5):
                # Klik menggunakan koordinat tengah secara dinamis
                bounds = berikutnya_btn.info.get('bounds')
                click_x = (bounds['left'] + bounds['right']) // 2
                click_y = (bounds['top'] + bounds['bottom']) // 2
                print(f"[BLOK II] Mengetuk koordinat tombol 'BERIKUTNYA': ({click_x}, {click_y})")
                d.click(click_x, click_y)
                
                # Verifikasi transisi halaman dengan mempolling tulisan 'BLOK III' (exact match)
                print("[BLOK II] Menunggu halaman BLOK III termuat...")
                for poll_attempt in range(30):
                    time.sleep(0.2)
                    try:
                        # Paksa refresh cache uiautomator2 agar melihat perubahan WebView
                        d.dump_hierarchy()
                    except Exception:
                        pass
                    if d(text="BLOK III").exists():
                        print("[BLOK II] Sukses berpindah ke halaman 'BLOK III' via tombol.")
                        transisi_sukses = True
                        time.sleep(0.2) # Jeda tambahan agar UI benar-benar stabil
                        break
            else:
                # Jika tombol tidak ada tapi text BLOK III sudah ada, artinya sudah di halaman tujuan
                if d(text="BLOK III").exists():
                    transisi_sukses = True
                    
            if not transisi_sukses:
                print("[WARNING] Gagal melakukan konfirmasi transisi halaman ke BLOK III via tombol.")

            # === PROSES BLOK III ===
            print("\n--- MEMULAI AUTOMASI BLOK III ---")
            
            # Ambil data dari temp_alamat.txt (kecuali Alamat dari Excel)
            alamat_data = baca_temp_alamat()
            
            provinsi = alamat_data["Provinsi"]
            kabupaten = alamat_data["Kabupaten"]
            kecamatan = alamat_data["Kecamatan"]
            desa = alamat_data["Desa/Kelurahan"]
            alamat_val = sheet.cell(row=row, column=4).value
            if alamat_val is None:
                alamat_val = ""

            # 1. Cek text "BLOK III"
            print("\n[BLOK III] [STEP 1] Memeriksa teks 'BLOK III'...")
            blok3_header = d(textContains="BLOK III")
            if blok3_header.exists(timeout=20):
                print("[BLOK III] [SUKSES] Berada di halaman/bagian 'BLOK III'")
            else:
                print("[BLOK III] [WARNING] Header 'BLOK III' tidak ditemukan di layar saat ini.")

            # Swipe ke atas kemudian kembalikan lagi dan ketuk koordinat statis (996, 588)
            print("[BLOK III] Melakukan force refresh (swipe up & down) dan mengetuk koordinat statis (996, 588)...")
            d.swipe(540, 1200, 540, 600, duration=0.05)
            time.sleep(0.1)
            d.swipe(540, 600, 540, 1200, duration=0.2)
            time.sleep(0.1)
            d.click(996, 588)
            time.sleep(SLEEP_MEDIUM)

            # 2. di label "a. Provinsi" masukkan data Provinsi
            print("\n[BLOK III] [STEP 2] Mengisi 'a. Provinsi'...")
            label_a = d(textContains="Provinsi")
            input_a = None
            for attempt in range(40):
                if label_a.exists():
                    input_a = label_a.down(className="android.widget.EditText")
                    if input_a and input_a.exists():
                        break
                time.sleep(0.5)
                
            if input_a and input_a.exists():
                provinsi_clean = ekstrak_nama_saja(provinsi)
                input_a.set_text(str(provinsi_clean))
                print(f"[BLOK III] Berhasil mengisi Provinsi: '{provinsi_clean}'")
                time.sleep(0.5)
                klik_opsi_dropdown(d, provinsi)
            else:
                raise Exception("Input text box untuk 'a. Provinsi' tidak ditemukan.")

            # 3. di label "b. Kabupaten/Kota" masukkan data Kabupaten
            print("\n[BLOK III] [STEP 3] Mengisi 'b. Kabupaten/Kota'...")
            label_b = d(textContains="Kabupaten/Kota")
            input_b = None
            for attempt in range(10):
                if label_b.exists():
                    input_b = label_b.down(className="android.widget.EditText")
                    if input_b and input_b.exists():
                        break
                time.sleep(0.5)
                
            if input_b and input_b.exists():
                kabupaten_clean = ekstrak_nama_saja(kabupaten)
                input_b.set_text(str(kabupaten_clean))
                print(f"[BLOK III] Berhasil mengisi Kabupaten/Kota: '{kabupaten_clean}'")
                time.sleep(0.5)
                klik_opsi_dropdown(d, kabupaten)
            else:
                raise Exception("Input text box untuk 'b. Kabupaten/Kota' tidak ditemukan.")

            # 4. di label "c. Kecamatan" masukkan data Kecamatan
            print("\n[BLOK III] [STEP 4] Mengisi 'c. Kecamatan'...")
            print("[BLOK III] Melakukan swipe ke bawah agar 'c. Kecamatan' terlihat penuh...")
            d.swipe(540, 1200, 540, 1000, duration=0.2)
            time.sleep(0.5)
                
            label_c = d(textContains="Kecamatan")
            input_c = None
            for attempt in range(10):
                if label_c.exists():
                    input_c = label_c.down(className="android.widget.EditText")
                    if input_c and input_c.exists():
                        break
                time.sleep(0.5)
                
            if input_c and input_c.exists():
                kecamatan_clean = ekstrak_nama_saja(kecamatan)
                input_c.set_text(str(kecamatan_clean))
                print(f"[BLOK III] Berhasil mengisi Kecamatan: '{kecamatan_clean}'")
                time.sleep(0.5)
                klik_opsi_dropdown(d, kecamatan)
            else:
                raise Exception("Input text box untuk 'c. Kecamatan' tidak ditemukan.")

            # 5. di label "d. Desa/Kelurahan" masukkan data Desa/Kelurahan
            print("\n[BLOK III] [STEP 5] Mengisi 'd. Desa/Kelurahan'...")
            print("[BLOK III] Melakukan swipe ke bawah agar 'd. Desa/Kelurahan' terlihat penuh...")
            d.swipe(540, 1200, 540, 600, duration=0.2)
            time.sleep(0.5)
                
            label_d = d(textContains="Desa/Kelurahan")
            input_d = None
            for attempt in range(10):
                if label_d.exists():
                    input_d = label_d.down(className="android.widget.EditText")
                    if input_d and input_d.exists():
                        break
                time.sleep(0.5)
                
            if input_d and input_d.exists():
                desa_clean = ekstrak_nama_saja(desa)
                desa_sukses = False
                max_attempts = max(1, len(desa_clean) - 3 + 1)
                
                for attempt_desa in range(max_attempts):
                    if attempt_desa == 0:
                        current_input = desa_clean
                    else:
                        current_input = desa_clean[:-attempt_desa]
                    
                    # Kosongkan textbox terlebih dahulu sebelum menginput teks baru
                    input_d.clear_text()
                    time.sleep(0.5)

                    input_d.set_text(str(current_input))
                    print(f"[BLOK III] Mengisi Desa/Kelurahan: '{current_input}' (Percobaan {attempt_desa + 1}/{max_attempts})")
                    time.sleep(0.5)
                    
                    if d(textContains="Pilihan tidak ditemukan").exists():
                        print(f"[BLOK III] Terdeteksi 'Pilihan tidak ditemukan' untuk '{current_input}'. Menutup overlay, menghapus input, dan mengulang...")
                        d.click(540, 300)
                        time.sleep(0.5)
                        input_d.clear_text()
                        time.sleep(0.5)
                        continue
                    
                    if klik_opsi_dropdown(d, current_input) or klik_opsi_dropdown(d, desa):
                        print(f"[BLOK III] Berhasil memilih opsi Desa/Kelurahan untuk '{current_input}'.")
                        desa_sukses = True
                        break
                    else:
                        print(f"[BLOK III] Opsi Desa/Kelurahan '{current_input}' tidak berhasil diklik. Menghapus input dan mencoba lagi...")
                        input_d.clear_text()
                        time.sleep(0.5)
                
                if not desa_sukses:
                    raise Exception(f"Error : tidak menemukan Desa/Kelurahan {desa_clean}")
            else:
                raise Exception("Input text box untuk 'd. Desa/Kelurahan' tidak ditemukan.")

            # 6. di label "e. Alamat" masukkan data Alamat
            print("\n[BLOK III] [STEP 6] Mengisi 'e. Alamat'...")
            d.click(540, 300)
            time.sleep(0.5)
            
            print("[BLOK III] Melakukan swipe ke bawah agar 'e. Alamat' terlihat penuh...")
            d.swipe(540, 1200, 540, 650, duration=0.2)
            time.sleep(0.5)
                
            label_alamat = d(textContains="Alamat")
            input_alamat = None
            for attempt in range(10):
                if label_alamat.exists():
                    input_alamat = label_alamat.down(className="android.widget.EditText")
                    if input_alamat and input_alamat.exists():
                        break
                time.sleep(0.5)
                
            if input_alamat and input_alamat.exists():
                input_alamat.set_text(str(alamat_val))
                print(f"[BLOK III] Berhasil mengisi Alamat: '{alamat_val}'")
                time.sleep(0.5)
            else:
                raise Exception("Input text box untuk 'e. Alamat' tidak ditemukan.")

            # 7. ketuk tombol kontrol Increment
            print("\n[BLOK III] [STEP 7] Mengetuk tombol kontrol 'Increment'...")
            
            # Cari tombol Increment dengan toleransi teks
            increment_btn = d(text="Increment")
            if not increment_btn.exists():
                increment_btn = d(textContains="Increment")
                
            if not increment_btn.exists():
                print("[BLOK III] [STEP 7] Melakukan scroll mencari tombol Increment...")
                try:
                    d(scrollable=True).scroll.to(text="Increment")
                except Exception:
                    try:
                        d(scrollable=True).scroll.to(textContains="Increment")
                    except Exception:
                        pass
                        
            # Fallback scroll manual jika scroll.to tidak menemukan/gagal
            if not increment_btn.exists():
                for _ in range(3):
                    d.swipe(540, 1200, 540, 600, duration=0.2)
                    time.sleep(0.5)
                    if d(text="Increment").exists():
                        increment_btn = d(text="Increment")
                        break
                    elif d(textContains="Increment").exists():
                        increment_btn = d(textContains="Increment")
                        break

            if increment_btn.exists():
                increment_btn.click()
                print("[BLOK III] Berhasil mengetuk tombol 'Increment'")
                time.sleep(1)
            else:
                raise Exception("Tombol 'Increment' tidak ditemukan.")

            # Transisi ke BLOK IV menggunakan tombol BERIKUTNYA secara dinamis
            print("[BLOK III] Mengetuk tombol 'BERIKUTNYA'...")
            transisi_sukses_blok4 = False
            
            # Cari tombol BERIKUTNYA BLOK IV di layar
            berikutnya_btn_blok4 = d(textContains="BERIKUTNYA BLOK IV")
            if not berikutnya_btn_blok4.exists(): berikutnya_btn_blok4 = d(textContains="BERIKUTNYA")
            
            if berikutnya_btn_blok4.exists(timeout=5):
                # Klik menggunakan koordinat tengah secara dinamis
                bounds = berikutnya_btn_blok4.info.get('bounds')
                click_x = (bounds['left'] + bounds['right']) // 2
                click_y = (bounds['top'] + bounds['bottom']) // 2
                print(f"[BLOK III] Mengetuk koordinat tombol 'BERIKUTNYA' (BLOK IV): ({click_x}, {click_y})")
                d.click(click_x, click_y)
                
                # Verifikasi transisi halaman dengan mempolling tulisan 'BLOK IV' (exact match)
                print("[BLOK III] Menunggu halaman BLOK IV termuat...")
                for poll_attempt in range(15):
                    time.sleep(1)
                    try:
                        # Paksa refresh cache uiautomator2 agar melihat perubahan WebView
                        d.dump_hierarchy()
                    except Exception:
                        pass
                    if d(text="BLOK IV").exists():
                        print("[BLOK III] Sukses berpindah ke halaman 'BLOK IV' via tombol.")
                        transisi_sukses_blok4 = True
                        time.sleep(1.0) # Jeda tambahan agar UI benar-benar stabil
                        break
            else:
                # Jika tombol tidak ada tapi text BLOK IV sudah ada, artinya sudah di halaman tujuan
                if d(text="BLOK IV").exists():
                    transisi_sukses_blok4 = True
                    
            if not transisi_sukses_blok4:
                print("[WARNING] Gagal melakukan konfirmasi transisi halaman ke BLOK IV via tombol.")

            # === PROSES BLOK IV ===
            print("\n--- MEMULAI AUTOMASI BLOK IV ---")
            
            print("\n[BLOK IV] [STEP 1] Mengisi catatan 'Catatan'...")
            label_catatan = d(textContains="Catatan")
            if not label_catatan.exists():
                print("[BLOK IV] Mencari label Catatan...")
                try:
                    d(scrollable=True).scroll.to(textContains="Catatan")
                except Exception:
                    pass
                    
            if not label_catatan.exists():
                for _ in range(3):
                    d.swipe(540, 1200, 540, 600, duration=0.2)
                    time.sleep(0.1)
                    if d(textContains="Catatan").exists():
                        label_catatan = d(textContains="Catatan")
                        break

            input_catatan = None
            if label_catatan.exists():
                for attempt in range(10):
                    input_catatan = label_catatan.down(className="android.widget.EditText")
                    if input_catatan and input_catatan.exists():
                        break
                    time.sleep(0.1)
                    
            if input_catatan and input_catatan.exists():
                input_catatan.set_text("-")
                print("[BLOK IV] Berhasil mengisi Catatan: '-'")
                time.sleep(0.2)
            else:
                # Alternatif jika down EditText gagal, cari EditText pertama
                print("[BLOK IV] [WARNING] EditText di bawah Catatan tidak ditemukan. Mencoba mengetik di EditText pertama di layar...")
                first_edit = d(className="android.widget.EditText")
                if first_edit.exists():
                    first_edit.set_text("-")
                    print("[BLOK IV] Berhasil mengisi Catatan (via first EditText): '-'")
                    time.sleep(0.2)
                else:
                    raise Exception("Input text box Catatan tidak ditemukan.")

            print("\n[BLOK IV] [STEP 2] Mengetuk tombol 'Ambil Waktu'...")
            ambil_waktu_btn = d(text="Ambil Waktu")
            if not ambil_waktu_btn.exists():
                try:
                    d(scrollable=True).scroll.to(text="Ambil Waktu")
                except Exception:
                    pass
                    
            if ambil_waktu_btn.exists(timeout=5):
                ambil_waktu_btn.click()
                print("[BLOK IV] Berhasil mengetuk 'Ambil Waktu'")
                time.sleep(0.2)
            else:
                raise Exception("Tombol 'Ambil Waktu' tidak ditemukan.")

            print("\n[BLOK IV] [STEP 3] Mengetuk konfirmasi 'Ya'...")
            ya_btn = d(text="Ya")
            if not ya_btn.exists(): ya_btn = d(text="YA")
            if not ya_btn.exists(): ya_btn = d(text="ya")
            if ya_btn.exists(timeout=5):
                ya_btn.click()
                print("[BLOK IV] Berhasil mengetuk 'Ya'")
                time.sleep(0.2)
            else:
                print("[BLOK IV] [WARNING] Konfirmasi dialog 'Ya' tidak muncul.")

            # Perulangan submit dari STEP 4 jika tombol 'Kirim' masih muncul setelah STEP 7
            for send_attempt in range(1, 6):
                print(f"\n[BLOK IV] [STEP 4] Mengetuk tombol 'Kirim' pertama (Percobaan {send_attempt})...")
                kirim_btn = d(text="Kirim")
                if not kirim_btn.exists(): kirim_btn = d(text="KIRIM")
                if not kirim_btn.exists(): kirim_btn = d(textContains="Kirim")
                
                if not kirim_btn.exists():
                    try:
                        d(scrollable=True).scroll.to(text="Kirim")
                    except Exception:
                        pass
                        
                if kirim_btn.exists(timeout=5):
                    kirim_btn.click()
                    print("[BLOK IV] Berhasil mengetuk tombol 'Kirim' pertama")
                    time.sleep(0.5)
                else:
                    raise Exception("Tombol 'Kirim' pertama tidak ditemukan.")

                print("\n[BLOK IV] [STEP 5] Mengetuk tombol 'Kirim' kedua...")
                # Polling tunggu teks "GALAT 0 Perlu diperbaiki" muncul (maksimal 5 detik)
                for attempt in range(10):
                    if d(textContains="GALAT 0 Perlu diperbaiki").exists() or d(textContains="GALAT 0").exists():
                        print("[BLOK IV] [STEP 5] Menemukan teks 'GALAT 0 Perlu diperbaiki'!")
                        break
                    time.sleep(0.1)
                    
                kirim_btn2 = d(text="Kirim", className="android.widget.Button")
                if not kirim_btn2.exists(): kirim_btn2 = d(text="KIRIM", className="android.widget.Button")
                if not kirim_btn2.exists(): kirim_btn2 = d(textContains="Kirim", className="android.widget.Button")
                
                if kirim_btn2.exists(timeout=5):
                    kirim_btn2.click()
                    print("[BLOK IV] Berhasil mengetuk tombol 'Kirim' kedua")
                    time.sleep(0.2)
                    # Jeda loading setelah kirim
                    time.sleep(0.2)
                    # Tunggu loading/progress dialog selesai jika ada
                    print("[BLOK IV] Menunggu loading selesai...")
                    try:
                        d(resourceId="id.go.bpsfasih:id/card_progress").wait_gone(timeout=20)
                    except Exception:
                        pass
                    time.sleep(0.5)  # Jeda pengaman agar dialog konfirmasi submit muncul dengan stabil
                else:
                    raise Exception("Tombol 'Kirim' kedua tidak ditemukan/tidak muncul.")

                print("\n[BLOK IV] [STEP 6] Mengetuk tombol 'Konfirmasi'...")
                konfirmasi_btn = d(text="Konfirmasi", className="android.widget.Button")
                if not konfirmasi_btn.exists():
                    konfirmasi_btn = d(textContains="Konfirmasi", className="android.widget.Button")
                    
                if konfirmasi_btn.exists(timeout=5):
                    konfirmasi_btn.click()
                    print("[BLOK IV] Berhasil mengetuk tombol 'Konfirmasi'")
                    time.sleep(0.2)
                    # Tunggu loading/progress dialog selesai
                    print("[BLOK IV] Menunggu loading selesai...")
                    try:
                        d(resourceId="id.go.bpsfasih:id/card_progress").wait_gone(timeout=20)
                    except Exception:
                        pass
                    time.sleep(0.5)  # Jeda pengaman agar dialog konfirmasi submit muncul dengan stabil
                else:
                    raise Exception("Tombol 'Konfirmasi' tidak ditemukan.")

                print("\n[BLOK IV] [STEP 7] Mengetuk tombol 'YA'...")
                ya_submit_btn = d(resourceId="id.go.bpsfasih:id/rButton_bottomDialog")
                if not ya_submit_btn.exists():
                    ya_submit_btn = d(text="YA", className="android.widget.Button")
                if not ya_submit_btn.exists():
                    ya_submit_btn = d(textContains="YA")

                if ya_submit_btn.exists(timeout=5):
                    ya_submit_btn.click()
                    print("[BLOK IV] Berhasil mengetuk tombol 'YA' Submit")
                    time.sleep(0.5)
                else:
                    raise Exception("Tombol 'YA' Submit tidak ditemukan.")

                # Scan apakah tombol "Kirim" masih ada setelah mengetuk tombol 'YA' Submit
                time.sleep(1.0)
                kirim_check = d(text="Kirim")
                if not kirim_check.exists(): kirim_check = d(text="KIRIM")
                if not kirim_check.exists(): kirim_check = d(textContains="Kirim")

                if kirim_check.exists():
                    print("[BLOK IV] [WARNING] Tombol 'Kirim' masih muncul setelah mengetuk 'YA' Submit. Mengulangi dari STEP 4...")
                    time.sleep(1.0)
                else:
                    print("[BLOK IV] Tombol 'Kirim' sudah tidak terdeteksi. Lanjut ke STEP 8...")
                    break

            #25 ketuk "OK" pada modal Submit diproses (retry 5x & fallback statis bounds 528, 1691)
            ketuk_ok_submit_diproses(max_attempts=5)

            #25 Scan teks "Halaman Upload" -> jika muncul maka tekan tombol "BACK" pada emulator
            print("[EMULATOR] Memeriksa apakah teks 'Halaman Upload' sudah muncul di layar...")
            is_halaman_upload = False
            for _ in range(5):
                if (d(textContains="Halaman Upload").exists or 
                    d(descriptionContains="Halaman Upload").exists or 
                    d.xpath("//*[contains(@text, 'Halaman Upload') or contains(@content-desc, 'Halaman Upload')]").exists):
                    is_halaman_upload = True
                    break
                time.sleep(0.5)

            if is_halaman_upload:
                print("[EMULATOR] Teks 'Halaman Upload' terdeteksi di layar! Menekan tombol Back pada emulator...")
                d.press("back")
                time.sleep(SLEEP_LONG)
            else:
                print("[EMULATOR] Teks 'Halaman Upload' tidak terdeteksi di layar.")

            print("\n[BLOK IV] [STEP 8] Menunggu loading submit selesai...")
            try:
                # Tunggu loading progress bar selesai jika muncul
                d(resourceId="id.go.bpsfasih:id/card_progress").wait_gone(timeout=30)
            except Exception:
                pass
                
            print("[BLOK IV] Menunggu halaman 'Daftar Assignment' termuat...")
            daftar_assignment_title = d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Daftar Assignment")
            if daftar_assignment_title.wait(exists=True, timeout=30):
                print("[BLOK IV] [SUKSES] Berhasil kembali ke halaman 'Daftar Assignment'")
            else:
                print("[BLOK IV] [WARNING] Halaman 'Daftar Assignment' tidak terdeteksi setelah 30 detik.")

            # === JEDA UNTUK VERIFIKASI MANUAL ===
            print("\n==================================================")
            # Tulis Status Sukses ke Excel
            sheet.cell(row=row, column=9).value = "SUKSES"
            sheet.cell(row=row, column=10).value = time.strftime("%Y-%m-%d %H:%M:%S")
            print(f"[STATUS] IDPEL {idpel} berhasil diproses.")

            

        except Exception as err:
            print(f"[FAILED] Baris {row} gagal diproses. Error: {err}")
            err_str = str(err)
            if "Error : SUDAH TERCATAT PADA SISTEM FASIH" in err_str:
                sheet.cell(row=row, column=9).value = "Error : SUDAH TERCATAT PADA SISTEM FASIH"
            elif err_str.startswith("Error :"):
                sheet.cell(row=row, column=9).value = err_str
            else:
                sheet.cell(row=row, column=9).value = f"GAGAL: {err_str}"
            sheet.cell(row=row, column=10).value = time.strftime("%Y-%m-%d %H:%M:%S")

            # --- PROSEDUR RECOVERY / RESET JIKA GAGAL ---
            print(f"[FAILED] Memulai pemulihan untuk Baris {row}...")
            dialog_found = False
            
            # Cek jika kita sudah di Dashboard Daftar Assignment terlebih dahulu
            if d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Daftar Assignment").exists():
                print("[FAILED] Sudah berada di Dashboard Daftar Assignment. Recovery dilewati.")
                dialog_found = True
                
            if not dialog_found:
                for i in range(5):
                    print(f"[FAILED] Menekan tombol BACK (Percobaan {i+1}/5)...")
                    if press_back_and_check_periode():
                        dialog_found = True
                        break
                    
                    # Cari dialog keluar konfirmasi
                    dialog_el = d(
                        resourceId="id.go.bpsfasih:id/deskripsi_bottomDialog",
                        text="Apakah Anda yakin akan keluar dari halaman ini ?"
                    )
                    
                    # Pengecekan dengan timeout 3 detik
                    if dialog_el.wait(exists=True, timeout=3.0):
                        print("[FAILED] Dialog konfirmasi keluar ditemukan!")
                        
                        # Ketuk tombol IYA
                        iya_btn = d(
                            resourceId="id.go.bpsfasih:id/rButton_bottomDialog",
                            text="IYA"
                        )
                        if iya_btn.exists(timeout=2):
                            iya_btn.click()
                            print("[FAILED] Tombol 'IYA' diketuk. Kembali ke Dashboard.")
                            dialog_found = True
                            time.sleep(SLEEP_LONG) # Beri waktu agar kembali ke Dashboard sepenuhnya
                            break
                        else:
                            print("[WARNING] Tombol 'IYA' tidak ditemukan.")
                    else:
                        # Cek juga jika kita sudah di Dashboard Daftar Assignment
                        if d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Daftar Assignment").exists():
                            print("[FAILED] Sudah berada di Dashboard Daftar Assignment.")
                            dialog_found = True
                            break
                        print("[FAILED] Dialog belum muncul, mencoba BACK kembali...")
            
            if not dialog_found:
                print("[FAILED] Gagal melakukan recovery (timeout). Menghentikan eksekusi script untuk mencegah kegagalan beruntun.")
                # Simpan Excel sebelum keluar
                try:
                    wb.save(EXCEL_FILE)
                except Exception:
                    pass
                tampilkan_ringkasan_akhir(sheet)
                sys.exit(1)

        finally:
            # Simpan Excel per baris demi keselamatan data dengan penanganan lock file
            while True:
                try:
                    wb.save(EXCEL_FILE)
                    break
                except PermissionError:
                    print(f"\n[WARNING] File '{EXCEL_FILE}' sedang dibuka di program lain (misal Microsoft Excel).")
                    print("Mohon segera TUTUP file tersebut agar bot dapat menyimpan status terbaru.")
                    print("Menunggu 5 detik sebelum mencoba menyimpan kembali...")
                    time.sleep(5)

    tampilkan_ringkasan_akhir(sheet)

    print("\n==================================================")
    print("PROSES BOT SCAPPER FASIH SELESAI!")
    print("==================================================")

if __name__ == "__main__":
    main()
