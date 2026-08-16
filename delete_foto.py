import os
import openpyxl

# ==========================================
# KONFIGURASI
# ==========================================
FOTO_DIRECTORY = r"C:\Users\BaliAga\Documents\XuanZhi9\Pictures"
EXCEL_FILES = ["data_tugas.xlsx", "data_tugas2.xlsx"]

def dapatkan_semua_idpel_dari_excel(excel_list):
    """Membaca semua IDPEL dari daftar file Excel (Kolom A)"""
    idpel_set = set()
    for excel_path in excel_list:
        if not os.path.exists(excel_path):
            print(f"[WARNING] File Excel '{excel_path}' tidak ditemukan, dilewati.")
            continue

        print(f"[EXCEL] Membaca IDPEL dari '{excel_path}'...")
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = wb.active
            count_local = 0
            for r in range(2, sheet.max_row + 1):
                val = sheet.cell(row=r, column=1).value
                if val is not None and str(val).strip() != "":
                    if isinstance(val, float):
                        val_str = str(int(val)).strip()
                    else:
                        val_str = str(val).strip()
                    if val_str:
                        idpel_set.add(val_str)
                        count_local += 1
            print(f"[EXCEL] Berhasil membaca {count_local} IDPEL dari '{excel_path}'.")
        except Exception as e:
            print(f"[ERROR] Gagal membaca Excel '{excel_path}': {e}")
            
    return idpel_set

def main():
    print("==================================================")
    print("PROGRAM PEMBERSIH FOTO (DELETE FOTO UNUSED)")
    print("==================================================")
    print(f"Direktori Target : {FOTO_DIRECTORY}")
    print(f"File Excel Target: {', '.join(EXCEL_FILES)}")
    print("==================================================\n")

    if not os.path.exists(FOTO_DIRECTORY):
        print(f"[ERROR] Folder foto '{FOTO_DIRECTORY}' tidak ditemukan!")
        return

    # 1. Kumpulkan semua IDPEL yang sah dari kedua file Excel
    idpel_valid = dapatkan_semua_idpel_dari_excel(EXCEL_FILES)
    print(f"\n[TOTAL] Total IDPEL unik dari seluruh Excel: {len(idpel_valid)} IDPEL\n")

    # 2. Pindai folder foto dan hapus foto yang tidak ada di Excel
    file_list = os.listdir(FOTO_DIRECTORY)
    total_files = len(file_list)
    deleted_count = 0
    kept_count = 0

    print(f"[SCAN] Memindai {total_files} file di '{FOTO_DIRECTORY}'...\n")

    for filename in file_list:
        file_path = os.path.join(FOTO_DIRECTORY, filename)

        # Hanya proses jika berupa file (bukan subfolder)
        if not os.path.isfile(file_path):
            continue

        # Ambil nama file tanpa ekstensi (misal: "51030612345.jpg" -> "51030612345")
        nama_tanpa_ext, ext = os.path.splitext(filename)

        # Periksa apakah IDPEL ada di set idpel_valid
        if nama_tanpa_ext in idpel_valid:
            kept_count += 1
        else:
            try:
                os.remove(file_path)
                deleted_count += 1
                print(f"[DELETE] Menghapus: {filename} (Tidak ada di Excel)")
            except Exception as e:
                print(f"[ERROR] Gagal menghapus '{filename}': {e}")

    print("\n==================================================")
    print("HASIL PEMBERSIHAN FOTO:")
    print(f"  - Total file diperiksa : {total_files}")
    print(f"  - File foto Dihapus    : {deleted_count}")
    print(f"  - File foto Dipertahankan : {kept_count}")
    print("==================================================")

if __name__ == "__main__":
    main()
