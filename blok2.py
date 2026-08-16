import uiautomator2 as u2
import openpyxl
import subprocess
import os
import time

LDPLAYER_ADB = r"C:\LDPlayer\LDPlayer9\adb.exe"
EXCEL_FILE = "data_tugas.xlsx"

# Jeda Waktu (Sleep Durations)
SLEEP_SHORT = 0.2   # Untuk ketik text, ENTER, scroll, radio button
SLEEP_MEDIUM = 0.5  # Untuk klik tombol normal, aksi loading cepat
SLEEP_LONG = 1.0    # Untuk transisi halaman, pemicu kamera/galeri, ambil GPS

def hubungkan_adb():
    print("Mencoba menghubungkan ADB ke emulator...")
    for port in ["5554", "5555"]:
        if os.path.exists(LDPLAYER_ADB):
            subprocess.run([LDPLAYER_ADB, "connect", f"127.0.0.1:{port}"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        
        try:
            d = u2.connect(f"127.0.0.1:{port}")
            device_info = d.info
            print(f"[SUKSES] Terhubung ke emulator di port {port}!")
            return d
        except Exception:
            continue
    print("[GAGAL] Tidak dapat terhubung ke emulator.")
    return None

def main():
    d = hubungkan_adb()
    if not d:
        return

    # Tentukan baris yang ingin ditarget untuk pengetesan (baris 2 untuk 1 idpel saja)
    row = 2

    # Baca data dari data_tugas.xlsx
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] File Excel '{EXCEL_FILE}' tidak ditemukan!")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        
        nama_penghuni = sheet.cell(row=row, column=3).value  # Baris 2, Kolom C (3)
        nik = sheet.cell(row=row, column=7).value            # Baris 2, Kolom G (7)
        telp = sheet.cell(row=row, column=8).value           # Baris 2, Kolom H (8)
        
        if nik is not None:
            if isinstance(nik, float):
                nik = str(int(nik)).strip()
            else:
                nik = str(nik).strip()
        if not nik or len(nik) < 15:
            nik = "9999999999999998"
            
        # Bersihkan nomor telepon jika berupa float/int
        if telp is not None:
            if isinstance(telp, float):
                telp = str(int(telp)).strip()
            else:
                telp = str(telp).strip()
        if not telp or telp == '0' or telp == "":
            telp = "-"
            
        print(f"[EXCEL] Baris {row} | Nama Penghuni: '{nama_penghuni}' | NIK: '{nik}' | Telp: '{telp}'")
    except Exception as e:
        print(f"[ERROR] Gagal membaca Excel: {e}")
        return

    # === TRANSISI KE HALAMAN BLOK II (KLIK TOMBOL BERIKUTNYA) ===
    print("\n--- TRANSISI KE BLOK II VIA TOMBOL BERIKUTNYA ---")
    transisi_sukses = False
    
    # Cari tombol BERIKUTNYA BLOK II di layar
    berikutnya_btn = d(textContains="BERIKUTNYA BLOK II")
    if not berikutnya_btn.exists(): berikutnya_btn = d(textContains="BERIKUTNYA")
    
    if berikutnya_btn.exists(timeout=5):
        # Klik menggunakan koordinat tengah secara dinamis
        bounds = berikutnya_btn.info.get('bounds')
        click_x = (bounds['left'] + bounds['right']) // 2
        click_y = (bounds['top'] + bounds['bottom']) // 2
        print(f"[NAVIGASI] Mengetuk koordinat tombol 'BERIKUTNYA': ({click_x}, {click_y})")
        d.click(click_x, click_y)
        
        # Verifikasi transisi halaman dengan mempolling field '201. Nama penghuni'
        print("[NAVIGASI] Menunggu halaman BLOK II termuat...")
        for poll_attempt in range(15):
            time.sleep(1)
            try:
                # Paksa refresh cache uiautomator2 agar melihat perubahan WebView
                d.dump_hierarchy()
            except Exception:
                pass
            if d(textContains="201. Nama penghuni").exists():
                print("[NAVIGASI] Sukses berpindah ke halaman 'BLOK II' via tombol.")
                transisi_sukses = True
                time.sleep(1.0) # Jeda tambahan agar UI benar-benar stabil sebelum proses selanjutnya
                break
    else:
        # Jika tombol tidak ada tapi text 201 sudah ada, artinya sudah di halaman tujuan
        if d(textContains="201. Nama penghuni").exists():
            transisi_sukses = True
            
    if not transisi_sukses:
        print("[WARNING] Gagal melakukan konfirmasi transisi halaman ke BLOK II via tombol.")

    # ==========================================
    # PENGISIAN BLOK II (KETERANGAN PERUMAHAN)
    # ==========================================
    print("\n[BLOK II] Memulai pemrosesan BLOK II...")
    
    # 1. Cek text "BLOK II" (Verifikasi sederhana tanpa scroll)
    blok2_header = d(textContains="BLOK II")
    if blok2_header.exists(timeout=5):
        print("[BLOK II] Berhasil berada di bagian 'BLOK II'.")
    else:
        print("[WARNING] Header 'BLOK II' tidak terdeteksi, melanjutkan pengisian...")

    # 2. Mengisi "201. Nama penghuni" menggunakan set_text pada EditText pertama secara langsung
    print("[BLOK II] Mengisi '201. Nama penghuni'...")
    try:
        first_edit_text = d(className="android.widget.EditText")
        if first_edit_text.exists(timeout=5):
            first_edit_text.set_text(str(nama_penghuni))
            print(f"[BLOK II] Berhasil mengisi Nama Penghuni: '{nama_penghuni}'")
            time.sleep(SLEEP_SHORT)
        else:
            raise Exception("Input text box '201' tidak ditemukan.")
    except Exception as text_err:
        raise Exception(f"Gagal mengisi Nama Penghuni: {text_err}")

    # 3. Mengisi "202. NIK penghuni"
    print("[BLOK II] Mengisi '202. NIK penghuni'...")
    label_202 = d(textContains="202. NIK penghuni")
    if not label_202.exists():
        d(scrollable=True).scroll.to(textContains="202. NIK penghuni")
        
    if label_202.exists(timeout=5):
        input_202 = label_202.down(className="android.widget.EditText")
        if input_202.exists():
            input_202.set_text(str(nik))
            print(f"[BLOK II] Berhasil mengisi NIK Penghuni: '{nik}'")
            time.sleep(SLEEP_SHORT)
        else:
            raise Exception("Input text box untuk '202. NIK penghuni' tidak ditemukan.")
    else:
        raise Exception("Label '202. NIK penghuni' tidak ditemukan.")

    # 4. Ketuk "Cek NIK"
    print("[BLOK II] Mengetuk tombol 'Cek NIK'...")
    cek_nik_btn = d(text="Cek NIK")
    if not cek_nik_btn.exists():
        d(scrollable=True).scroll.to(text="Cek NIK")
        
    if cek_nik_btn.exists(timeout=5):
        # Klik koordinat tengah agar aman
        bounds = cek_nik_btn.info.get('bounds')
        click_x = (bounds['left'] + bounds['right']) // 2
        click_y = (bounds['top'] + bounds['bottom']) // 2
        d.click(click_x, click_y)
        print("[BLOK II] Tombol 'Cek NIK' diketuk. Menunggu respon Cek NIK selesai...")
        time.sleep(1)
        d(resourceId="id.go.bpsfasih:id/card_progress").wait_gone(timeout=20)
        time.sleep(SLEEP_MEDIUM)
    else:
        raise Exception("Tombol 'Cek NIK' tidak ditemukan.")

    # 5. Mengisi "202a. NIK penghuni (NIK versi lama)"
    nik_tidak_ditemukan = d(textContains="TIDAK DITEMUKAN").exists()
    if nik_tidak_ditemukan:
        print("[BLOK II] NIK berstatus 'TIDAK DITEMUKAN'. Mengisi '202a. NIK penghuni (NIK versi lama)'...")
        label_202a = d(textContains="202a. NIK penghuni")
        if not label_202a.exists():
            d(scrollable=True).scroll.to(textContains="202a. NIK penghuni")
            
        if label_202a.exists(timeout=5):
            input_202a = label_202a.down(className="android.widget.EditText")
            if input_202a.exists():
                input_202a.set_text(str(nik))
                print(f"[BLOK II] Berhasil mengisi NIK lama: '{nik}'")
                time.sleep(SLEEP_SHORT)
                
                # Scroll sampai mentok ke bawah
                print("[BLOK II] Men-scroll ke bawah sampai mentok...")
                try:
                    for _ in range(3):
                        d.swipe(540, 1200, 540, 300, duration=0.2)
                        time.sleep(0.2)
                except Exception as scroll_err:
                    print(f"[WARNING] Gagal scroll setelah isi 202a: {scroll_err}")
            else:
                raise Exception("Input text box untuk '202a. NIK penghuni (NIK versi lama)' tidak ditemukan.")
        else:
            raise Exception("Label '202a. NIK penghuni (NIK versi lama)' tidak ditemukan.")
    else:
        print("[BLOK II] NIK berstatus 'DITEMUKAN' atau tidak ada pesan 'TIDAK DITEMUKAN'. Melewati pengisian '202a. NIK penghuni (NIK versi lama)'.")

    # 6. Mengisi "203. Nomor telepon/HP penghuni"
    print("[BLOK II] Mengisi '203. Nomor telepon/HP penghuni'...")
    
    # Scroll statis sedikit agar posisi menu 203 bergeser naik
    try:
        d.swipe(540, 1000, 540, 750, duration=0.15)
        time.sleep(SLEEP_SHORT)
    except Exception as swipe_err:
        print(f"[WARNING] Gagal melakukan scroll statis step 6: {swipe_err}")
        
    label_203 = d(textContains="203. Nomor telepon")
    if not label_203.exists():
        d(scrollable=True).scroll.to(textContains="203. Nomor telepon")
        
    if label_203.exists(timeout=5):
        input_203 = label_203.down(className="android.widget.EditText")
        if input_203.exists():
            input_203.set_text(str(telp))
            print(f"[BLOK II] Berhasil mengisi Nomor telepon: '{telp}'")
            time.sleep(SLEEP_SHORT)
        else:
            raise Exception("Input text box untuk '203. Nomor telepon/HP penghuni' tidak ditemukan.")
    else:
        raise Exception("Label '203. Nomor telepon/HP penghuni' tidak ditemukan.")

    # Lakukan scroll sekali setelah step 6
    print("[BLOK II] Melakukan scroll ke bawah sekali setelah Step 6...")
    try:
        d(scrollable=True).scroll.forward()
        time.sleep(SLEEP_SHORT)
    except Exception as scroll_err:
        print(f"[WARNING] Gagal scroll setelah step 6: {scroll_err}")

    # 7. Mengisi "204. Status kepemilikan bangunan tempat tinggal" (Centang "1. Milik sendiri")
    print("[BLOK II] Mengisi '204. Status kepemilikan bangunan tempat tinggal'...")
    label_204 = d(textContains="204. Status kepemilikan")
    if not label_204.exists():
        d(scrollable=True).scroll.to(textContains="204. Status kepemilikan")
        
    clicked_204 = False
    try:
        label_el = None
        for pattern in ["1. Milik sendiri", "1.  Milik sendiri", "Milik sendiri"]:
            elements = d.xpath(f"//*[contains(@text, '{pattern}')]").all()
            for el in elements:
                bounds_str = el.attrib.get('bounds')
                import re
                pts = [int(x) for x in re.findall(r'\d+', bounds_str)]
                if len(pts) == 4:
                    x1, y1, x2, y2 = pts
                    if (x2 - x1) > 50:
                        label_el = el
                        break
            if label_el:
                break

        if label_el:
            label_bounds_str = label_el.attrib.get('bounds')
            import re
            l_pts = [int(x) for x in re.findall(r'\d+', label_bounds_str)]
            label_left, label_top, label_right, label_bottom = l_pts
            
            label_text = label_el.text
            parent = d.xpath(f"//*[contains(@text, '{label_text}')]/..")
            
            if parent.exists:
                parent_bounds = parent.all()[0].attrib.get('bounds')
                p_pts = [int(x) for x in re.findall(r'\d+', parent_bounds)]
                parent_left, parent_top, parent_right, parent_bottom = p_pts
                
                click_x = parent_left + (label_left - parent_left) // 2
                click_y = label_top + (label_bottom - label_top) // 2
                
                print(f"[KLIK] Parent Left: {parent_left}, Label Left: {label_left}")
                print(f"[KLIK] Mengklik koordinat dinamis radio button 204: ({click_x}, {click_y})")
                d.click(click_x, click_y)
                time.sleep(SLEEP_SHORT)
                clicked_204 = True
    except Exception as click_err:
        print(f"[WARNING] Percobaan klik koordinat dinamis 204 gagal: {click_err}")

    if not clicked_204:
        try:
            matched_elements = d.xpath("//*[contains(@text, 'Milik sendiri') or contains(@text, 'milik sendiri')]").all()
            if matched_elements:
                for el in matched_elements:
                    print(f"[FALLBACK] Mengklik elemen Milik sendiri: '{el.text}'")
                    el.click()
                    time.sleep(SLEEP_SHORT)
                clicked_204 = True
        except Exception as xpath_err:
            print(f"[WARNING] Fallback xpath 204 gagal: {xpath_err}")
            
    if not clicked_204:
        print("[ERROR] Opsi radio button '1. Milik sendiri' tidak ditemukan.")
    else:
        print("[BLOK II] Berhasil memilih opsi '1. Milik sendiri'")

    # 8. Transisi ke BLOK III menggunakan tombol BERIKUTNYA secara dinamis
    print("\n[BLOK II] Mengetuk tombol 'BERIKUTNYA'...")
    transisi_sukses = False
    
    # Cari tombol BERIKUTNYA BLOK III di layar
    berikutnya_btn = d(textContains="BERIKUTNYA BLOK III")
    if not berikutnya_btn.exists(): berikutnya_btn = d(textContains="BERIKUTNYA")
    
    if berikutnya_btn.exists(timeout=5):
        # Klik menggunakan koordinat tengah secara dinamis
        bounds = berikutnya_btn.info.get('bounds')
        click_x = (bounds['left'] + bounds['right']) // 2
        click_y = (bounds['top'] + bounds['bottom']) // 2
        print(f"[BLOK II] Mengetuk koordinat tombol 'BERIKUTNYA': ({click_x}, {click_y})")
        d.click(click_x, click_y)
        
        # Verifikasi transisi halaman dengan mempolling tulisan 'BLOK III' (exact match)
        print("[BLOK II] Menunggu halaman BLOK III termuat...")
        for poll_attempt in range(15):
            time.sleep(1)
            try:
                # Paksa refresh cache uiautomator2 agar melihat perubahan WebView
                d.dump_hierarchy()
            except Exception:
                pass
            if d(text="BLOK III").exists():
                print("[BLOK II] Sukses berpindah ke halaman 'BLOK III' via tombol.")
                transisi_sukses = True
                time.sleep(1.0) # Jeda tambahan agar UI benar-benar stabil
                break
    else:
        # Jika tombol tidak ada tapi text BLOK III sudah ada, artinya sudah di halaman tujuan
        if d(text="BLOK III").exists():
            transisi_sukses = True
            
    if not transisi_sukses:
        print("[WARNING] Gagal melakukan konfirmasi transisi halaman ke BLOK III via tombol.")

    print("\n==================================================")
    print("PROSES PENGETESAN BLOK II SELESAI!")
    print("==================================================")

if __name__ == "__main__":
    main()
