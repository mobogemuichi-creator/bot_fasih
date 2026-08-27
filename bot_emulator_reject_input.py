import os
import time
import subprocess
import openpyxl
import uiautomator2 as u2
import re
import sys
from datetime import datetime

from konfigurasi import (
    LDPLAYER_ADB,
    EXCEL_FILE_REJECT as EXCEL_FILE,
    EMULATOR_PORTS_1 as EMULATOR_PORTS,
    SLEEP_SHORT,
    SLEEP_MEDIUM,
    SLEEP_LONG_REJECT as SLEEP_LONG,
)

d = None

def hubungkan_emulator():
    """Menghubungkan ke emulator via uiautomator2 secara cepat"""
    global d
    print("[KONEKSI] Mencoba menghubungkan ke emulator LDPlayer...")
    
    for port in EMULATOR_PORTS:
        try:
            if os.path.exists(LDPLAYER_ADB):
                subprocess.run([LDPLAYER_ADB, "connect", f"127.0.0.1:{port}"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=2)
            temp_d = u2.connect(f"127.0.0.1:{port}")
            info = temp_d.info
            d = temp_d
            print(f"[KONEKSI] Berhasil terhubung ke emulator di port {port}!")
            print(f"[EMULATOR] Layar: {info.get('displayWidth')}x{info.get('displayHeight')} ({info.get('productName', 'Android')})")
            return True
        except Exception:
            continue

    # Fallback default connect (jika ADB sudah auto detect)
    try:
        temp_d = u2.connect()
        info = temp_d.info
        d = temp_d
        print("[KONEKSI] Berhasil terhubung ke emulator via default connection!")
        return True
    except Exception:
        pass
            
    print("[ERROR] Gagal terhubung ke emulator. Pastikan LDPlayer sudah berjalan.")
    return False


def check_exists(el):
    """Fungsi helper universal untuk mengecek keberadaan elemen Selector maupun XPath uiautomator2"""
    if el is None:
        return False
    try:
        if hasattr(el, 'exists'):
            if callable(el.exists):
                return el.exists()
            return bool(el.exists)
    except Exception:
        pass
    return False


def baca_data_reject(file_path=EXCEL_FILE):
    """Membaca data IDPEL, NIK, NAMA, dan NO TELP dari file Excel data_reject.xlsx, melewati baris yang sudah 'SUKSES DIUPDATE'"""
    if not os.path.exists(file_path):
        print(f"[ERROR] File Excel '{file_path}' tidak ditemukan.")
        return []

    print(f"[EXCEL] Membaca data dari '{file_path}'...")
    data_list = []
    skipped_count = 0
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Baca header
        headers = [cell.value for cell in ws[1]]
        
        # Cari indeks kolom IDPEL, NOMETER_BARU, NIK, NAMA, NO TELP, STATUS
        idpel_idx = -1
        nometer_idx = -1
        nik_idx = -1
        nama_idx = -1
        no_telp_idx = -1
        status_idx = -1

        for idx, h in enumerate(headers):
            if h is None:
                continue
            h_str = str(h).strip().upper()
            if h_str == "IDPEL":
                idpel_idx = idx
            elif h_str in ["NOMETER_BARU", "NOMETER", "NO METER", "NO_METER"]:
                nometer_idx = idx
            elif h_str == "NIK":
                nik_idx = idx
            elif h_str == "NAMA":
                nama_idx = idx
            elif h_str in ["NO TELP", "NO TELP/HP", "TELP", "TELEPON"]:
                no_telp_idx = idx
            elif h_str == "STATUS":
                status_idx = idx

        if idpel_idx == -1:
            print("[ERROR] Kolom 'IDPEL' tidak ditemukan di header Excel.")
            return []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            idpel_val = row[idpel_idx] if idpel_idx < len(row) else None
            nometer_val = row[nometer_idx] if (nometer_idx != -1 and nometer_idx < len(row)) else None
            nik_val = row[nik_idx] if (nik_idx != -1 and nik_idx < len(row)) else None
            nama_val = row[nama_idx] if (nama_idx != -1 and nama_idx < len(row)) else None
            no_telp_val = row[no_telp_idx] if (no_telp_idx != -1 and no_telp_idx < len(row)) else None
            status_val = row[status_idx] if (status_idx != -1 and status_idx < len(row)) else None

            # Skip baris yang statusnya sudah memuat 'SUKSES', 'SUBMIT', 'TIDAK DITEMUKAN', 'ERROR', atau 'KOORDINAT & FOTO TIDAK ADA' (kecuali 'ALAMAT TIDAK DITEMUKAN')
            if status_val is not None:
                status_str = str(status_val).strip().upper()
                if "ALAMAT TIDAK DITEMUKAN" not in status_str:
                    if any(x in status_str for x in ["SUKSES", "SUBMIT", "TIDAK DITEMUKAN", "ERROR", "KOORDINAT & FOTO TIDAK ADA", "KOORDINAT"]):
                        skipped_count += 1
                        continue

            if idpel_val is not None:
                idpel_str = str(idpel_val).strip()
                if idpel_str and idpel_str.lower() != "none":
                    nometer_str = str(nometer_val).strip() if nometer_val is not None else ""
                    nik_str = str(nik_val).strip() if nik_val is not None else ""
                    nama_str = str(nama_val).strip() if nama_val is not None else ""
                    no_telp_str = str(no_telp_val).strip() if no_telp_val is not None else "-"
                    data_list.append({
                        "row": row_idx,
                        "idpel": idpel_str,
                        "nometer": nometer_str,
                        "nik": nik_str,
                        "nama": nama_str,
                        "no_telp": no_telp_str
                    })

        print(f"[EXCEL] Berhasil membaca {len(data_list)} entri data yang perlu diproses ({skipped_count} data berstatus 'SUKSES DIUPDATE' dilewati).")
    except Exception as e:
        print(f"[ERROR] Gagal membaca file Excel: {e}")

    return data_list


def simpan_status_excel(row_num, status_text="SUKSES DIUPDATE", file_path=EXCEL_FILE):
    """
    Menyimpan status pengerjaan data ke kolom 'STATUS' dan 'TGL UPDATE STATUS' pada file Excel:
    - row_num: Nomor baris di Excel (1-based, misal 2, 3, dst)
    - status_text: Pesan status (default: 'SUKSES DIUPDATE')
    """
    if not os.path.exists(file_path):
        print(f"[ERROR] File Excel '{file_path}' tidak ditemukan untuk simpan status.")
        return False

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Cari kolom 'STATUS' dan 'TGL UPDATE STATUS' di baris 1
        status_col = -1
        tgl_col = -1

        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value is not None:
                cell_val = str(cell.value).strip().upper()
                if cell_val == "STATUS":
                    status_col = col_idx
                elif "TGL" in cell_val or "TANGGAL" in cell_val or "UPDATE" in cell_val:
                    tgl_col = col_idx

        # Jika kolom STATUS belum ada, buat di kolom paling kanan
        if status_col == -1:
            status_col = ws.max_column + 1
            ws.cell(row=1, column=status_col, value="STATUS")
            print(f"[EXCEL] Membuat kolom baru 'STATUS' pada kolom ke-{status_col}")

        # Tulis nilai status pada baris yang bersangkutan
        ws.cell(row=row_num, column=status_col, value=status_text)

        # Tulis tanggal update status jika kolom 'TGL UPDATE STATUS' ada/ditemukan
        waktu_sekarang = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if tgl_col != -1:
            ws.cell(row=row_num, column=tgl_col, value=waktu_sekarang)
            print(f"[EXCEL] Berhasil menyimpan status '{status_text}' & tanggal '{waktu_sekarang}' pada baris {row_num} di Excel.")
        else:
            print(f"[EXCEL] Berhasil menyimpan status '{status_text}' pada baris {row_num} di Excel.")

        wb.save(file_path)
        wb.close()
        return True
    except PermissionError:
        print(f"[ERROR EXCEL] File '{file_path}' sedang dibuka di Microsoft Excel! Harap TUTUP file Excel di komputer Anda agar status dapat tersimpan.")
        return False
    except Exception as err:
        print(f"[ERROR EXCEL] Gagal menyimpan status ke Excel pada baris {row_num}: {err}")
        return False


def tunggu_loading(target_text=None, timeout=30, sleep_before=0.1):
    """
    Fungsi universal untuk menunggu loading screen / progress bar di Fasih (Referensi bot_emulator_idpel.py):
    - target_text: Teks elemen yang ditunggu setelah loading selesai (misal: 'Daftar Assignment', 'DITEMUKAN')
    - timeout: Batas maksimum waktu tunggu dalam detik (default: 30s)
    - sleep_before: Jeda sebelum mengecek agar progress bar sempat muncul (default: 0.1s)
    """
    t_start = time.time()
    waktu_str = datetime.now().strftime('%H:%M:%S.%f')[:-3]
    print(f"[{waktu_str}] [LOADING] Menunggu loading / progress bar selesai...")

    if sleep_before > 0:
        time.sleep(sleep_before)

    # 1. Tunggu progress bar (id.go.bpsfasih:id/card_progress) hilang dari layar
    try:
        if check_exists(d(resourceId="id.go.bpsfasih:id/card_progress")):
            d(resourceId="id.go.bpsfasih:id/card_progress").wait_gone(timeout=timeout)
    except Exception:
        pass

    # 2. Cek apakah limit API check-idpln tercapai
    limit_msg = "Permintaan API check-idpln sudah terlampaui (limit)."
    try:
        if check_exists(d(textContains=limit_msg)):
            waktu_limit = datetime.now().strftime('%H:%M:%S.%f')[:-3]
            print(f"[{waktu_limit}] [WARNING] PERMINTAAN API CEK ID MELEWATI LIMIT!")
            return False
    except Exception:
        pass

    # 3. Tunggu elemen target_text termuat jika diberikan (menggunakan active polling cepat)
    success = True
    if target_text:
        waktu_target = datetime.now().strftime('%H:%M:%S.%f')[:-3]
        print(f"[{waktu_target}] [LOADING] Menunggu halaman / teks '{target_text}' termuat di layar...")
        
        found = False
        t_end = time.time() + timeout
        while time.time() < t_end:
            try:
                if (check_exists(d(resourceId="id.go.bpsfasih:id/title_toolbar", text=target_text)) or 
                    check_exists(d(text=target_text)) or 
                    check_exists(d(textContains=target_text)) or 
                    check_exists(d(descriptionContains=target_text)) or
                    check_exists(d.xpath(f"//*[contains(@text, '{target_text}') or contains(@content-desc, '{target_text}')]"))):
                    found = True
                    break
            except Exception:
                pass
            time.sleep(0.2)

        if found:
            waktu_ok = datetime.now().strftime('%H:%M:%S.%f')[:-3]
            print(f"[{waktu_ok}] [SUKSES] Berhasil memuat halaman / teks '{target_text}'")
        else:
            waktu_warn = datetime.now().strftime('%H:%M:%S.%f')[:-3]
            print(f"[{waktu_warn}] [WARNING] Teks '{target_text}' tidak terdeteksi setelah {timeout} detik.")
            success = False

    durasi = time.time() - t_start
    waktu_selesai = datetime.now().strftime('%H:%M:%S.%f')[:-3]
    print(f"[{waktu_selesai}] [LOADING] Progress / loading selesai (+{durasi:.2f}s).")
    return success


def tunggu_loading_screen(target_text=None, timeout=30):
    """Alias universal untuk tunggu_loading()"""
    return tunggu_loading(target_text=target_text, timeout=timeout)


def tunggu_loading_cek_nik(timeout=30, sleep_before=0.3):
    """
    Fungsi khusus untuk menunggu loading screen / progress bar setelah mengetuk 'Cek NIK':
    - Menunggu progress bar (card_progress / ProgressBar) muncul & hilang dari layar
    - Polling hingga salah satu indikator hasil ('Hasil Pemadanan NIK', 'NIK tidak valid', dll) terdeteksi di layar.
    """
    t_start = time.time()
    waktu_str = datetime.now().strftime('%H:%M:%S.%f')[:-3]
    print(f"[{waktu_str}] [LOADING CEK NIK] Menunggu proses 'Cek NIK' / loading screen selesai...")

    if sleep_before > 0:
        time.sleep(sleep_before)

    # 1. Tunggu progress bar (card_progress atau ProgressBar) hilang dari layar
    try:
        if check_exists(d(resourceId="id.go.bpsfasih:id/card_progress")):
            d(resourceId="id.go.bpsfasih:id/card_progress").wait_gone(timeout=timeout)
        elif check_exists(d(className="android.widget.ProgressBar")):
            d(className="android.widget.ProgressBar").wait_gone(timeout=timeout)
    except Exception:
        pass

    # 2. Active polling hingga salah satu indikator selesai muncul di layar
    indikator_selesai = [
        "Hasil Pemadanan NIK",
        "NIK tidak valid",
        "periksa digit",
        "DITEMUKAN",
        "TIDAK DITEMUKAN",
        "NIK :"
    ]
    t_end = time.time() + timeout
    found = False
    while time.time() < t_end:
        try:
            for item in indikator_selesai:
                if (check_exists(d(textContains=item)) or 
                    check_exists(d(descriptionContains=item)) or
                    check_exists(d.xpath(f"//*[contains(@text, '{item}') or contains(@content-desc, '{item}')]"))):
                    found = True
                    break
            if found:
                break
        except Exception:
            pass
        time.sleep(0.2)

    durasi = time.time() - t_start
    waktu_selesai = datetime.now().strftime('%H:%M:%S.%f')[:-3]
    if found:
        print(f"[{waktu_selesai}] [SUKSES CEK NIK] Loading 'Cek NIK' selesai (+{durasi:.2f}s).")
    else:
        print(f"[{waktu_selesai}] [LOADING CEK NIK] Selesai menunggu (+{durasi:.2f}s).")
    return found


def is_textbox_disabled(label_text="202. NIK penghuni"):
    """
    Mengecek apakah form/textbox '202. NIK penghuni' terkunci / sudah disubmit dari awal (Referensi dump.xml L165-L175):
    - Jika clickable="true" & enabled="true" (L173: text="9999999999999998") -> EDITABLE / BISA DIISI (kembalikan False)
    - Jika clickable="false" / enabled="false" / 'DITEMUKAN' terdeteksi -> TERKUNCI / SUDAH SUBMIT (kembalikan True)
    """
    # 1. Teks 'Hasil Pemadanan NIK' atau 'DITEMUKAN' sudah tampil di layar dari awal
    try:
        if check_exists(d(textContains="Hasil Pemadanan NIK")) or check_exists(d(textContains="DITEMUKAN")):
            print("[CHECK] Terdeteksi 'Hasil Pemadanan NIK' / 'DITEMUKAN' sudah tampil di layar. Form sudah disubmit dari awal!")
            return True
    except Exception:
        pass

    # 2. Tombol 'Cek NIK' ada tetapi enabled="false"
    try:
        if check_exists(d(text="Cek NIK")):
            btn_info = d(text="Cek NIK").info
            if isinstance(btn_info, dict) and not btn_info.get("enabled", True):
                print("[CHECK] Tombol 'Cek NIK' dalam keadaan disabled (enabled=False). Form sudah disubmit dari awal!")
                return True
    except Exception:
        pass

    # 3. Node XPath EditText di bawah label (mengecek atribut clickable & enabled secara presisi)
    try:
        xp = f"//*[contains(@text, '{label_text}') or contains(@content-desc, '{label_text}')]/following::android.widget.EditText[1]"
        if check_exists(d.xpath(xp)):
            xp_obj = d.xpath(xp)
            node = xp_obj.get() if hasattr(xp_obj, 'get') else None
            info = node.info if (node and hasattr(node, 'info')) else (xp_obj.info if hasattr(xp_obj, 'info') else {})
            attrib = node.attrib if (node and hasattr(node, 'attrib')) else (xp_obj.attrib if hasattr(xp_obj, 'attrib') else {})

            # Cek status clickable (referensi dump.xml L173: clickable="true", enabled="true")
            is_clickable = info.get("clickable", True)
            if isinstance(is_clickable, str):
                is_clickable = is_clickable.lower() == "true"

            is_enabled = info.get("enabled", True)
            if isinstance(is_enabled, str):
                is_enabled = is_enabled.lower() == "true"

            is_focusable = info.get("focusable", True)
            if isinstance(is_focusable, str):
                is_focusable = is_focusable.lower() == "true"

            # Jika clickable=True dan enabled=True -> EDITABLE / BISA DIISI
            if is_clickable and is_enabled and is_focusable:
                print(f"[CHECK] Field '{label_text}' ber-status clickable=True, enabled=True (Referensi dump.xml L173) -> FORM EDITABLE / BISA DIISI.")
                return False
            else:
                print(f"[CHECK] Field '{label_text}' ber-status TERKUNCI (clickable={is_clickable}, enabled={is_enabled}). Form sudah disubmit dari awal!")
                return True
    except Exception:
        pass

    # 4. Selector uiautomator2 langsung pada EditText
    try:
        label_el = d(textContains=label_text)
        if check_exists(label_el):
            tb = label_el.down(className="android.widget.EditText")
            if check_exists(tb):
                info = tb.info
                if isinstance(info, dict):
                    clk = info.get("clickable", True)
                    en = info.get("enabled", True)
                    foc = info.get("focusable", True)
                    if clk and en and foc:
                        print(f"[CHECK] Field '{label_text}' via selector ber-status clickable=True -> FORM EDITABLE / BISA DIISI.")
                        return False
                    else:
                        print(f"[CHECK] Field '{label_text}' via selector TERKUNCI (clickable={clk}, enabled={en}). Form sudah disubmit dari awal!")
                        return True
    except Exception:
        pass

    # Default: anggap Form EDITABLE (BISA DIISI)
    return False







def kembali_ke_daftar_assignment(max_retry=5):
    """
    Menekan tombol BACK berulang kali hingga kembali ke halaman 'Daftar Assignment' (Referensi bot_emulator_idpel.py):
    - Jika muncul dialog konfirmasi 'Apakah Anda yakin akan keluar dari halaman ini ?', ketuk 'IYA'.
    """
    print("[RECOVERY] Memulai prosedur kembali ke halaman 'Daftar Assignment'...")
    for i in range(1, max_retry + 1):
        if (check_exists(d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Daftar Assignment")) or 
            check_exists(d(text="Daftar Assignment"))):
            print("[RECOVERY] Sudah berada di halaman 'Daftar Assignment'.")
            return True

        # Ketuk tombol OK statis (bounds 528, 1691) jika ada modal OK yang menghalangi sebelum BACK
        print(f"[RECOVERY] Mengetuk tombol OK statis (bounds 528, 1691) sebelum BACK...")
        try:
            btn_close = d(resourceId="id.go.bpsfasih:id/btn_submit_progress_close")
            if btn_close.exists:
                btn_close.click()
            elif d(text="OK").exists:
                d(text="OK").click()
            else:
                d.click(528, 1691)
        except Exception:
            d.click(528, 1691)
        time.sleep(0.5)

        print(f"[RECOVERY] Menekan tombol BACK (Percobaan {i}/{max_retry})...")
        try:
            d.press("back")
        except Exception:
            pass
        time.sleep(0.5)

        # Cek dialog konfirmasi keluar "Apakah Anda yakin akan keluar dari halaman ini ?"
        dialog_text = "Apakah Anda yakin akan keluar dari halaman ini ?"
        if check_exists(d(textContains=dialog_text)) or check_exists(d(resourceId="id.go.bpsfasih:id/deskripsi_bottomDialog")):
            print("[RECOVERY] Dialog konfirmasi keluar terdeteksi! Mengetuk 'IYA'...")
            try:
                if check_exists(d(resourceId="id.go.bpsfasih:id/rButton_bottomDialog", text="IYA")):
                    d(resourceId="id.go.bpsfasih:id/rButton_bottomDialog", text="IYA").click()
                elif check_exists(d(text="IYA")):
                    d(text="IYA").click()
                print("[RECOVERY] Tombol 'IYA' diketuk. Kembali ke Dashboard.")
                time.sleep(SLEEP_LONG)
            except Exception as e:
                print(f"[WARNING] Gagal mengetuk 'IYA': {e}")

        if (check_exists(d(resourceId="id.go.bpsfasih:id/title_toolbar", text="Daftar Assignment")) or 
            check_exists(d(text="Daftar Assignment"))):
            print("[RECOVERY] Berhasil kembali ke halaman 'Daftar Assignment'.")
            return True

    print("[WARNING] Gagal kembali ke 'Daftar Assignment' setelah mencoba BACK.")

    # Fallback jika berada di halaman 'Surveys'
    try:
        if check_exists(d(textContains="Surveys")) or check_exists(d(descriptionContains="Surveys")):
            print("[RECOVERY SURVEYS] Terdeteksi teks 'Surveys'! Mengeksekusi alur submit recovery (PRABAYAR -> Submit -> Submit)...")
            
            # 1. Ketuk teks 'PRABAYAR'
            ketuk("PRABAYAR", sleep_after=SLEEP_SHORT)
            time.sleep(SLEEP_SHORT)

            # 2. Ketuk tombol 'Submit'
            ketuk("Submit", sleep_after=SLEEP_SHORT)
            time.sleep(SLEEP_SHORT)

            # 3. Ketuk tombol 'Submit' konfirmasi lagi jika ada
            ketuk("Submit", sleep_after=SLEEP_SHORT)
            time.sleep(SLEEP_SHORT)

            # 4. Menunggu kembali ke 'Daftar Assignment'
            res = tunggu_loading("Daftar Assignment", timeout=15)
            if res:
                print("[RECOVERY SURVEYS] Berhasil kembali ke halaman 'Daftar Assignment' via alur Surveys.")
                return True
    except Exception as err:
        print(f"[WARNING] Gagal alur recovery Surveys: {err}")

    return False


def cari_dan_ketuk_search_box(idpel):
    """Mencari text box search berdasarkan hirarki node dump.xml, mengklik, memasukkan IDPEL, dan menekan Enter"""
    print(f"\n[SEARCH] Mencari text box search di layar...")
    
    search_input = None

    # 1. Target spesifik via hint="Search:" dan className EditText
    try:
        if d(className="android.widget.EditText", hint="Search:").exists():
            search_input = d(className="android.widget.EditText", hint="Search:")
            print("[SEARCH] Ditemukan via className 'android.widget.EditText' dan hint 'Search:'")
    except Exception:
        pass

    # 2. Target via XPath hint="Search:"
    if not search_input:
        try:
            xp_el = d.xpath("//android.widget.EditText[@hint='Search:']")
            if xp_el.exists:
                search_input = xp_el
                print("[SEARCH] Ditemukan via XPath //android.widget.EditText[@hint='Search:']")
        except Exception:
            pass

    # 3. Target via Sibling dari TextView "Search:"
    if not search_input:
        try:
            if d(text="Search:").exists():
                search_input = d(text="Search:").sibling(className="android.widget.EditText")
                print("[SEARCH] Ditemukan via sibling dari TextView 'Search:'")
            elif d(textContains="Search:").exists():
                search_input = d(textContains="Search:").sibling(className="android.widget.EditText")
                print("[SEARCH] Ditemukan via sibling dari TextView textContains 'Search:'")
        except Exception:
            pass

    # 4. Target via XPath sibling TextView "Search:"
    if not search_input:
        try:
            xp_sibling = d.xpath("//*[@text='Search:' or contains(@text, 'Search:')]/following-sibling::android.widget.EditText[1]")
            if xp_sibling.exists:
                search_input = xp_sibling
                print("[SEARCH] Ditemukan via XPath sibling //*[@text='Search:']/following-sibling::android.widget.EditText")
        except Exception:
            pass

    # 5. Target EditText pertama di layar
    if not search_input:
        try:
            if d(className="android.widget.EditText").exists():
                search_input = d(className="android.widget.EditText")
                print("[SEARCH] Ditemukan via EditText pertama di layar")
        except Exception:
            pass

    def check_exists(el):
        if el is None:
            return False
        try:
            if hasattr(el, 'exists'):
                if callable(el.exists):
                    return el.exists()
                return bool(el.exists)
        except Exception:
            pass
        return False

    # 6. Fallback jika selector tidak menemukan, klik koordinat tengah bounds [336,780][900,873] -> (618, 826)
    if not check_exists(search_input):
        print("[WARNING] Text box search tidak terdeteksi via selector. Mengklik koordinat bounds [336,780][900,873] (618, 826)...")
        d.click(618, 826)
        time.sleep(SLEEP_SHORT)
        if d(className="android.widget.EditText").exists():
            search_input = d(className="android.widget.EditText")

    if check_exists(search_input):
        print("[SEARCH] Mengetuk text box search...")
        try:
            search_input.click()
        except Exception:
            d.click(618, 826)
        time.sleep(SLEEP_SHORT)

        print(f"[INPUT] Memasukkan IDPEL: '{idpel}'...")
        try:
            if hasattr(search_input, 'set_text'):
                search_input.set_text(str(idpel))
            else:
                d.send_keys(str(idpel))
        except Exception:
            d.send_keys(str(idpel))
        time.sleep(SLEEP_SHORT)

        print("[ENTER] Menekan tombol Enter...")
        d.press("enter")
        time.sleep(SLEEP_MEDIUM)
        return True
    else:
        # Fallback klik koordinat langsung & send_keys
        print("[SEARCH] Mencoba klik koordinat (618, 826) & send_keys langsung...")
        d.click(618, 826)
        time.sleep(SLEEP_SHORT)
        d.send_keys(str(idpel))
        time.sleep(SLEEP_SHORT)
        d.press("enter")
        time.sleep(SLEEP_MEDIUM)
        return True


class Swipe:
    """Class Object Oriented untuk operasi swipe dinamis pada emulator"""
    def __init__(self, device=None):
        self.device = device

    def get_device(self):
        return self.device if self.device is not None else d

    def loop_swipe_dinamis(self, tengah_layar=None, delta_y=-700, target_text=None, max_retry=15, duration=0.05):
        dev = self.get_device()
        
        # 1. Penentuan koordinat titik tengah layar awal
        if tengah_layar is None:
            width = dev.info.get('displayWidth', 1080)
            height = dev.info.get('displayHeight', 1920)
            cx = width // 2
            cy = height // 2
        else:
            cx, cy = tengah_layar

        target_y = cy + delta_y
        print(f"[SWIPE] Swipe dinamis ({cx}, {cy} -> {cx}, {target_y}) | Target text: '{target_text}'")

        # Jika target_text tidak ditentukan, lakukan 1x swipe langsung
        if not target_text:
            try:
                dev.swipe(cx, cy, cx, target_y, duration=duration)
            except Exception:
                duration_ms = int(duration * 1000)
                try:
                    dev.shell(f"input swipe {cx} {cy} {cx} {target_y} {duration_ms}")
                except Exception:
                    pass
            return True

        # Loop swipe sampai target_text terdeteksi di layar
        for attempt in range(1, max_retry + 1):
            found = False
            try:
                if dev(textContains=target_text).exists() or dev(descriptionContains=target_text).exists():
                    found = True
                else:
                    xp = f"//*[contains(@text, '{target_text}') or contains(@content-desc, '{target_text}')]"
                    if dev.xpath(xp).exists:
                        found = True
            except Exception:
                pass

            if found:
                print(f"[SWIPE] Text target '{target_text}' terdeteksi di layar pada percobaan ke-{attempt}!")
                return True

            try:
                dev.swipe(cx, cy, cx, target_y, duration=duration)
            except Exception:
                duration_ms = int(duration * 1000)
                try:
                    dev.shell(f"input swipe {cx} {cy} {cx} {target_y} {duration_ms}")
                except Exception:
                    pass

            time.sleep(0.05)

        print(f"[WARNING] Target text '{target_text}' tidak terdeteksi setelah {max_retry}x swipe.")
        return False

    def loop_swipe_statis(self, delta_y=700, loop=3, tengah_layar=None, duration=0.05):
        dev = self.get_device()
        
        if tengah_layar is None:
            width = dev.info.get('displayWidth', 1080)
            height = dev.info.get('displayHeight', 1920)
            cx = width // 2
            cy = height // 2
        else:
            cx, cy = tengah_layar

        target_y = cy + delta_y
        print(f"[SWIPE STATIS] Melakukan {loop}x swipe statis ({cx}, {cy} -> {cx}, {target_y})...")

        for idx in range(1, loop + 1):
            try:
                dev.swipe(cx, cy, cx, target_y, duration=duration)
            except Exception:
                duration_ms = int(duration * 1000)
                try:
                    dev.shell(f"input swipe {cx} {cy} {cx} {target_y} {duration_ms}")
                except Exception:
                    pass
            time.sleep(0.05)
        return True

    def custom(self, fx, fy, tx, ty, duration=0.1):
        dev = self.get_device()
        try:
            dev.swipe(fx, fy, tx, ty, duration=duration)
        except Exception:
            duration_ms = int(duration * 1000)
            try:
                dev.shell(f"input swipe {fx} {fy} {tx} {ty} {duration_ms}")
            except Exception:
                pass


def loop_swipe_dinamis(tengah_layar=None, delta_y=-700, target_text=None, max_retry=15, duration=0.05):
    """
    Fungsi universal OOP untuk swipe dinamis:
    - tengah_layar: tuple (cx, cy) koordinat awal swipe (default: tengah layar)
    - delta_y: jarak y swipe (default: -700)
    - target_text: teks yang dicari dalam loop (default: None)
    """
    s = Swipe(d)
    return s.loop_swipe_dinamis(tengah_layar=tengah_layar, delta_y=delta_y, target_text=target_text, max_retry=max_retry, duration=duration)


def loop_swipe_statis(delta_y=700, loop=3, tengah_layar=None, duration=0.05):
    """
    Fungsi universal OOP untuk swipe statis:
    - delta_y: jarak offset y (default: 700)
    - loop: jumlah perulangan swipe statis (default: 3)
    - tengah_layar: tuple (cx, cy) titik awal swipe (default: tengah layar)
    """
    s = Swipe(d)
    return s.loop_swipe_statis(delta_y=delta_y, loop=loop, tengah_layar=tengah_layar, duration=duration)


def swipe_aman(fx, fy, tx, ty, duration=0.1):
    """Helper swipe_aman yang memanggil method OOP Swipe"""
    s = Swipe(d)
    s.custom(fx, fy, tx, ty, duration=duration)


def ketuk(target_text, exact=False, sleep_after=SLEEP_SHORT):
    """
    Fungsi universal & reusable untuk mencari dan mengetuk elemen teks/description/XPath.
    Memprioritaskan elemen berjenis android.widget.Button atau yang bernilai clickable=True.
    """
    print(f"[KLIK] Mencari dan mengetuk '{target_text}'...")
    target_btn = None

    # 1. Prioritas 1: Target spesifik android.widget.Button
    try:
        btn_el = d(className="android.widget.Button", text=target_text)
        if not btn_el.exists():
            btn_el = d(className="android.widget.Button", textContains=target_text)
        if btn_el.exists():
            target_btn = btn_el
            print(f"[KLIK] Ditemukan via android.widget.Button: '{target_text}'")
    except Exception:
        pass

    # 2. Prioritas 2: Target elemen dengan clickable=True
    if not target_btn:
        try:
            if exact:
                click_el = d(text=target_text, clickable=True)
                if not click_el.exists():
                    click_el = d(description=target_text, clickable=True)
            else:
                click_el = d(textContains=target_text, clickable=True)
                if not click_el.exists():
                    click_el = d(descriptionContains=target_text, clickable=True)

            if click_el.exists():
                target_btn = click_el
                print(f"[KLIK] Ditemukan via clickable=True: '{target_text}'")
        except Exception:
            pass

    # 3. Prioritas 3: Target via XPath Button / clickable=True
    if not target_btn:
        try:
            if exact:
                xp = f"//android.widget.Button[@text='{target_text}' or @content-desc='{target_text}'] | //*[@clickable='true' and (@text='{target_text}' or @content-desc='{target_text}')]"
            else:
                xp = f"//android.widget.Button[contains(@text, '{target_text}') or contains(@content-desc, '{target_text}')] | //*[@clickable='true' and (contains(@text, '{target_text}') or contains(@content-desc, '{target_text}'))]"

            if d.xpath(xp).exists:
                target_btn = d.xpath(xp)
                print(f"[KLIK] Ditemukan via XPath clickable: '{xp}'")
        except Exception:
            pass

    # 4. Fallback selector umum jika di atas tidak ketemu
    if not target_btn:
        if exact:
            if d(text=target_text).exists():
                target_btn = d(text=target_text)
            elif d(description=target_text).exists():
                target_btn = d(description=target_text)
        else:
            if d(textContains=target_text).exists():
                target_btn = d(textContains=target_text)
                try:
                    print(f"[KLIK] Ditemukan via textContains: '{target_btn.info.get('text', target_text)}'")
                except Exception:
                    pass
            elif d(descriptionContains=target_text).exists():
                target_btn = d(descriptionContains=target_text)

    if not target_btn:
        try:
            if exact:
                xp = f"//*[@text='{target_text}' or @content-desc='{target_text}']"
            else:
                xp = f"//*[contains(@text, '{target_text}') or contains(@content-desc, '{target_text}')]"

            if d.xpath(xp).exists:
                target_btn = d.xpath(xp)
                print(f"[KLIK] Ditemukan via XPath fallback: '{xp}'")
        except Exception:
            pass

    # 5. Eksekusi Klik
    success = False
    if target_btn:
        try:
            target_btn.click()
            print(f"[KLIK] Berhasil mengetuk '{target_text}'")
            success = True
        except Exception:
            try:
                if exact:
                    d.xpath(f"//android.widget.Button[@text='{target_text}'] | //*[@text='{target_text}']").click()
                else:
                    d.xpath(f"//android.widget.Button[contains(@text, '{target_text}')] | //*[contains(@text, '{target_text}')]").click()
                print(f"[KLIK] Berhasil mengetuk '{target_text}' via XPath fallback click")
                success = True
            except Exception as err:
                print(f"[WARNING] Gagal mengetuk '{target_text}': {err}")

    # 6. Fallback khusus untuk tombol 'Aksi', 'Buka' / 'BUKA', dan 'sidebar-toggle' jika belum berhasil diklik
    if not success and target_text.upper() in ["BUKA", "AKSI", "SIDEBAR-TOGGLE"]:
        if target_text.upper() == "BUKA":
            print("[KLIK] Mengetuk tombol 'BUKA' via resourceId 'id.go.bpsfasih:id/openAssignment_b'...")
            try:
                if d(resourceId="id.go.bpsfasih:id/openAssignment_b").exists():
                    d(resourceId="id.go.bpsfasih:id/openAssignment_b").click()
                    success = True
            except Exception:
                pass
            if not success:
                print("[KLIK] Mengetuk tombol 'BUKA' via koordinat bounds [63,905][1017,1014] -> (540, 959)...")
                try:
                    d.click(540, 959)
                    success = True
                except Exception as e:
                    print(f"[ERROR] Gagal klik koordinat BUKA: {e}")
        elif target_text == "Aksi":
            print("[KLIK] Mengetuk tombol 'Aksi' via koordinat bounds [306,1296][468,1401] -> (387, 1348)...")
            try:
                d.click(387, 1348)
                success = True
            except Exception as e:
                print(f"[ERROR] Gagal klik koordinat Aksi: {e}")
        elif target_text == "sidebar-toggle":
            print("[KLIK] Mengetuk tombol 'sidebar-toggle' via koordinat bounds [24,114][120,210] -> (72, 162)...")
            try:
                d.click(72, 162)
                success = True
            except Exception as e:
                print(f"[ERROR] Gagal klik koordinat sidebar-toggle: {e}")
        elif target_text.upper() == "BLOK II":
            print("[KLIK] Mengetuk 'BLOK II' via koordinat sidebar bounds [570, 792] -> (350, 681)...")
            try:
                d.click(350, 681)
                success = True
            except Exception as e:
                print(f"[ERROR] Gagal klik koordinat BLOK II: {e}")
        elif target_text.upper() == "BATAL":
            print("[KLIK] Mengetuk tombol 'Batal' via koordinat bounds [564,1629][729,1725] -> (646, 1677)...")
            try:
                d.click(646, 1677)
                success = True
            except Exception as e:
                print(f"[ERROR] Gagal klik koordinat Batal: {e}")

    if sleep_after > 0:
        time.sleep(sleep_after)
    return success


def check_dan_tutup_pengaturan():
    """
    Memeriksa apakah halaman/dialog 'Pengaturan' muncul secara tidak sengaja (Referensi dump.xml).
    Jika muncul, mengetuk tombol 'Batal' (via selector atau koordinat bounds [564,1629][729,1725] -> (646, 1677)) untuk menutupnya.
    """
    try:
        is_pengaturan = (
            check_exists(d(text="Pengaturan")) or
            check_exists(d(textContains="Pengaturan")) or
            check_exists(d(descriptionContains="Pengaturan")) or
            check_exists(d.xpath("//*[contains(@text, 'Pengaturan') or contains(@content-desc, 'Pengaturan')]"))
        )
        if is_pengaturan:
            print("[PENGATURAN] Terdeteksi dialog/halaman 'Pengaturan'! Mengetuk 'Batal'...")
            clicked = False
            try:
                btn_batal = d(text="Batal")
                if not check_exists(btn_batal):
                    btn_batal = d(textContains="Batal")
                if check_exists(btn_batal):
                    btn_batal.click()
                    clicked = True
                elif check_exists(d.xpath("//*[contains(@text, 'Batal') or contains(@content-desc, 'Batal')]")):
                    d.xpath("//*[contains(@text, 'Batal') or contains(@content-desc, 'Batal')]").click()
                    clicked = True
            except Exception:
                pass

            if not clicked:
                print("[PENGATURAN] Mengetuk 'Batal' via koordinat statis bounds [564,1629][729,1725] -> (646, 1677)...")
                d.click(646, 1677)

            time.sleep(SLEEP_SHORT)
            print("[PENGATURAN] Melakukan 1x swipe down setelah mengetuk 'Batal'...")
            loop_swipe_statis(delta_y=700, loop=1)
            time.sleep(SLEEP_SHORT)
            return True
    except Exception as e:
        print(f"[WARNING] Error pengecekan Pengaturan: {e}")
    return False


def pilih_blok(nama_blok, max_attempts=3):
    """
    Fungsi universal untuk beralih ke BLOK I, BLOK II, BLOK III, atau BLOK IV di Fasih:
    - Referensi koordinat fallback dari dump.xml:
      * BLOK I   : (350, 450)  [bounds: -840,339][-27,561]
      * BLOK II  : (350, 681)  [bounds: -840,570][-27,792]
      * BLOK III : (350, 912)  [bounds: -840,801][-27,1023]
      * BLOK IV  : (350, 1119) [bounds: -840,1032][-27,1206]
    - Jika malah mengetuk 'Pengaturan' dan muncul dialog 'Pengaturan',
      akan mengetuk tombol 'Batal' dan mengulangi pengetukan blok hingga max_attempts kali.
    """
    if not nama_blok:
        print("[BLOK] Parameter nama_blok kosong.")
        return False

    # Normalisasi masukan (misal: '1' -> 'BLOK I', '2' -> 'BLOK II', 'III' -> 'BLOK III')
    target = str(nama_blok).strip().upper()
    if target in ["1", "I"]:
        target = "BLOK I"
    elif target in ["2", "II"]:
        target = "BLOK II"
    elif target in ["3", "III"]:
        target = "BLOK III"
    elif target in ["4", "IV"]:
        target = "BLOK IV"
    elif not target.startswith("BLOK"):
        target = f"BLOK {target}"

    # Mapping koordinat presisi sidebar menu dari dump.xml
    coords_map = {
        "BLOK I": (350, 450),
        "BLOK II": (350, 681),
        "BLOK III": (350, 912),
        "BLOK IV": (350, 1119)
    }

    for attempt in range(1, max_attempts + 1):
        print(f"[BLOK] Mencari dan beralih ke '{target}' (Percobaan {attempt}/{max_attempts})...")

        # Cek jika dialog Pengaturan sedang terbuka sebelum mengetuk
        if check_dan_tutup_pengaturan():
            time.sleep(0.5)

        clicked = False

        # 1. Coba klik via uiautomator2 text / description / textContains
        try:
            if check_exists(d(text=target)):
                d(text=target).click()
                print(f"[BLOK] Berhasil mengetuk '{target}' via text")
                clicked = True
            elif check_exists(d(textContains=target)):
                d(textContains=target).click()
                print(f"[BLOK] Berhasil mengetuk '{target}' via textContains")
                clicked = True
            elif check_exists(d(descriptionContains=target)):
                d(descriptionContains=target).click()
                print(f"[BLOK] Berhasil mengetuk '{target}' via descriptionContains")
                clicked = True
        except Exception:
            pass

        # 2. Coba klik via XPath
        if not clicked:
            try:
                xp = f"//*[contains(@text, '{target}') or contains(@content-desc, '{target}')]"
                if check_exists(d.xpath(xp)):
                    d.xpath(xp).click()
                    print(f"[BLOK] Berhasil mengetuk '{target}' via XPath")
                    clicked = True
            except Exception:
                pass

        # 3. Fallback koordinat presisi sidebar menu dari dump.xml
        if not clicked and target in coords_map:
            cx, cy = coords_map[target]
            print(f"[BLOK] Mengetuk '{target}' via koordinat fallback sidebar ({cx}, {cy})...")
            try:
                d.click(cx, cy)
                print(f"[BLOK] Berhasil mengetuk koordinat '{target}' ({cx}, {cy})")
                clicked = True
            except Exception as err:
                print(f"[ERROR] Gagal mengetuk '{target}' via koordinat: {err}")

        time.sleep(SLEEP_MEDIUM)

        # 4. Pengecekan setelah ketuk: Apakah muncul halaman/dialog 'Pengaturan'?
        if check_dan_tutup_pengaturan():
            print(f"[BLOK] [RETRY] Pengetukan '{target}' terlempar ke halaman 'Pengaturan'. Mengulangi pengetukan {target}...")
            time.sleep(0.5)
            continue
        else:
            if clicked:
                return True

    print(f"[WARNING] Gagal beralih ke '{target}' setelah {max_attempts}x percobaan.")
    return False


def blok(nama_blok):
    """Alias universal untuk pilih_blok()"""
    return pilih_blok(nama_blok)


def ketuk_blok_ii():
    """Wrapper ketuk untuk 'BLOK II'"""
    return pilih_blok("BLOK II")


def pilih_radio_button(option_text, exact=False, sleep_after=SLEEP_SHORT):
    """
    Fungsi universal untuk mengetuk RadioButton secara dinamis berdasarkan teks opsi:
    - Menggunakan XPath dinamis relative ke teks (//*[contains(@text, '{target}')]/..//*[@clickable='true'])
    - Tidak meng-hardcode koordinat/bounds sehingga selalu akurat di posisi swipe mana pun.
    - Mencari lingkaran RadioButton yang clickable (NAF=true, class=RadioButton) dalam satu baris container.

    Parameter:
        option_text: Teks atau nomor opsi (misal: '1. Berhasil didata', '3. Responden menolak', '1. Milik sendiri')
        exact: True jika pencarian teks harus sama persis
        sleep_after: Waktu tunggu dalam detik setelah mengetuk
    Returns:
        bool: True jika berhasil mengetuk, False jika gagal
    """
    if not option_text:
        print("[RADIO] Parameter option_text kosong.")
        return False

    target = str(option_text).strip()
    print(f"[RADIO] Mencari dan mengetuk RadioButton '{target}'...")

    # 1. Strategi Utama: XPath Dinamis mencari elemen @clickable='true' pada parent row teks target
    try:
        xp = f"//*[contains(@text, '{target}') or contains(@content-desc, '{target}')]/..//*[@clickable='true']"
        if exact:
            xp = f"//*[@text='{target}' or @content-desc='{target}']/..//*[@clickable='true']"

        elems = d.xpath(xp).all()
        if elems:
            for elem in elems:
                info = elem.info
                bounds = info.get("bounds", {})
                w = bounds.get("right", 0) - bounds.get("left", 0)
                h = bounds.get("bottom", 0) - bounds.get("top", 0)
                
                # Pastikan ini lingkaran RadioButton visual (bukan node dummy 3x6 px)
                if w >= 20 and h >= 20:
                    elem.click()
                    print(f"[RADIO] Berhasil mengetuk RadioButton '{target}' via dynamic XPath (@clickable='true')")
                    if sleep_after > 0:
                        time.sleep(sleep_after)
                    return True

            # Fallback jika elemen pertama node dummy, ketuk elemen kedua / parent
            elems[0].click()
            print(f"[RADIO] Berhasil mengetuk RadioButton '{target}' via XPath element")
            if sleep_after > 0:
                time.sleep(sleep_after)
            return True
    except Exception as e:
        print(f"[RADIO] Gagal via dynamic XPath sibling: {e}")

    # 2. Strategi 2: Ketuk parent row container (android.view.View) secara dinamis
    try:
        xp_row = f"//*[contains(@text, '{target}') or contains(@content-desc, '{target}')]/parent::*"
        if check_exists(d.xpath(xp_row)):
            d.xpath(xp_row).click()
            print(f"[RADIO] Berhasil mengetuk RadioButton '{target}' via dynamic parent row")
            if sleep_after > 0:
                time.sleep(sleep_after)
            return True
    except Exception as e:
        print(f"[RADIO] Gagal via dynamic parent row: {e}")

    # 3. Strategi 3: Selector u2 langsung
    try:
        el = d(text=target) if exact else d(textContains=target)
        if check_exists(el):
            el.click()
            print(f"[RADIO] Berhasil mengetuk RadioButton '{target}' via u2 text selector")
            if sleep_after > 0:
                time.sleep(sleep_after)
            return True
    except Exception as e:
        print(f"[RADIO] Gagal via u2 selector: {e}")

    print(f"[WARNING] Gagal menemukan/mengetuk RadioButton '{target}'.")
    return False


def cek_radio_button_tercentang(option_text, exact=False):
    """
    Memeriksa apakah RadioButton dengan teks tertentu sedang berstatus tercentang/terpilih.
    
    Pada app WebView/Flutter (Fasih), atribut checked/selected/focused selalu false.
    Deteksi dilakukan dengan memeriksa apakah elemen RadioButton visual (NAF="true", index=1)
    yang bersaudara dengan label teks target memiliki child element (android.widget.TextView)
    di dalamnya. Jika ada child = tercentang (filled circle), jika tidak ada = belum tercentang.
    
    Returns: bool (True jika tercentang, False jika belum)
    """
    if not option_text:
        return False
    
    target = str(option_text).strip()
    import xml.etree.ElementTree as ET
    
    try:
        xml_data = d.dump_hierarchy()
        root = ET.fromstring(xml_data)
        
        # Strategi 1: Cari parent View yang berisi RadioButton dengan teks target 
        # dan cek sibling RadioButton NAF=true apakah punya child (= tercentang)
        for parent in root.iter():
            children = list(parent)
            if len(children) < 2:
                continue
            
            # Cek apakah parent ini mengandung RadioButton dengan teks target
            has_target_radio = False
            naf_radio = None
            
            for child in children:
                child_class = child.attrib.get('class', '')
                child_text = (child.attrib.get('text') or '').strip()
                child_naf = child.attrib.get('NAF', '')
                
                # Cek apakah ini RadioButton dengan teks target (index=0 biasanya)
                if 'RadioButton' in child_class and child_text:
                    if exact:
                        if child_text == target:
                            has_target_radio = True
                    else:
                        if target in child_text:
                            has_target_radio = True
                
                # Cek apakah ini RadioButton visual (NAF=true, tanpa teks)
                if 'RadioButton' in child_class and child_naf == 'true' and not child_text:
                    naf_radio = child
            
            # Jika tidak ditemukan RadioButton berteks, cek juga View dengan teks target
            if not has_target_radio:
                for child in children:
                    child_class = child.attrib.get('class', '')
                    child_text = (child.attrib.get('text') or '').strip()
                    if 'View' in child_class and child_text:
                        if exact and child_text == target:
                            has_target_radio = True
                            break
                        elif not exact and target in child_text:
                            has_target_radio = True
                            break
            
            if has_target_radio and naf_radio is not None:
                # Cek apakah RadioButton NAF=true punya child element
                naf_children = list(naf_radio)
                has_inner = len(naf_children) > 0
                
                if has_inner:
                    print(f"[RADIO CEK] '{target}' TERDETEKSI terpilih! (RadioButton NAF=true memiliki {len(naf_children)} child element)")
                    return True
                else:
                    print(f"[RADIO CEK] '{target}' TIDAK terpilih (RadioButton NAF=true tidak memiliki child element)")
                    return False
        
        # Strategi 2 (fallback): Cek atribut standar checked/focused/selected 
        for node in root.iter():
            node_text = (node.attrib.get('text') or '').strip()
            node_class = node.attrib.get('class', '')
            
            if 'RadioButton' not in node_class or not node_text:
                continue
            
            match = False
            if exact:
                match = (node_text == target)
            else:
                match = (target in node_text)
            
            if match:
                focused = node.attrib.get('focused', 'false') == 'true'
                checked = node.attrib.get('checked', 'false') == 'true'
                selected = node.attrib.get('selected', 'false') == 'true'
                print(f"[RADIO CEK] Fallback: Node '{node_text}': focused={focused}, checked={checked}, selected={selected}")
                if focused or checked or selected:
                    print(f"[RADIO CEK] '{target}' TERDETEKSI terpilih via fallback!")
                    return True
        
        print(f"[RADIO CEK] '{target}' TIDAK terdeteksi terpilih pada semua pengecekan.")
    except Exception as e:
        print(f"[RADIO] Error saat cek status tercentang '{target}': {e}")
    
    return False


def input_textbox(label_text, value, bounds_fallback=None, exact=False, sleep_after=SLEEP_MEDIUM):
    """
    Fungsi universal & reusable untuk mengisi field EditText berdasarkan label teks (TextView):
    - label_text: Teks label penanda (misal: '202. NIK penghuni' atau '202.')
    - value: Nilai string yang akan diisikan ke EditText
    - bounds_fallback: Tuple (cx, cy) koordinat fallback jika selector tidak ditemukan
    - exact: True jika pencarian label harus persis (exact match)
    """
    if value is None:
        print(f"[INPUT] Nilai untuk '{label_text}' kosong, melewati pengisian.")
        return False

    val_str = str(value).strip()
    print(f"[INPUT] Mencari field '{label_text}' untuk menginput: '{val_str}'...")

    target_input = None

    # 1. Target via .down() / .sibling() dari label
    try:
        if exact:
            label = d(text=label_text)
            if not check_exists(label):
                label = d(description=label_text)
        else:
            label = d(textContains=label_text)
            if not check_exists(label):
                label = d(descriptionContains=label_text)

        if check_exists(label):
            temp_input = label.down(className="android.widget.EditText")
            if check_exists(temp_input):
                target_input = temp_input
            else:
                temp_sibling = label.sibling(className="android.widget.EditText")
                if check_exists(temp_sibling):
                    target_input = temp_sibling
    except Exception:
        pass

    # 2. Target via XPath (following EditText)
    if not check_exists(target_input):
        try:
            if exact:
                xp = f"//*[@text='{label_text}' or @content-desc='{label_text}']/following::android.widget.EditText[1]"
            else:
                xp = f"//*[contains(@text, '{label_text}') or contains(@content-desc, '{label_text}')]/following::android.widget.EditText[1]"
            
            if check_exists(d.xpath(xp)):
                target_input = d.xpath(xp)
                print(f"[INPUT] Ditemukan via XPath: {xp}")
        except Exception:
            pass

    # 3. Eksekusi Set Text
    success = False
    if check_exists(target_input):
        try:
            target_input.click()
            time.sleep(SLEEP_SHORT)
            try:
                if hasattr(target_input, 'clear_text') and callable(target_input.clear_text):
                    target_input.clear_text()
            except Exception:
                pass
            
            if hasattr(target_input, 'set_text') and callable(target_input.set_text):
                target_input.set_text(val_str)
            else:
                d.send_keys(val_str)
            print(f"[INPUT] Berhasil mengisi '{label_text}' dengan: '{val_str}'")
            success = True
        except Exception as e:
            print(f"[WARNING] Gagal set_text pada '{label_text}': {e}")

    # 4. Fallback koordinat jika bounds_fallback diberikan atau default khusus '202'
    if not success:
        if bounds_fallback:
            cx, cy = bounds_fallback
            print(f"[INPUT] Mengetuk fallback koordinat ({cx}, {cy}) untuk '{label_text}'...")
            try:
                d.click(cx, cy)
                time.sleep(SLEEP_SHORT)
                d.send_keys(val_str)
                print(f"[INPUT] Berhasil mengisi '{label_text}' via koordinat ({cx}, {cy}): '{val_str}'")
                success = True
            except Exception as err:
                print(f"[ERROR] Gagal mengisi '{label_text}' via koordinat: {err}")
        elif "202" in label_text or "NIK" in label_text:
            print(f"[INPUT] Mengetuk fallback koordinat (540, 1156) untuk '{label_text}'...")
            try:
                d.click(540, 1156)
                time.sleep(SLEEP_SHORT)
                d.send_keys(val_str)
                print(f"[INPUT] Berhasil mengisi '{label_text}' via koordinat (540, 1156): '{val_str}'")
                success = True
            except Exception as err:
                print(f"[ERROR] Gagal mengisi '{label_text}' via koordinat: {err}")

    if sleep_after > 0:
        time.sleep(sleep_after)
    return success

def normalisasi_nama_desa(val):
    """
    Normalisasi nama Desa/Kelurahan khusus variasi Padang Sambian & Pemecutan:
    - 'PADANG SAMBIAN KAJA' -> 'PADANGSAMBIAN KAJA'
    - 'PADANG SAMBIAN KELOD' / 'PADANGSAMBIAN KELOD' -> 'PADANGSAMBIAN KLOD'
    - 'PADANG SAMBIAN' -> 'PADANGSAMBIAN'
    - 'PEMECUTAN KELOD' -> 'PEMECUTAN KLOD'
    """
    if not val:
        return val
    replacements = [
        ("PADANG SAMBIAN KELOD", "PADANGSAMBIAN KLOD"),
        ("PADANGSAMBIAN KELOD", "PADANGSAMBIAN KLOD"),
        ("Padang Sambian Kelod", "PADANGSAMBIAN KLOD"),
        ("Padangsambian Kelod", "PADANGSAMBIAN KLOD"),
        ("PADANG SAMBIAN KAJA", "PADANGSAMBIAN KAJA"),
        ("Padang Sambian Kaja", "PADANGSAMBIAN KAJA"),
        ("PADANG SAMBIAN", "PADANGSAMBIAN"),
        ("Padang Sambian", "PADANGSAMBIAN"),
        ("PEMECUTAN KELOD", "PEMECUTAN KLOD"),
        ("Pemecutan Kelod", "PEMECUTAN KLOD"),
    ]
    for old_str, new_str in replacements:
        if old_str in val:
            val = val.replace(old_str, new_str)
    return val

def verifikasi_nama_penghuni_terisi(d_dev):
    """
    Melakukan scan halaman untuk memverifikasi apakah '201. Nama penghuni' sudah terisi di UI
    (tidak kosong / bukan 'Wajib diisi'), tanpa mengecek/mencocokkan dengan data nama dari Excel.
    """
    val_existing = ""
    try:
        label = d_dev(textContains="201. Nama penghuni")
        if not check_exists(label):
            label = d_dev(descriptionContains="201. Nama penghuni")

        if check_exists(label):
            input_el = label.down(className="android.widget.EditText")
            if not check_exists(input_el):
                input_el = label.sibling(className="android.widget.EditText")
            if check_exists(input_el):
                val_existing = (input_el.get_text() or "").strip()

        if not val_existing:
            xp = "//*[contains(@text, '201. Nama penghuni') or contains(@content-desc, '201. Nama penghuni')]/following::android.widget.EditText[1]"
            if check_exists(d_dev.xpath(xp)):
                val_existing = (d_dev.xpath(xp).info.get('text', '') or "").strip()
    except Exception as e:
        print(f"[SCAN NAMA] Exception saat verifikasi nama: {e}")

    print(f"[SCAN NAMA] Hasil scan field '201. Nama penghuni': '{val_existing}'")
    if val_existing and val_existing.lower() not in ["wajib diisi", "pilih salah satu...", ""]:
        return True
    return False

def cek_nik_tidak_valid(d_dev):
    """
    Melakukan scan halaman untuk mengonfirmasi apakah ada pesan 'NIK tidak valid' di layar.
    Returns:
        tuple (is_invalid: bool, pesan_error: str)
    """
    invalid_keywords = [
        "NIK tidak valid",
        "periksa digit",
        "digit pertama",
        "digit ketiga",
        "digit ke-7"
    ]
    try:
        for kw in invalid_keywords:
            if check_exists(d_dev(textContains=kw)) or check_exists(d_dev(descriptionContains=kw)):
                txt = d_dev(textContains=kw).info.get('text', '') if check_exists(d_dev(textContains=kw)) else kw
                return True, txt
        
        xp = "//*[contains(@text, 'NIK tidak valid') or contains(@content-desc, 'NIK tidak valid') or contains(@text, 'periksa digit')]"
        if check_exists(d_dev.xpath(xp)):
            nodes = d_dev.xpath(xp).all()
            if nodes:
                txt = nodes[0].text if hasattr(nodes[0], 'text') and nodes[0].text else "NIK tidak valid"
                return True, txt
    except Exception as e:
        print(f"[SCAN NIK] Exception saat memindai NIK tidak valid: {e}")

    return False, ""


def cari_dan_scroll_ke_tombol_cek_nik(max_swipes=10):
    """
    Mencari terlebih dahulu tombol 'Cek NIK' sebelum mengetuknya.
    Jika belum ketemu, melakukan loop_swipe_statis(delta_y=-200, loop=1)
    dan mengulangi maksimal 10 kali sampai ketemu.
    """
    for attempt in range(1, max_swipes + 1):
        found = False
        try:
            btn_el = d(className="android.widget.Button", text="Cek NIK")
            if not check_exists(btn_el):
                btn_el = d(className="android.widget.Button", textContains="Cek NIK")
            if not check_exists(btn_el):
                btn_el = d(text="Cek NIK")
            if not check_exists(btn_el):
                btn_el = d(textContains="Cek NIK")
            if not check_exists(btn_el):
                btn_el = d(description="Cek NIK")
            if not check_exists(btn_el):
                btn_el = d(descriptionContains="Cek NIK")
            if not check_exists(btn_el):
                xp = "//*[contains(@text, 'Cek NIK') or contains(@content-desc, 'Cek NIK')]"
                btn_el = d.xpath(xp)

            if check_exists(btn_el):
                found = True
        except Exception as e:
            print(f"[CEK NIK] Exception saat mencari tombol 'Cek NIK': {e}")

        if found:
            print(f"[CEK NIK] Tombol 'Cek NIK' ditemukan di layar (Pencarian ke-{attempt}).")
            return True

        print(f"[CEK NIK] Tombol 'Cek NIK' belum ditemukan (Pencarian ke-{attempt}/{max_swipes}). Melakukan swipe ke bawah...")
        loop_swipe_statis(delta_y=-200, loop=1)
        time.sleep(0.5)

    print(f"[CEK NIK] Tombol 'Cek NIK' tidak ditemukan setelah {max_swipes}x swipe.")
    return False


def baca_nilai_field_nik(d_dev):
    """
    Membaca nilai yang sedang terisi di field '202. NIK penghuni' pada UI.
    Returns: string isi field (kosong jika tidak ditemukan/gagal).
    """
    val_existing = ""
    try:
        label = d_dev(textContains="202. NIK penghuni")
        if not check_exists(label):
            label = d_dev(descriptionContains="202. NIK penghuni")

        if check_exists(label):
            input_el = label.down(className="android.widget.EditText")
            if not check_exists(input_el):
                input_el = label.sibling(className="android.widget.EditText")
            if check_exists(input_el):
                val_existing = (input_el.get_text() or "").strip()

        if not val_existing:
            xp = "//*[contains(@text, '202. NIK penghuni') or contains(@content-desc, '202. NIK penghuni')]/following::android.widget.EditText[1]"
            if check_exists(d_dev.xpath(xp)):
                val_existing = (d_dev.xpath(xp).info.get('text', '') or "").strip()
    except Exception as e:
        print(f"[SCAN NIK] Exception saat membaca nilai field NIK: {e}")

    print(f"[SCAN NIK] Nilai field '202. NIK penghuni' di UI: '{val_existing}'")
    return val_existing


def ambil_data_alamat(file_output="temp_alamat.txt", idpel=""):
    """
    Fungsi untuk men-scroll dan mengambil data alamat (Provinsi, Kabupaten, Kecamatan, Desa/Kelurahan, Alamat)
    dari BLOK I pada form Fasih berdasarkan referensi dump.xml (L225-L327) & bot_emulator_idpel2.py (L710-L855).

    Format output di temp_alamat.txt:
        Provinsi: [51] BALI
        Kabupaten: [5103] KAB. BADUNG
        Kecamatan: [510305] KUTA SELATAN
        Desa/Kelurahan: [5103051005] TANJUNG BENOA
        Alamat: TURUS LUMBUNG LSTR
    """
    print("\n[BLOK I] Men-scroll ke bawah secara dinamis hingga data alamat terlihat...")
    max_swipes = 20
    for swipe_idx in range(1, max_swipes + 1):
        if d(textContains="103.").exists() or d(textContains="Nama pada ID Pelanggan").exists():
            print(f"[BLOK I] Teks alamat ditemukan di layar (pemeriksaan ke-{swipe_idx}).")
            break
        
        print(f"[BLOK I] Performing dynamic swipe {swipe_idx} (540, 800 -> 540, 155)...")
        try:
            swipe_aman(540, 700, 540, 400, duration=0.1)
            time.sleep(0.2)
        except Exception as scroll_err:
            print(f"[WARNING] Gagal swipe ke bawah pada percobaan {swipe_idx}: {scroll_err}")
            break
    time.sleep(0.2)

    # Sistem pencarian & verifikasi label "a. Provinsi"
    def cek_label_provinsi():
        patterns = ["a. Provinsi", "a.  Provinsi", "a.Provinsi", "Provinsi", "provinsi"]
        for p in patterns:
            if d(textContains=p).exists():
                return True
        try:
            if d.xpath("//*[contains(@text, 'Provinsi') or contains(@text, 'provinsi')]").exists:
                return True
        except Exception:
            pass
        return False

    if not cek_label_provinsi():
        print("[BLOK I] Label 'a. Provinsi' tidak terdeteksi. Men-scroll ke atas (max 10x)...")
        for swipe_up_idx in range(1, 11):
            if cek_label_provinsi():
                print(f"[BLOK I] Label 'a. Provinsi' ditemukan di layar (pemeriksaan ke-{swipe_up_idx}).")
                break
            try:
                swipe_aman(540, 400, 540, 600, duration=0.1)
                time.sleep(0.2)
            except Exception as scroll_up_err:
                print(f"[WARNING] Gagal swipe ke atas pada percobaan {swipe_up_idx}: {scroll_up_err}")
                break
        time.sleep(0.2)

    # === AMBIL DATA ALAMAT & SIMPAN KE temp_alamat.txt ===
    print("[BLOK I] Mengambil data alamat dari screen...")
    info_alamat = {
        "Provinsi": "",
        "Kabupaten": "",
        "Kecamatan": "",
        "Desa/Kelurahan": "",
        "Alamat": ""
    }
    
    mapping_keys = [
        ("Provinsi", ["a. Provinsi", "Provinsi"]),
        ("Kabupaten", ["b. Kabupaten/Kota", "Kabupaten/Kota", "Kabupaten"]),
        ("Kecamatan", ["c. Kecamatan", "Kecamatan"]),
        ("Desa/Kelurahan", ["d. Desa/Kelurahan", "Desa/Kelurahan", "Desa"]),
        ("Alamat", ["e. Alamat", "Alamat"])
    ]

    for key, patterns in mapping_keys:
        # 1. Coba cari elemen EditText/TextView via label pattern
        for p in patterns:
            el = d(textContains=p)
            if el.exists():
                txt = el.info.get('text', '').strip()
                if ":" in txt and len(txt) > len(p) + 2:
                    info_alamat[key] = f"{key}: {txt.split(':', 1)[1].strip()}"
                    break
                else:
                    # Cari sibling/down EditText (sesuai dump.xml: <node text="..." class="android.widget.EditText" />)
                    sibling = el.sibling(className="android.widget.EditText")
                    if not sibling.exists():
                        sibling = el.sibling(className="android.widget.TextView")
                    if not sibling.exists():
                        sibling = el.down(className="android.widget.EditText")
                    if not sibling.exists():
                        sibling = el.down(className="android.widget.TextView")

                    if sibling.exists():
                        val = sibling.info.get('text', '').strip()
                        if val and not any(val.startswith(x) for x in ["a. ", "b. ", "c. ", "d. ", "e. "]):
                            info_alamat[key] = f"{key}: {val}"
                            break

        # 2. Jika belum ditemukan, coba via XPath
        if not info_alamat[key]:
            try:
                for p in patterns:
                    xpath_el = d.xpath(f"//*[contains(@text, '{p}')]")
                    if xpath_el.exists:
                        parent_siblings = d.xpath(f"//*[contains(@text, '{p}')]/..//*[contains(@class, 'EditText') or contains(@class, 'TextView')]").all()
                        for sib in parent_siblings:
                            sib_text = sib.text.strip()
                            if sib_text and not any(p_check in sib_text for p_check in ["Provinsi", "Kabupaten", "Kecamatan", "Desa", "Alamat"]) and not any(sib_text.startswith(x) for x in ["a. ", "b. ", "c. ", "d. ", "e. "]):
                                info_alamat[key] = f"{key}: {sib_text}"
                                break
                        if info_alamat[key]:
                            break
            except Exception as e:
                print(f"[WARNING] Gagal mencari via XPath untuk key '{key}': {e}")

        # Fallback jika tidak ditemukan
        if not info_alamat[key]:
            info_alamat[key] = f"{key}: (tidak ditemukan)"

    # Normalisasi nama Desa/Kelurahan jika ada variasi Padang Sambian
    if info_alamat.get("Desa/Kelurahan"):
        info_alamat["Desa/Kelurahan"] = normalisasi_nama_desa(info_alamat["Desa/Kelurahan"])

    # Validasi jika data alamat server kosong / hanya kurung siku "[]"
    alamat_kosong = False
    if d(text="[]").exists() or d(text="[ ]").exists():
        alamat_kosong = True
    else:
        nilai_bersih = []
        for k, v in info_alamat.items():
            clean_v = v.split(":", 1)[1].strip() if ":" in v else v.strip()
            nilai_bersih.append(clean_v)
        if any(v in ["[]", "[ ]"] for v in nilai_bersih) or all(v in ["", "[]", "[ ]", "(tidak ditemukan)"] for v in nilai_bersih):
            alamat_kosong = True

    if alamat_kosong:
        print(f"[BLOK I] Data alamat server kosong (berisi '[]') untuk IDPEL {idpel}.")

    # Simpan ke file file_output (temp_alamat.txt)
    try:
        with open(file_output, "w", encoding="utf-8") as f_temp:
            for key in ["Provinsi", "Kabupaten", "Kecamatan", "Desa/Kelurahan", "Alamat"]:
                val = info_alamat[key]
                if not val.startswith(key):
                    val = f"{key}: {val}"
                f_temp.write(val + "\n")
        print(f"[BLOK I] Data alamat berhasil disimpan ke '{file_output}':")
        for key, val in info_alamat.items():
            print(f"  - {val}")
    except Exception as write_err:
        print(f"[WARNING] Gagal menulis ke '{file_output}': {write_err}")

    return info_alamat

def isi_nik_penghuni(nik):
    """Wrapper universal untuk pengisian NIK (202. NIK penghuni)"""
    return input_textbox("202. NIK penghuni", nik, bounds_fallback=(540, 1156))


def ketuk_sidebar_toggle():
    """
    Mengetuk tombol 'sidebar-toggle' (burger icon pada toolbar Fasih):
    Node: <node text="sidebar-toggle" class="android.widget.Button" bounds="[24,114][120,210]" />
    """
    return ketuk("sidebar-toggle")


def cek_dan_ketuk_plus_minus(nometer="", idpel="", sleep_after=SLEEP_SHORT, bounds_str="[48,1248][435,1482]"):
    """
    Fungsi universal untuk mengecek dan mengetuk simbol plus (+) atau minus (-) pada item Fasih:
    - Referensi mengecek dan mengetuk simbol plus (+): dump.xml (L87-L99)
      Node: <node text="+ 86257964485" class="android.view.View" clickable="true" bounds="[48,1248][435,1482]" />
    - Referensi mengecek simbol minus (-): dump copy.xml (L55-L101)
      Node: <node text="- 86257964485" class="android.view.View" clickable="true" bounds="[48,513][435,747]" />

    Logika:
    1. Memeriksa apakah simbol minus (-) terdeteksi (contoh: '- 86257964485'), yang menandakan baris detail sudah terbuka -> Melewati ketuk.
    2. Jika belum terbuka, mencari dan mengetuk simbol plus (+) (contoh: '+ 86257964485') untuk membuka baris detail.
    3. Jika selector teks belum berhasil, menggunakan selector XPath bounds '[48,1248][435,1482]' atau koordinat (241, 1365).
    """
    val_str = str(nometer).strip() if (nometer and str(nometer).strip()) else str(idpel).strip()
    target_plus_text = f"+ {val_str}" if val_str else "+"
    target_minus_text = f"- {val_str}" if val_str else "-"

    # 1. Cek apakah simbol minus (-) terdeteksi (Referensi dump copy.xml L55-L101)
    is_minus_detected = False
    try:
        if (check_exists(d(textContains=target_minus_text)) or 
            check_exists(d(descriptionContains=target_minus_text)) or
            check_exists(d.xpath(f"//*[contains(@text, '{target_minus_text}') or contains(@content-desc, '{target_minus_text}')]"))):
            is_minus_detected = True
        elif val_str and (check_exists(d(textContains=f"- {val_str}")) or check_exists(d(descriptionContains=f"- {val_str}"))):
            is_minus_detected = True
        elif not val_str and (check_exists(d(textContains="-")) or check_exists(d(descriptionContains="-"))):
            is_minus_detected = True
    except Exception:
        pass

    if is_minus_detected:
        print(f"[PLUS/MINUS] Baris detail sudah terbuka (terdeteksi simbol minus '{target_minus_text}'). Melewati ketuk... (Referensi dump copy.xml L55-L101)")
        return True

    # 2. Jika belum terbuka, ketuk simbol plus (+) (Referensi dump.xml L87-L99)
    print(f"[PLUS/MINUS] Mengetuk simbol plus '{target_plus_text}' / bounds '{bounds_str}' untuk membuka detail... (Referensi dump.xml L87-L99)")
    sukses_plus = False
    try:
        if d(textContains=target_plus_text).exists():
            d(textContains=target_plus_text).click()
            sukses_plus = True
        elif d(text=target_plus_text).exists():
            d(text=target_plus_text).click()
            sukses_plus = True
        elif d.xpath(f"//*[contains(@text, '{target_plus_text}') or contains(@content-desc, '{target_plus_text}')]").exists:
            d.xpath(f"//*[contains(@text, '{target_plus_text}') or contains(@content-desc, '{target_plus_text}')]").click()
            sukses_plus = True
        elif d.xpath("//*[starts-with(@text, '+') or starts-with(@content-desc, '+')]").exists:
            d.xpath("//*[starts-with(@text, '+') or starts-with(@content-desc, '+')]").click()
            sukses_plus = True
        elif bounds_str and d.xpath(f"//*[@bounds='{bounds_str}']").exists:
            d.xpath(f"//*[@bounds='{bounds_str}']").click()
            sukses_plus = True
    except Exception:
        pass

    # 3. Fallback jika selector teks & XPath bounds tidak berhasil
    if not sukses_plus:
        # Fallback koordinat dari bounds [48,1248][435,1482] -> cx = 241, cy = 1365
        print(f"[PLUS/MINUS] Fallback mengetuk koordinat bounds {bounds_str} -> (241, 1365)...")
        try:
            d.click(241, 1365)
            sukses_plus = True
        except Exception:
            target_fallback = val_str if val_str else "+"
            print(f"[PLUS/MINUS] Fallback mengetuk via identifier '{target_fallback}'...")
            sukses_plus = ketuk(target_fallback, sleep_after=sleep_after)

    if sleep_after > 0:
        time.sleep(sleep_after)

    return sukses_plus


def ketuk_plus(nometer="", idpel=""):
    """Wrapper ketuk untuk '+' / cek_dan_ketuk_plus_minus()"""
    return cek_dan_ketuk_plus_minus(nometer=nometer, idpel=idpel)


def ketuk_aksi():
    """Wrapper ketuk untuk 'Aksi'"""
    return ketuk("Aksi")

def konfirmasi_stop_atau_lanjut(pesan="Tekan ENTER / Y untuk melanjutkan, atau ketik T untuk menghentikan proses"):
    """
    Fungsi interaktif untuk menghentikan atau melanjutkan proses bot:
    Returns:
        True  -> Pengguna memilih melanjutkan proses
        False -> Pengguna memilih menghentikan / mengakhiri proses
    """
    print(f"\n==================================================")
    print(f"[KONFIRMASI PROSES] {pesan}")
    print(f"==================================================")
    print("  [Y / Enter] : Melanjutkan proses ke data berikutnya")
    print("  [T / Stop]  : Mengakhiri / menghentikan seluruh proses bot")
    
    try:
        jawaban = input(">> Pilihan Anda (Y/t): ").strip().lower()
        if jawaban in ["t", "n", "stop", "exit", "q", "keluar", "no"]:
            print("[STOP] Pengguna memilih untuk MENGHENTIKAN proses.")
            return False
        print("[LANJUT] Pengguna memilih MELANJUTKAN proses...")
        return True
    except (KeyboardInterrupt, EOFError):
        print("\n[STOP] Interupsi terdeteksi. Mengakhiri proses.")
        return False


def baca_temp_alamat():
    """Membaca data alamat dari file temp_alamat.txt yang dihasilkan oleh ambil_data_alamat()."""
    data = {
        "Provinsi": "[51] BALI",
        "Kabupaten": "[5103] KAB. BADUNG",
        "Kecamatan": "[510306] KUTA UTARA",
        "Desa/Kelurahan": "[5103061002] KEROBOKAN",
        "Alamat": "JL. GN. MANDALA"
    }

    file_path = "temp_alamat.txt"
    if os.path.exists(file_path):
        print(f"[EXCEL/TXT] Membaca data alamat dari {file_path}...")
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                for line in f:
                    if ":" in line:
                        key, val = line.split(":", 1)
                        key = key.strip()
                        val = val.strip()
                        if "Desa" in key or "Kelurahan" in key:
                            data["Desa/Kelurahan"] = normalisasi_nama_desa(val)
                        elif "Kabupaten" in key or "Kota" in key:
                            data["Kabupaten"] = val
                        elif key in data:
                            data[key] = val
            print("[SUKSES] Berhasil membaca data alamat:")
            for k, v in data.items():
                print(f"  - {k}: '{v}'")
        except Exception as e:
            print(f"[WARNING] Gagal membaca {file_path}: {e}. Menggunakan default.")
    else:
        print(f"[WARNING] File {file_path} tidak ditemukan. Menggunakan nilai default.")

    return data


def ekstrak_nama_saja(teks):
    """Hilangkan bagian '[...]' dari teks (misal: '[51] BALI' -> 'BALI')."""
    import re as _re
    clean = _re.sub(r'\[.*?\]', '', teks)
    return clean.strip()


def klik_opsi_dropdown(d_dev, teks_pilihan):
    """Klik opsi dropdown yang cocok dengan teks_pilihan dari ListView/Dialog."""
    nama_saja = ekstrak_nama_saja(teks_pilihan)
    print(f"[DROPDOWN] Mencari opsi untuk clean name: '{nama_saja}' (dari '{teks_pilihan}')...")

    for attempt in range(8):
        try:
            xpath_list = [
                f"//android.widget.ListView//*[contains(@text, '{teks_pilihan}')]",
                f"//android.widget.ListView//*[contains(@text, '{nama_saja}') and contains(@text, '[')]",
                f"//android.widget.ListView//*[contains(@text, '{nama_saja}')]",
                f"//android.app.Dialog//*[contains(@text, '{teks_pilihan}')]",
                f"//android.app.Dialog//*[contains(@text, '{nama_saja}') and contains(@text, '[')]",
                f"//android.app.Dialog//*[contains(@text, '{nama_saja}')]",
            ]

            for xp in xpath_list:
                xpath_el = d_dev.xpath(xp)
                if xpath_el.exists:
                    matching_elements = xpath_el.all()
                    target_el = None

                    for el in matching_elements:
                        txt = el.text or ""
                        if "EditText" in (el.info.get("className", "") or ""):
                            continue
                        if txt == teks_pilihan or txt.endswith(f"] {nama_saja}") or txt.endswith(f" {nama_saja}") or txt == nama_saja:
                            target_el = el
                            break

                    if not target_el:
                        for el in matching_elements:
                            if "EditText" not in (el.info.get("className", "") or ""):
                                target_el = el
                                break

                    if target_el:
                        print(f"[DROPDOWN] Menemukan opsi via XPath ('{xp}'): '{target_el.text}' -> Mengklik...")
                        target_el.click()
                        time.sleep(0.4)
                        return True
        except Exception as xpath_err:
            print(f"[DROPDOWN] [WARNING] Gagal mengevaluasi XPath: {xpath_err}")

        time.sleep(0.5)

    print(f"[DROPDOWN] [WARNING] Opsi '{teks_pilihan}' tidak ditemukan di layar. Menutup overlay.")
    d_dev.click(540, 300)
    time.sleep(0.5)
    return False


def check_wajib_diisi_field(label_text):
    """
    Memeriksa apakah teks 'Wajib diisi' muncul KHUSUS pada container field tertentu (misal: 'Provinsi', 'Kabupaten/Kota', 'Kecamatan', 'Desa/Kelurahan').
    """
    try:
        if d.xpath(f"//*[contains(@text, '{label_text}')]/..//*[contains(@text, 'Wajib diisi')]").exists:
            return True
        lbl = d(textContains=label_text)
        if lbl.exists:
            p_top = lbl.info.get('bounds', {}).get('top', 0)
            p_bottom_limit = p_top + 250
            w_els = d(textContains="Wajib diisi").all() if hasattr(d(textContains="Wajib diisi"), 'all') else []
            for w_el in w_els:
                b_str = w_el.attrib.get('bounds', '')
                import re
                match_y = re.search(r'\[\d+,(\d+)\]\[\d+,\d+\]', b_str)
                if match_y and (p_top <= int(match_y.group(1)) <= p_bottom_limit):
                    return True
    except Exception as w_err:
        print(f"[WARNING] Error pengecekan spesifik Wajib diisi {label_text}: {w_err}")
    return False


def isi_blok_iii(alamat_val=""):
    """
    Mengisi seluruh field BLOK III:
      - a. Provinsi
      - b. Kabupaten/Kota
      - c. Kecamatan
      - d. Desa/Kelurahan
      - e. Alamat
    Data wilayah dibaca dari temp_alamat.txt (hasil ambil_data_alamat()).
    Parameter:
        alamat_val: Nilai alamat spesifik (e. Alamat) dari Excel, default kosong.
    """
    alamat_data = baca_temp_alamat()

    provinsi = alamat_data["Provinsi"]
    kabupaten = alamat_data["Kabupaten"]
    kecamatan = alamat_data["Kecamatan"]
    desa = alamat_data["Desa/Kelurahan"]
    if not alamat_val:
        alamat_val = alamat_data.get("Alamat", "")

    # 1. Cek text "BLOK III"
    print("\n[BLOK III] [STEP 1] Memeriksa teks 'BLOK III'...")
    blok3_header = d(textContains="BLOK III")
    if blok3_header.exists(timeout=20):
        print("[BLOK III] [SUKSES] Berada di halaman/bagian 'BLOK III'")
    else:
        print("[BLOK III] [WARNING] Header 'BLOK III' tidak ditemukan di layar saat ini.")

    # Force refresh (swipe up & down) lalu ketuk koordinat statis (996, 588)
    print("[BLOK III] Melakukan force refresh (swipe up & down) dan mengetuk koordinat statis (996, 588)...")
    d.swipe(540, 1200, 540, 600, duration=0.05)
    time.sleep(0.1)
    d.swipe(540, 600, 540, 1200, duration=0.2)
    time.sleep(0.1)
    d.click(996, 588)
    time.sleep(SLEEP_MEDIUM)

    # 2. Label "a. Provinsi"
    print("\n[BLOK III] [STEP 2] Mengisi 'a. Provinsi'...")
    label_a = d(textContains="Provinsi")
    input_a = None
    for attempt in range(40):
        if label_a.exists():
            input_a = label_a.down(className="android.widget.EditText")
            if input_a and input_a.exists():
                break
        time.sleep(0.5)

    if not (input_a and input_a.exists()):
        print("[BLOK III] [RETRY NAVIGASI] Field 'a. Provinsi' tidak terdeteksi. Mencoba berpindah kembali ke BLOK III via sidebar...")
        for nav_retry in range(1, 4):
            ketuk_sidebar_toggle()
            time.sleep(SLEEP_SHORT)
            pilih_blok("III")
            time.sleep(SLEEP_MEDIUM)
            label_a = d(textContains="Provinsi")
            if label_a.exists():
                input_a = label_a.down(className="android.widget.EditText")
                if input_a and input_a.exists():
                    print(f"[BLOK III] [SUKSES NAVIGASI] Berhasil masuk ke BLOK III pada percobaan ke-{nav_retry}.")
                    break

    if input_a and input_a.exists():
        provinsi_clean = ekstrak_nama_saja(provinsi)
        for prov_attempt in range(1, 3):
            input_a.set_text(str(provinsi_clean))
            print(f"[BLOK III] Berhasil mengisi Provinsi: '{provinsi_clean}' (Percobaan ke-{prov_attempt})")
            time.sleep(0.5)
            klik_opsi_dropdown(d, provinsi)
            time.sleep(0.5)

            if check_wajib_diisi_field("Provinsi"):
                print(f"[BLOK III] [RETRY] Terdeteksi teks 'Wajib diisi' KHUSUS pada Provinsi (percobaan ke-{prov_attempt}). Mengulangi pengisian Provinsi...")
                time.sleep(0.5)
            else:
                print(f"[BLOK III] Provinsi '{provinsi_clean}' berhasil dipilih.")
                break
    else:
        raise Exception("Input text box untuk 'a. Provinsi' tidak ditemukan.")

    # 3. Label "b. Kabupaten/Kota"
    print("\n[BLOK III] [STEP 3] Mengisi 'b. Kabupaten/Kota'...")
    label_b = d(textContains="Kabupaten/Kota")
    input_b = None
    for attempt in range(10):
        if label_b.exists():
            input_b = label_b.down(className="android.widget.EditText")
            if input_b and input_b.exists():
                break
        time.sleep(0.5)

    if input_b and input_b.exists():
        kabupaten_clean = ekstrak_nama_saja(kabupaten)
        for kab_attempt in range(1, 3):
            input_b.set_text(str(kabupaten_clean))
            print(f"[BLOK III] Berhasil mengisi Kabupaten/Kota: '{kabupaten_clean}' (Percobaan ke-{kab_attempt})")
            time.sleep(0.5)
            klik_opsi_dropdown(d, kabupaten)
            time.sleep(0.5)

            if check_wajib_diisi_field("Kabupaten/Kota"):
                print(f"[BLOK III] [RETRY] Terdeteksi teks 'Wajib diisi' KHUSUS pada Kabupaten/Kota (percobaan ke-{kab_attempt}). Mengulangi pengisian Kabupaten/Kota...")
                time.sleep(0.5)
            else:
                print(f"[BLOK III] Kabupaten/Kota '{kabupaten_clean}' berhasil dipilih.")
                break
    else:
        raise Exception("Input text box untuk 'b. Kabupaten/Kota' tidak ditemukan.")

    # 4. Label "c. Kecamatan"
    print("\n[BLOK III] [STEP 4] Mengisi 'c. Kecamatan'...")
    print("[BLOK III] Melakukan swipe ke bawah agar 'c. Kecamatan' terlihat penuh...")
    d.swipe(540, 1200, 540, 1000, duration=0.2)
    time.sleep(0.5)

    label_c = d(textContains="Kecamatan")
    input_c = None
    for attempt in range(10):
        if label_c.exists():
            input_c = label_c.down(className="android.widget.EditText")
            if input_c and input_c.exists():
                break
        time.sleep(0.5)

    if input_c and input_c.exists():
        kecamatan_clean = ekstrak_nama_saja(kecamatan)
        for kec_attempt in range(1, 3):
            input_c.set_text(str(kecamatan_clean))
            print(f"[BLOK III] Berhasil mengisi Kecamatan: '{kecamatan_clean}' (Percobaan ke-{kec_attempt})")
            time.sleep(0.5)
            klik_opsi_dropdown(d, kecamatan)
            time.sleep(0.5)

            if check_wajib_diisi_field("Kecamatan"):
                print(f"[BLOK III] [RETRY] Terdeteksi teks 'Wajib diisi' KHUSUS pada Kecamatan (percobaan ke-{kec_attempt}). Mengulangi pengisian Kecamatan...")
                time.sleep(0.5)
            else:
                print(f"[BLOK III] Kecamatan '{kecamatan_clean}' berhasil dipilih.")
                break
    else:
        raise Exception("Input text box untuk 'c. Kecamatan' tidak ditemukan.")

    # 5. Label "d. Desa/Kelurahan"
    print("\n[BLOK III] [STEP 5] Mengisi 'd. Desa/Kelurahan'...")
    print("[BLOK III] Melakukan swipe ke bawah agar 'd. Desa/Kelurahan' terlihat penuh...")
    d.swipe(540, 1200, 540, 600, duration=0.2)
    time.sleep(0.5)

    label_d = d(textContains="Desa/Kelurahan")
    input_d = None
    for attempt in range(10):
        if label_d.exists():
            input_d = label_d.down(className="android.widget.EditText")
            if input_d and input_d.exists():
                break
        time.sleep(0.5)

    if input_d and input_d.exists():
        desa_clean = ekstrak_nama_saja(desa)
        desa_sukses = False
        max_attempts_desa = max(1, len(desa_clean) - 3 + 1)

        for attempt_desa in range(max_attempts_desa):
            current_input = desa_clean if attempt_desa == 0 else desa_clean[:-attempt_desa]

            input_d.clear_text()
            time.sleep(0.5)
            input_d.set_text(str(current_input))
            print(f"[BLOK III] Mengisi Desa/Kelurahan: '{current_input}' (Percobaan {attempt_desa + 1}/{max_attempts_desa})")
            time.sleep(0.5)

            if d(textContains="Pilihan tidak ditemukan").exists():
                print(f"[BLOK III] Terdeteksi 'Pilihan tidak ditemukan' untuk '{current_input}'. Menutup overlay, menghapus input, dan mengulang...")
                d.click(540, 300)
                time.sleep(0.5)
                input_d.clear_text()
                time.sleep(0.5)
                continue

            for retry_wajib in range(1, 3):
                if retry_wajib > 1:
                    print(f"[BLOK III] [RETRY DESA] Mengulangi pengisian Desa/Kelurahan '{current_input}' karena terdeteksi 'Wajib diisi'...")
                    input_d.clear_text()
                    time.sleep(0.5)
                    input_d.set_text(str(current_input))
                    time.sleep(0.5)

                if klik_opsi_dropdown(d, current_input) or klik_opsi_dropdown(d, desa):
                    time.sleep(0.5)
                    if check_wajib_diisi_field("Desa/Kelurahan"):
                        print(f"[BLOK III] [RETRY] Terdeteksi 'Wajib diisi' KHUSUS pada Desa/Kelurahan untuk '{current_input}' (Percobaan {retry_wajib})...")
                        time.sleep(0.5)
                    else:
                        print(f"[BLOK III] Berhasil memilih opsi Desa/Kelurahan untuk '{current_input}'.")
                        desa_sukses = True
                        break
                else:
                    print(f"[BLOK III] Opsi Desa/Kelurahan '{current_input}' tidak berhasil diklik.")
                    break

            if desa_sukses:
                break

        if not desa_sukses:
            raise Exception(f"Error : tidak menemukan Desa/Kelurahan {desa_clean}")
    else:
        raise Exception("Input text box untuk 'd. Desa/Kelurahan' tidak ditemukan.")

    # 6. Label "e. Alamat"
    print("\n[BLOK III] [STEP 6] Mengisi 'e. Alamat'...")
    d.click(540, 300)
    time.sleep(0.5)
    print("[BLOK III] Melakukan swipe ke bawah agar 'e. Alamat' terlihat penuh...")
    d.swipe(540, 1200, 540, 650, duration=0.2)
    time.sleep(0.5)

    label_alamat = d(textContains="Alamat")
    input_alamat = None
    for attempt in range(10):
        if label_alamat.exists():
            input_alamat = label_alamat.down(className="android.widget.EditText")
            if input_alamat and input_alamat.exists():
                break
        time.sleep(0.5)

    if input_alamat and input_alamat.exists():
        input_alamat.set_text(str(alamat_val))
        print(f"[BLOK III] Berhasil mengisi Alamat: '{alamat_val}'")
        time.sleep(0.5)
    else:
        raise Exception("Input text box untuk 'e. Alamat' tidak ditemukan.")



def ketuk_tombol_increment():
    """Mencari dan mengetuk tombol 'Increment' di BLOK III jika jumlah keluarga belum terisi."""
    print("\n[BLOK III] [STEP 7] Memeriksa & mencari tombol 'Increment' di BLOK III...")
    
    increment_btn = d(text="Increment")
    if not increment_btn.exists():
        increment_btn = d(textContains="Increment")
        
    if not increment_btn.exists():
        print("[BLOK III] [STEP 7] Melakukan scroll mencari tombol Increment / label 302...")
        try:
            d(scrollable=True).scroll.to(text="Increment")
        except Exception:
            try:
                d(scrollable=True).scroll.to(textContains="Increment")
            except Exception:
                pass
                
    # Fallback scroll manual jika scroll.to tidak menemukan/gagal
    if not increment_btn.exists():
        for _ in range(3):
            d.swipe(540, 1200, 540, 600, duration=0.2)
            time.sleep(0.5)
            if d(text="Increment").exists():
                increment_btn = d(text="Increment")
                break
            elif d(textContains="Increment").exists():
                increment_btn = d(textContains="Increment")
                break

    # Setelah berada di posisi tombol Increment/302, periksa nilai textbox 302
    label_302 = d(textContains="Berapa jumlah keluarga")
    if not label_302.exists():
        label_302 = d(textContains="302")

    input_302 = None
    if label_302.exists():
        input_302 = label_302.down(className="android.widget.EditText")

    if not input_302 or not input_302.exists():
        if increment_btn and increment_btn.exists():
            input_302 = increment_btn.left(className="android.widget.EditText")

    if input_302 and input_302.exists():
        val_302 = (input_302.get_text() or "").strip()
        print(f"[BLOK III] Status textbox 302 (jumlah keluarga): '{val_302}'")
        if val_302 == "1" or (val_302 != "" and val_302 != "0"):
            print(f"[BLOK III] [SKIP] Textbox 302 (jumlah keluarga) sudah berisi '{val_302}'. Melewati pengetukan tombol 'Increment'.")
            return True

    if increment_btn and increment_btn.exists():
        increment_btn.click()
        print("[BLOK III] Berhasil mengetuk tombol 'Increment'")
        time.sleep(1)
        return True
    else:
        raise Exception("Tombol 'Increment' tidak ditemukan.")


def ketuk_ok_submit_diproses(max_attempts=5):
    """
    Memeriksa dan mengetuk tombol 'OK' pada modal 'Submit diproses'.
    Jika setelah diketuk modal masih tetap muncul ('Submit diproses' masih ada di layar),
    akan mengulang pengetukan hingga `max_attempts` kali (default 5).
    Jika setelah 5x masih tidak bisa, menggunakan fallback pengetukan statis pada koordinat bounds (528, 1691).
    """
    print("\n[SUBMIT DIPROSES] Memeriksa modal 'Submit diproses'...")
    
    # Koordinat tengah tombol OK dari bounds [137,1639][919,1744] (id: btn_submit_progress_close)
    ok_bounds_x, ok_bounds_y = 528, 1691
    
    for attempt in range(1, max_attempts + 1):
        is_submit_modal = (
            check_exists(d(textContains="Submit diproses")) or
            check_exists(d(resourceId="id.go.bpsfasih:id/tv_submit_progress_title")) or
            check_exists(d(resourceId="id.go.bpsfasih:id/btn_submit_progress_close")) or
            check_exists(d(text="OK"))
        )
        
        if not is_submit_modal:
            print("[SUBMIT DIPROSES] Modal 'Submit diproses' tidak terdeteksi (sudah tertutup).")
            return True

        print(f"[SUBMIT DIPROSES] Terdeteksi modal 'Submit diproses'. Percobaan ketuk 'OK' ke-{attempt}/{max_attempts}...")
        
        clicked = False
        try:
            btn_close = d(resourceId="id.go.bpsfasih:id/btn_submit_progress_close")
            if btn_close.exists():
                btn_close.click()
                clicked = True
            elif d(text="OK").exists():
                d(text="OK").click()
                clicked = True
            else:
                clicked = ketuk("OK", sleep_after=SLEEP_SHORT)
        except Exception as e:
            print(f"[WARNING] Klik tombol OK via elemen gagal: {e}")

        if not clicked:
            print(f"[SUBMIT DIPROSES] Klik via elemen gagal. Mengetuk koordinat statis bounds ({ok_bounds_x}, {ok_bounds_y})...")
            d.click(ok_bounds_x, ok_bounds_y)

        time.sleep(2)

        still_exists = (
            check_exists(d(textContains="Submit diproses")) or
            check_exists(d(resourceId="id.go.bpsfasih:id/tv_submit_progress_title"))
        )
        if not still_exists:
            print(f"[SUBMIT DIPROSES] [SUKSES] Modal 'Submit diproses' berhasil ditutup pada percobaan ke-{attempt}.")
            return True
        else:
            print(f"[SUBMIT DIPROSES] [RETRY] Modal 'Submit diproses' masih ada di layar setelah percobaan ke-{attempt}.")

    # Fallback jika 5x percobaan elemen tidak berhasil menutup modal
    print(f"[SUBMIT DIPROSES] [FALLBACK STATIS] Sudah {max_attempts}x percobaan dan modal masih ada. Mengetuk koordinat statis bounds OK ({ok_bounds_x}, {ok_bounds_y})...")
    d.click(ok_bounds_x, ok_bounds_y)
    time.sleep(2)
    return True



def proses_update_reject_nik():
    """Fungsi utama memproses list data reject untuk update NIK"""
    data_reject = baca_data_reject(EXCEL_FILE)
    if not data_reject:
        print("[HALT] Tidak ada data reject yang dapat diproses.")
        return

    total = len(data_reject)
    print(f"\n==================================================")
    print(f"       MEMULAI PROSES INPUT DATA REJECT ({total} DATA)    ")
    print(f"==================================================")

    for idx, item in enumerate(data_reject, start=1):
        idpel = item["idpel"]
        nometer = item.get("nometer", "")
        nik = item["nik"]
        nama = item["nama"]
        row = item["row"]

        # Validasi NAMA: Harus alphabet saja (a-z A-Z dan spasi) jika NAMA tidak kosong. Skip jika mengandung karakter selain alphabet.
        if nama and not all(c.isalpha() or c.isspace() for c in nama):
            print(f"[SKIP] Baris {row} | IDPEL {idpel} dilewati karena NAMA '{nama}' mengandung karakter selain alphabet a-z A-Z.")
            simpan_status_excel(row, "Error : nama tidak murni alphabeth")
            continue

        # Loop percobaan pengulangan untuk baris IDPEL yang sama (maksimal 3x)
        sukses_baris = False
        for row_attempt in range(1, 4):
            print(f"\n--------------------------------------------------")
            print(f"[{idx}/{total}] Memproses Baris {row} (Percobaan {row_attempt}) | IDPEL: {idpel} | NOMETER: {nometer} | NAMA: {nama} | NIK: {nik}")
            print(f"--------------------------------------------------")

            #1 swipe keatas statis untuk memastikan halaman berada di paling atas
            loop_swipe_statis(delta_y=-700, loop=1)
            loop_swipe_statis(delta_y=800, loop=1)
            time.sleep(SLEEP_SHORT)

            #2. Tunggu halaman Search (jika timeout 15s -> akhiri seluruh proses bot)
            if not tunggu_loading("Search", timeout=15):
                print("[HALT] Halaman 'Search' tidak ditemukan setelah 15 detik. Mengakhiri seluruh proses bot.")
                return

            #2 Ketuk search box -> masukkan IDPEL -> press Enter
            sukses_l1 = cari_dan_ketuk_search_box(idpel)
            if not sukses_l1:
                print(f"[SKIP] Gagal di #1 (Search IDPEL {idpel}). Melanjutkan ke percobaan berikutnya...")
                continue

            #2.5 Cek jika terdeteksi 'No matching records found'
            is_no_matching = False
            try:
                if (check_exists(d(textContains="No matching records")) or 
                    check_exists(d(descriptionContains="No matching records")) or
                    check_exists(d.xpath("//*[contains(@text, 'No matching records') or contains(@content-desc, 'No matching records')]"))):
                    is_no_matching = True
            except Exception:
                pass

            if is_no_matching:
                print(f"[NOT FOUND] Terdeteksi 'No matching records found' untuk IDPEL {idpel}. Menyimpan status & lanjut ke baris berikutnya...")
                simpan_status_excel(row, "DATA TIDAK DITEMUKAN")
                sukses_baris = True
                break

            #3-8 Eksekusi buka detail form
            #3 Cek dan ketuk simbol plus/minus (+ / -) untuk membuka detail (Referensi dump.xml L87-L99 & dump copy.xml L55-L101)
            cek_dan_ketuk_plus_minus(nometer=nometer, idpel=idpel, sleep_after=SLEEP_SHORT)

            #4 Loop swipe statis ke bawah
            loop_swipe_statis(delta_y=-700, loop=1)
            time.sleep(SLEEP_SHORT)

            #4.5 Cek status REJECTED vs SUBMITTED / SUBMIT (PENDING)
            is_submitted_status = False
            is_rejected_status = False
            try:
                if (check_exists(d(textContains="SUBMITTED")) or 
                    check_exists(d(descriptionContains="SUBMITTED")) or
                    check_exists(d(textContains="SUBMIT (PENDING)")) or 
                    check_exists(d(descriptionContains="SUBMIT (PENDING)")) or
                    check_exists(d.xpath("//*[contains(@text, 'SUBMITTED') or contains(@content-desc, 'SUBMITTED') or contains(@text, 'SUBMIT (PENDING)') or contains(@content-desc, 'SUBMIT (PENDING)')]"))):
                    is_submitted_status = True

                if (check_exists(d(textContains="REJECTED")) or 
                    check_exists(d(descriptionContains="REJECTED")) or
                    check_exists(d(textContains="Responden menolak")) or 
                    check_exists(d(descriptionContains="Responden menolak")) or
                    check_exists(d.xpath("//*[contains(@text, 'REJECTED') or contains(@content-desc, 'REJECTED') or contains(@text, 'Responden menolak') or contains(@content-desc, 'Responden menolak')]"))):
                    is_rejected_status = True
            except Exception:
                pass

            if is_submitted_status:
                print(f"[STATUS] Terdeteksi status 'SUBMITTED / SUBMIT (PENDING)' pada IDPEL {idpel}. Menyimpan status Excel & kembali ke Search Box...")
                simpan_status_excel(row, "SUKSES DARI AWAL SUDAH SUBMIT")
                sukses_baris = True
                break
            elif is_rejected_status:
                print(f"[STATUS] Terdeteksi status 'REJECTED' pada IDPEL {idpel}. Melanjutkan proses input...")
            else:
                print(f"[STATUS] Melanjutkan ke langkah tombol 'Aksi' untuk IDPEL {idpel}...")

            #5 Ketuk tombol 'Aksi'
            ketuk("Aksi", sleep_after=SLEEP_SHORT)
            time.sleep(SLEEP_MEDIUM)

            #6 Ketuk tombol 'Buka'
            ketuk("BUKA", sleep_after=SLEEP_SHORT)
            time.sleep(SLEEP_MEDIUM)

            #7 ketuk tombol 'YA'
            ketuk("YA", sleep_after=SLEEP_LONG)
            time.sleep(SLEEP_LONG)

            # Cek jika muncul text/dialog "Perhatian"
            if check_exists(d(text="Perhatian")) or check_exists(d(textContains="Perhatian")) or check_exists(d(resourceId="id.go.bpsfasih:id/judul_bottomDialog")):
                print("[DIALOG] Terdeteksi text/dialog 'Perhatian'. Mengetuk 'DOWNLOAD SEKARANG'...")
                if check_exists(d(resourceId="id.go.bpsfasih:id/rButton_bottomDialog")):
                    d(resourceId="id.go.bpsfasih:id/rButton_bottomDialog").click()
                else:
                    ketuk("DOWNLOAD SEKARANG", sleep_after=SLEEP_SHORT)
                print("[LOADING] Menunggu proses download / loading selesai...")
                tunggu_loading(timeout=30)
                time.sleep(SLEEP_LONG)

                # Mengulangi ketuk 'Aksi' -> 'BUKA' -> 'YA' setelah download selesai
                print("[RETRY] Mengulangi ketuk 'Aksi' -> 'BUKA' -> 'YA' setelah download selesai...")
                ketuk("Aksi", sleep_after=SLEEP_SHORT)
                time.sleep(SLEEP_MEDIUM)
                ketuk("BUKA", sleep_after=SLEEP_SHORT)
                time.sleep(SLEEP_MEDIUM)
                ketuk("YA", sleep_after=SLEEP_SHORT)
                time.sleep(SLEEP_LONG)

            # Scan tombol Kirim terlebih dahulu untuk mengecek posisi tombol "Kirim" menggunakan bounds
            print("[SCAN KIRIM] Memindai posisi tombol 'Kirim' menggunakan bounds...")
            kirim_bounds = None
            try:
                btn_kirim = d(className="android.widget.Button", text="Kirim")
                if not btn_kirim.exists():
                    btn_kirim = d(text="Kirim", clickable=True)
                if not btn_kirim.exists():
                    btn_kirim = d(text="Kirim")

                if check_exists(btn_kirim):
                    info = btn_kirim.info
                    b = info.get("bounds") if isinstance(info, dict) else None
                    if b and isinstance(b, dict):
                        cx = (b.get("left", 0) + b.get("right", 0)) // 2
                        cy = (b.get("top", 0) + b.get("bottom", 0)) // 2
                        kirim_bounds = (cx, cy)
                        print(f"[SCAN KIRIM] Posisi tombol 'Kirim' terdeteksi pada bounds [{b.get('left')},{b.get('top')}][{b.get('right')},{b.get('bottom')}] -> Titik tengah ({cx}, {cy})")
            except Exception as e:
                print(f"[SCAN KIRIM] Gagal memindai bounds tombol Kirim: {e}")

            if kirim_bounds:
                d.click(kirim_bounds[0], kirim_bounds[1])
                time.sleep(SLEEP_SHORT)
            else:
                ketuk("Kirim", sleep_after=SLEEP_SHORT)
                time.sleep(SLEEP_SHORT)

            # Ketuk teks yang mengandung kata "GALAT"
            ketuk("GALAT", exact=False, sleep_after=SLEEP_SHORT)
            time.sleep(SLEEP_SHORT)

            # Cek jika muncul kata "Koordinat lokasi meteran" atau "Foto rumah tampak depan"
            is_galat_koordinat_foto = False
            try:
                if (check_exists(d(textContains="Koordinat lokasi meteran")) or 
                    check_exists(d(descriptionContains="Koordinat lokasi meteran")) or
                    check_exists(d.xpath("//*[contains(@text, 'Koordinat lokasi meteran') or contains(@content-desc, 'Koordinat lokasi meteran')]")) or
                    check_exists(d(textContains="Foto rumah tampak depan")) or
                    check_exists(d(textContains="Foto ruimah tampak depan")) or
                    check_exists(d(descriptionContains="Foto rumah tampak depan")) or
                    check_exists(d(descriptionContains="Foto ruimah tampak depan")) or
                    check_exists(d.xpath("//*[contains(@text, 'tampak depan') or contains(@content-desc, 'tampak depan')]"))):
                    is_galat_koordinat_foto = True
            except Exception:
                pass

            if is_galat_koordinat_foto:
                print(f"[GALAT CHECK] [SKIP] Terdeteksi 'Koordinat lokasi meteran' / 'Foto rumah tampak depan' pada GALAT untuk IDPEL {idpel}. Menyimpan status & berpindah ke baris berikutnya...")
                simpan_status_excel(row, "koordinat & foto tidak ada")
                kembali_ke_daftar_assignment()
                sukses_baris = True
                break

            time.sleep(SLEEP_LONG)
            alamat_dict = ambil_data_alamat(file_output="temp_alamat.txt", idpel=idpel)
            time.sleep(SLEEP_SHORT)

            # Cek jika data alamat mengandung kata 'null' (case-insensitive), '[]', atau 'tidak ditemukan'
            is_alamat_null = False
            is_tidak_ditemukan = False
            for k, v in (alamat_dict or {}).items():
                v_str = str(v)
                v_str_lower = v_str.lower()
                if "tidak ditemukan" in v_str_lower:
                    is_tidak_ditemukan = True
                    print(f"[BLOK I] [WARNING] Terdeteksi kata 'tidak ditemukan' pada field alamat '{k}': '{v}' (IDPEL: {idpel}).")
                elif "null" in v_str_lower or "[]" in v_str:
                    is_alamat_null = True
                    print(f"[BLOK I] [WARNING] Terdeteksi kata 'null' / '[]' pada field alamat '{k}': '{v}' (IDPEL: {idpel}).")

            if is_tidak_ditemukan:
                print(f"[BLOK I] [SKIP] Data alamat untuk IDPEL {idpel} terdeteksi 'tidak ditemukan'. Menyimpan status 'Error : Alamat tidak ditemukan' & berpindah ke baris berikutnya...")
                simpan_status_excel(row, "Error : Alamat tidak ditemukan")
                kembali_ke_daftar_assignment()
                sukses_baris = True
                break

            if is_alamat_null:
                print(f"[BLOK I] [SKIP] Data alamat untuk IDPEL {idpel} tidak valid ('null' / '[]'). Menyimpan status 'Error: Alamat NULL' & berpindah ke baris berikutnya...")
                simpan_status_excel(row, "Error: Alamat NULL")
                kembali_ke_daftar_assignment()
                sukses_baris = True
                break

            loop_swipe_dinamis(delta_y=-400, target_text="105. Koordinat lokasi meteran", duration=0.5)
            time.sleep(SLEEP_LONG)

            # Pengecekan status RadioButton: Ketuk '1. Berhasil didata' HANYA jika belum tercentang
            radio_success = False
            if cek_radio_button_tercentang("1. Berhasil didata", exact=False):
                print("[RADIO CHECK] RadioButton '1. Berhasil didata' sudah tercentang.")
                radio_success = True
            else:
                print("[RADIO CHECK] RadioButton '1. Berhasil didata' BELUM tercentang. Memulai pengetukan...")
                max_radio_attempts = 5
                for r_attempt in range(1, max_radio_attempts + 1):
                    clicked = False

                    # Strategi 1: Kalkulasi koordinat dinamis dari parent/label bounds
                    try:
                        print("strategi 1")
                        label_el = None
                        for pattern in ["1. Berhasil didata", "1.  Berhasil didata", "Berhasil didata"]:
                            elements = d.xpath(f"//*[contains(@text, '{pattern}')]").all()
                            for el in elements:
                                bounds_str = el.attrib.get('bounds')
                                import re
                                pts = [int(x) for x in re.findall(r'\d+', bounds_str)]
                                if len(pts) == 4:
                                    x1, y1, x2, y2 = pts
                                    if (x2 - x1) > 50:
                                        label_el = el
                                        break
                            if label_el:
                                break

                        if label_el:
                            label_bounds_str = label_el.attrib.get('bounds')
                            import re
                            l_pts = [int(x) for x in re.findall(r'\d+', label_bounds_str)]
                            label_left, label_top, label_right, label_bottom = l_pts

                            label_text = label_el.text
                            parent = d.xpath(f"//*[contains(@text, '{label_text}')]/..")

                            if parent.exists:
                                parent_bounds = parent.all()[0].attrib.get('bounds')
                                p_pts = [int(x) for x in re.findall(r'\d+', parent_bounds)]
                                parent_left, parent_top, parent_right, parent_bottom = p_pts

                                click_x = parent_left + (label_left - parent_left) // 2
                                click_y = label_top + (label_bottom - label_top) // 2

                                print(f"[KLIK] Parent Left: {parent_left}, Label Left: {label_left}")
                                print(f"[KLIK] Mengklik koordinat dinamis radio button: ({click_x}, {click_y})")
                                d.click(click_x, click_y)
                                time.sleep(SLEEP_SHORT)
                                clicked = True
                    except Exception as click_err:
                        print(f"[WARNING] Percobaan klik koordinat dinamis gagal: {click_err}")

                    if not clicked:
                        print(f"[WARNING] Strategi 1 klik gagal pada percobaan ke-{r_attempt}.")

                    # Verifikasi apakah sudah tercentang
                    time.sleep(0.2)
                    if cek_radio_button_tercentang("1. Berhasil didata", exact=True):
                        print(f"[RADIO SUCCESS] Berhasil memverifikasi '1. Berhasil didata' tercentang pada percobaan ke-{r_attempt}.")
                        radio_success = True
                        break
                    print(f"[RETRY RADIO] '1. Berhasil didata' belum tercentang (percobaan ke-{r_attempt}/{max_radio_attempts}). Mengulangi...")
                    time.sleep(SLEEP_SHORT)
                    loop_swipe_statis(delta_y=-200, loop=1)
                    time.sleep(SLEEP_SHORT)

            if not radio_success:
                print(f"[RADIO CHECK] [SKIP] RadioButton '1. Berhasil didata' belum tercentang setelah {max_radio_attempts}x percobaan (IDPEL: {idpel}). Menyimpan status Excel & berpindah ke baris berikutnya...")
                simpan_status_excel(row, "Error : RadioButton 1. Berhasil didata tidak tercentang")
                kembali_ke_daftar_assignment()
                sukses_baris = True
                break
            
            

            #13.B BLOK II
            ketuk_sidebar_toggle()
            time.sleep(SLEEP_SHORT)
            pilih_blok("II")
            time.sleep(SLEEP_LONG)

            # Cek halaman apakah sudah di "Blok II" atau belum
            is_blok_ii = (
                check_exists(d(textContains="BLOK II")) or
                check_exists(d(textContains="Blok II")) or
                check_exists(d(descriptionContains="BLOK II")) or
                check_exists(d(descriptionContains="Blok II")) or
                check_exists(d(textContains="201. Nama penghuni")) or
                check_exists(d.xpath("//*[contains(@text, 'BLOK II') or contains(@text, 'Blok II') or contains(@content-desc, 'BLOK II') or contains(@content-desc, 'Blok II') or contains(@text, '201. Nama penghuni')]"))
            )

            if not is_blok_ii:
                print(f"[BLOK II] [SKIP] Halaman tidak berada di 'Blok II' setelah pilih_blok('II') (IDPEL: {idpel}). Menyimpan status Excel & berpindah ke baris berikutnya...")
                simpan_status_excel(row, "Error : Gagal masuk Blok II")
                kembali_ke_daftar_assignment()
                sukses_baris = True
                break

            # Pengisian '201. Nama penghuni' dengan verifikasi scan halaman
            max_nama_attempts = 5
            nama_terisi_sukses = False

            # Cek terlebih dahulu apakah '201. Nama penghuni' sudah terisi di UI
            if verifikasi_nama_penghuni_terisi(d):
                print(f"[SCAN NAMA] [SUKSES] Field '201. Nama penghuni' sudah terisi di UI. Mengabaikan pengetikan ulang.")
                nama_terisi_sukses = True
            elif not nama:
                print(f"[SCAN NAMA] Kolom NAMA di Excel kosong, melanjutkan proses tanpa mengisi field NAMA.")
                nama_terisi_sukses = True
            else:
                for nama_attempt in range(1, max_nama_attempts + 1):
                    print(f"[INPUT NAMA] Memasukkan '201. Nama penghuni': '{nama}' (Percobaan {nama_attempt}/{max_nama_attempts})...")
                    input_textbox(label_text="201. Nama penghuni", value=nama, bounds_fallback=None, exact=False, sleep_after=SLEEP_SHORT)
                    time.sleep(0.5)

                    if verifikasi_nama_penghuni_terisi(d):
                        print(f"[SCAN NAMA] [SUKSES] Field '201. Nama penghuni' terisi dengan sukses pada percobaan ke-{nama_attempt}.")
                        nama_terisi_sukses = True
                        break
                    else:
                        print(f"[SCAN NAMA] [GAGAL] Field '201. Nama penghuni' belum terisi (percobaan ke-{nama_attempt}/{max_nama_attempts}). Mengulangi pengisian...")
                        time.sleep(0.5)

            if not nama_terisi_sukses:
                print(f"[SCAN NAMA] [ERROR] Field '201. Nama penghuni' tetap gagal terisi setelah {max_nama_attempts}x percobaan.")

            # Pengisian '202. NIK penghuni' dan 'Cek NIK' dengan verifikasi scan halaman untuk 'NIK tidak valid'
            max_nik_attempts = 5
            nik_sukses = False

            # Cek terlebih dahulu apakah field NIK berisi nilai dummy '9999999999999998'
            nik_existing = baca_nilai_field_nik(d)
            if nik_existing == "9999999999999998":
                print(f"[SCAN NIK] Field '202. NIK penghuni' berisi dummy '9999999999999998'. Akan diganti dengan NIK dari Excel: '{nik}'")

            for nik_attempt in range(1, max_nik_attempts + 1):
                # Baca nilai field NIK saat ini (kecuali percobaan pertama yang sudah dibaca di atas)
                if nik_attempt > 1:
                    nik_existing = baca_nilai_field_nik(d)

                # Tentukan apakah perlu mengisi ulang NIK:
                # - Jika field berisi dummy '9999999999999998'
                # - Jika field kosong
                # - Jika field berbeda dengan NIK dari Excel
                perlu_isi = False
                if not nik_existing:
                    print(f"[SCAN NIK] Field NIK kosong. Perlu diisi dengan NIK dari Excel: '{nik}' (Percobaan {nik_attempt}/{max_nik_attempts})")
                    perlu_isi = True
                elif nik_existing == "9999999999999998":
                    print(f"[SCAN NIK] Field NIK masih berisi dummy '9999999999999998'. Perlu diganti dengan NIK dari Excel: '{nik}' (Percobaan {nik_attempt}/{max_nik_attempts})")
                    perlu_isi = True
                elif nik_existing != nik:
                    print(f"[SCAN NIK] Field NIK berisi '{nik_existing}' (berbeda dengan NIK Excel '{nik}'). Perlu diisi ulang (Percobaan {nik_attempt}/{max_nik_attempts})")
                    perlu_isi = True
                else:
                    print(f"[SCAN NIK] Field NIK sudah berisi '{nik_existing}' sesuai dengan NIK Excel. Lanjut ke Cek NIK.")

                if perlu_isi:
                    print(f"[INPUT NIK] Memasukkan '202. NIK penghuni': '{nik}' (Percobaan {nik_attempt}/{max_nik_attempts})...")
                    input_textbox(label_text="202. NIK penghuni", value=nik, bounds_fallback=None, exact=False, sleep_after=SLEEP_SHORT)
                    time.sleep(0.5)

                    # Verifikasi setelah pengisian: baca ulang field NIK
                    nik_setelah_isi = baca_nilai_field_nik(d)
                    if not nik_setelah_isi or nik_setelah_isi != nik:
                        print(f"[SCAN NIK] [VERIFIKASI GAGAL] Setelah pengisian, field NIK berisi '{nik_setelah_isi}' (seharusnya '{nik}'). Melakukan swipe ke bawah & ke atas...")
                        loop_swipe_statis(delta_y=-400, loop=1)
                        time.sleep(0.3)
                        loop_swipe_statis(delta_y=400, loop=1)
                        time.sleep(0.5)
                        continue  # Ulangi loop tanpa menekan Cek NIK
                    else:
                        print(f"[SCAN NIK] [VERIFIKASI OK] Field NIK berisi '{nik_setelah_isi}' sesuai dengan NIK Excel.")

                # Cari terlebih dahulu tombol "Cek NIK" sebelum mengetuknya (swipe max 10x jika belum ketemu)
                cari_dan_scroll_ke_tombol_cek_nik(max_swipes=10)
                print(f"[KLIK] Mengetuk 'Cek NIK' (Percobaan {nik_attempt}/{max_nik_attempts})...")
                ketuk("Cek NIK")
                tunggu_loading_cek_nik(timeout=30, sleep_before=0.3)

                is_nik_invalid, msg_error = cek_nik_tidak_valid(d)
                if is_nik_invalid:
                    print(f"[SCAN NIK] [RETRY NIK] Terdeteksi '{msg_error}' (percobaan ke-{nik_attempt}/{max_nik_attempts}). Melakukan swipe ke bawah & ke atas...")
                    loop_swipe_statis(delta_y=-400, loop=1)
                    time.sleep(0.3)
                    loop_swipe_statis(delta_y=400, loop=1)
                    time.sleep(SLEEP_SHORT)
                else:
                    print(f"[SCAN NIK] [SUKSES] NIK '{nik}' berhasil dicek (tidak ada pesan NIK tidak valid) pada percobaan ke-{nik_attempt}.")
                    nik_sukses = True
                    break

            if not nik_sukses:
                print(f"[SCAN NIK] [GAGAL] NIK '{nik}' tetap 'NIK tidak valid' setelah {max_nik_attempts}x percobaan. Menyimpan status Excel & berpindah ke baris berikutnya...")
                simpan_status_excel(row, "Error : Nik tidak valid")
                kembali_ke_daftar_assignment()
                sukses_baris = True
                break

            loop_swipe_dinamis(delta_y=-700, target_text="204. Status kepemilikan")
            input_textbox(label_text="203. Nomor telepon", value='-', bounds_fallback=None, exact=False, sleep_after=SLEEP_SHORT)
            
            # Pengecekan status RadioButton: Ketuk '1. Milik sendiri' HANYA jika belum tercentang
            if cek_radio_button_tercentang("1. Milik sendiri", exact=False):
                print("[RADIO CHECK] RadioButton '1. Milik sendiri' sudah tercentang.")
            else:
                print("[RADIO CHECK] RadioButton '1. Milik sendiri' BELUM tercentang. Memulai pengetukan...")
                max_radio_attempts = 5
                for r_attempt in range(1, max_radio_attempts + 1):
                    clicked = False

                    # Strategi 1: Kalkulasi koordinat dinamis dari parent/label bounds
                    try:
                        print("strategi 1")
                        label_el = None
                        for pattern in ["1. Milik sendiri", "1.  Milik sendiri", "Milik sendiri"]:
                            elements = d.xpath(f"//*[contains(@text, '{pattern}')]").all()
                            for el in elements:
                                bounds_str = el.attrib.get('bounds')
                                import re
                                pts = [int(x) for x in re.findall(r'\d+', bounds_str)]
                                if len(pts) == 4:
                                    x1, y1, x2, y2 = pts
                                    if (x2 - x1) > 50:
                                        label_el = el
                                        break
                            if label_el:
                                break

                        if label_el:
                            label_bounds_str = label_el.attrib.get('bounds')
                            import re
                            l_pts = [int(x) for x in re.findall(r'\d+', label_bounds_str)]
                            label_left, label_top, label_right, label_bottom = l_pts

                            label_text = label_el.text
                            parent = d.xpath(f"//*[contains(@text, '{label_text}')]/..")

                            if parent.exists:
                                parent_bounds = parent.all()[0].attrib.get('bounds')
                                p_pts = [int(x) for x in re.findall(r'\d+', parent_bounds)]
                                parent_left, parent_top, parent_right, parent_bottom = p_pts

                                click_x = parent_left + (label_left - parent_left) // 2
                                click_y = label_top + (label_bottom - label_top) // 2

                                print(f"[KLIK] Parent Left: {parent_left}, Label Left: {label_left}")
                                print(f"[KLIK] Mengklik koordinat dinamis radio button: ({click_x}, {click_y})")
                                d.click(click_x, click_y)
                                time.sleep(SLEEP_SHORT)
                                clicked = True
                    except Exception as click_err:
                        print(f"[WARNING] Percobaan klik koordinat dinamis gagal: {click_err}")

                    if not clicked:
                        print(f"[WARNING] Strategi 1 klik gagal pada percobaan ke-{r_attempt}.")

                    # Verifikasi apakah sudah tercentang
                    time.sleep(0.5)
                    if cek_radio_button_tercentang("1. Milik sendiri", exact=False):
                        print(f"[RADIO SUCCESS] Berhasil memverifikasi '1. Milik sendiri' tercentang pada percobaan ke-{r_attempt}.")
                        break
                    print(f"[RETRY RADIO] '1. Milik sendiri' belum tercentang (percobaan ke-{r_attempt}/{max_radio_attempts}). Mengulangi...")
                    time.sleep(SLEEP_SHORT)
                    loop_swipe_statis(delta_y=-100, loop=1)
                    time.sleep(SLEEP_SHORT)


            # Blok III
            ketuk_sidebar_toggle()
            time.sleep(SLEEP_LONG)
            pilih_blok("III")
            time.sleep(SLEEP_LONG)
            
            # 301. alamat saat ini
            isi_blok_iii()
            time.sleep(SLEEP_LONG)

            # ketuk tombol kontrol Increment
            ketuk_tombol_increment()

            # Blok IV
            ketuk_sidebar_toggle()
            time.sleep(SLEEP_LONG)
            pilih_blok("IV")
            time.sleep(SLEEP_LONG)

            #Catatan
            input_textbox(label_text="Catatan", value='-', bounds_fallback=None, exact=False, sleep_after=SLEEP_SHORT)

            #18 ketuk tombol "Kirim" pertama
            ketuk("Kirim", sleep_after=SLEEP_SHORT)
            time.sleep(SLEEP_SHORT)

            #19 ketuk tombol "YA" (jika muncul konfirmasi YA)
            if check_exists(d(text="YA")) or check_exists(d(textContains="YA")):
                ketuk("YA", sleep_after=SLEEP_SHORT)
                time.sleep(SLEEP_SHORT)

            # Cek apakah modal ringkasan validasi menampilkan 'GALAT 0' (validasi error = 0)
            is_galat_0 = False
            print("[SUBMIT] Memeriksa status 'GALAT 0' pada modal ringkasan validasi...")
            for galat_attempt in range(15):
                if (check_exists(d(textContains="GALAT 0 Perlu diperbaiki")) or 
                    check_exists(d(textContains="GALAT 0")) or 
                    check_exists(d(descriptionContains="GALAT 0")) or 
                    check_exists(d.xpath("//*[contains(@text, 'GALAT 0') or contains(@content-desc, 'GALAT 0')]"))):
                    is_galat_0 = True
                    break
                time.sleep(0.3)

            if not is_galat_0:
                print(f"[SUBMIT] [SKIP] Teks 'GALAT 0' tidak terdeteksi (terdapat error validasi) untuk IDPEL {idpel}. Menyimpan status Excel 'koordinat & foto tidak ada' & berpindah ke baris berikutnya...")
                simpan_status_excel(row, "koordinat & foto tidak ada")
                kembali_ke_daftar_assignment()
                sukses_baris = True
                break

            print("[SUBMIT] [SUKSES] Terdeteksi 'GALAT 0'! Mengetuk tombol 'Kirim' kedua...")
            #20 ketuk tombol "Kirim" kedua di dalam modal
            ketuk("Kirim", sleep_after=SLEEP_SHORT)
            time.sleep(SLEEP_SHORT)

            #21 ketuk tombol "Konfirmasi"
            ketuk("Konfirmasi", sleep_after=SLEEP_SHORT)
            time.sleep(SLEEP_SHORT)

            #23 ketuk tombol "Konfirmasi"
            ketuk("Konfirmasi", sleep_after=SLEEP_SHORT)
            time.sleep(SLEEP_SHORT)

            #24 ketuk tombol "YA" (ulangi jika 'Submit diproses' belum muncul)
            max_ya_attempts = 5
            submit_muncul = False
            for ya_attempt in range(1, max_ya_attempts + 1):
                print(f"[SUBMIT] Mengetuk tombol 'YA' (Percobaan {ya_attempt}/{max_ya_attempts})...")
                ketuk("YA", sleep_after=SLEEP_SHORT)
                time.sleep(SLEEP_MEDIUM)

                is_submit_modal = (
                    check_exists(d(textContains="Submit diproses")) or
                    check_exists(d(resourceId="id.go.bpsfasih:id/tv_submit_progress_title")) or
                    check_exists(d(resourceId="id.go.bpsfasih:id/btn_submit_progress_close")) or
                    check_exists(d(text="OK"))
                )

                if is_submit_modal:
                    print(f"[SUBMIT] [SUKSES] Modal 'Submit diproses' terdeteksi setelah ketuk 'YA' pada percobaan ke-{ya_attempt}.")
                    submit_muncul = True
                    break
                else:
                    print(f"[SUBMIT] [RETRY] Modal 'Submit diproses' belum muncul (percobaan ke-{ya_attempt}/{max_ya_attempts}). Mengulangi ketuk 'YA'...")
                    time.sleep(SLEEP_SHORT)

            if not submit_muncul:
                print(f"[WARNING] Modal 'Submit diproses' tetap tidak terdeteksi setelah {max_ya_attempts}x mengetuk 'YA'. Memeriksa modal secara langsung...")

            #25 ketuk "OK" pada modal Submit diproses (retry 5x & fallback statis bounds 528, 1691)
            ketuk_ok_submit_diproses(max_attempts=5)

            #25 Scan teks "Halaman Upload" -> jika muncul maka tekan tombol "BACK" pada emulator
            print("[EMULATOR] Memeriksa apakah teks 'Halaman Upload' sudah muncul di layar...")
            is_halaman_upload = False
            for _ in range(5):
                if (check_exists(d(textContains="Halaman Upload")) or 
                    check_exists(d(descriptionContains="Halaman Upload")) or 
                    check_exists(d.xpath("//*[contains(@text, 'Halaman Upload') or contains(@content-desc, 'Halaman Upload')]"))):
                    is_halaman_upload = True
                    break
                time.sleep(0.5)

            if is_halaman_upload:
                print("[EMULATOR] Teks 'Halaman Upload' terdeteksi di layar! Menekan tombol Back pada emulator...")
                d.press("back")
                time.sleep(SLEEP_LONG)
            else:
                print("[EMULATOR] Teks 'Halaman Upload' tidak terdeteksi di layar.")

            #24. Tunggu kembali ke halaman 'Daftar Assignment'
            sukses_da = tunggu_loading("Daftar Assignment", timeout=15)
            if not sukses_da:
                print(f"[WARNING] Teks 'Daftar Assignment' tidak terdeteksi setelah 15 detik untuk baris {row} (IDPEL: {idpel}).")
                print(f"[RECOVERY] Mengurungkan simpan_status_excel, mengeksekusi kembali_ke_daftar_assignment()...")
                kembali_ke_daftar_assignment()
                if row_attempt < 3:
                    print(f"[RETRY BARIS] Mengulangi pemrosesan baris {row} ({idpel}) dari awal (percobaan ke-{row_attempt + 1})...")
                    continue
                else:
                    print(f"[RETRY GAGAL] Sudah mencoba 3x untuk baris {row} ({idpel}). Melanjutkan ke baris berikutnya.")
                    break
            else:
                #25 Simpan status di Excel 'SUKSES DIUPDATE' HANYA jika sukses_da True
                simpan_status_excel(row, "SUKSES DIUPDATE")
                sukses_baris = True
                break

            # # Pause proses dengan pilihan stop seluruh proses atau lanjutkan proses
            # if not konfirmasi_stop_atau_lanjut("Pause sebelum mengirim data (Kirim). Lanjutkan atau Stop seluruh proses?"):
            #     print(f"[HALT] Seluruh proses bot dihentikan secara manual oleh pengguna pada baris {row} (IDPEL: {idpel}).")
            #     return

def main():
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

    print("==================================================")
    print("    BOT EMULATOR REJECT INPUT - FASIH SCRAPPER    ")
    print("==================================================")

    if not hubungkan_emulator():
        return

    proses_update_reject_nik()

    print("\n==================================================")
    print("   PROSES BOT REJECT INPUT SELESAI DENGAN SUKSES  ")
    print("==================================================")


if __name__ == "__main__":
    main()
