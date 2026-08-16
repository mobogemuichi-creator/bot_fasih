"""
Script Otomatisasi Cek Status Nomor Meter di FASIH (cek_nometer_1.py)
-------------------------------------------------------------------
Membaca Nomor Meter dari kolom A pada file Excel 'data_cek.xlsx',
memeriksa status keberadaan Nomor Meter di aplikasi FASIH,
dan menyimpan hasilnya ke file 'cek_nik.txt' dengan format 'nometer|status_exist'.
Jika Nomor Meter sudah pernah tercatat di 'cek_nik.txt', script akan otomatis di-skip.
"""
import os
import time
import subprocess
import openpyxl
import uiautomator2 as u2
import sys
import datetime

from konfigurasi import (
    LDPLAYER_ADB,
    EMULATOR_PORTS_1 as EMULATOR_PORTS,
    SLEEP_SHORT,
    SLEEP_MEDIUM,
    SLEEP_LONG,
)

d = None

def hubungkan_emulator():
    """Menghubungkan ke emulator via uiautomator2"""
    global d
    print("[KONEKSI] Mencoba menghubungkan ke emulator LDPlayer...")
    
    for port in EMULATOR_PORTS:
        try:
            if os.path.exists(LDPLAYER_ADB):
                subprocess.run([LDPLAYER_ADB, "connect", f"127.0.0.1:{port}"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=2)
            temp_d = u2.connect(f"127.0.0.1:{port}")
            info = temp_d.info
            d = temp_d
            print(f"[KONEKSI] Berhasil terhubung ke emulator di port {port}!")
            print(f"[EMULATOR] Layar: {info.get('displayWidth')}x{info.get('displayHeight')} ({info.get('productName', 'Android')})")
            return True
        except Exception:
            continue

    try:
        temp_d = u2.connect()
        info = temp_d.info
        d = temp_d
        print("[KONEKSI] Berhasil terhubung ke emulator via default connection!")
        return True
    except Exception:
        pass
            
    print("[ERROR] Gagal terhubung ke emulator. Pastikan LDPlayer sudah berjalan.")
    return False


def swipe_ke_bawah_cepat(times=3):
    """Men-scroll layar ke bawah dengan cepat menggunakan swipe"""
    for _ in range(times):
        try:
            d.swipe(540, 1300, 540, 200, duration=0.05)
            time.sleep(0.05)
        except Exception:
            pass


def baca_nometer_terproses(txt_path="cek_nik.txt"):
    """
    Membaca file txt 'cek_nik.txt' jika sudah ada,
    dan mengembalikan set dari Nomor Meter yang sudah dipproses.
    Format file txt per baris: nometer|status_exist
    """
    terproses = set()
    if os.path.exists(txt_path):
        try:
            with open(txt_path, "r", encoding="utf-8") as f:
                for line in f:
                    line_str = line.strip()
                    if line_str and "|" in line_str:
                        parts = line_str.split("|")
                        meter_key = parts[0].strip()
                        if meter_key:
                            terproses.add(meter_key)
            print(f"[TXT] Ditemukan {len(terproses)} Nomor Meter yang sudah tercatat di '{txt_path}'.")
        except Exception as e:
            print(f"[WARNING] Gagal membaca '{txt_path}': {e}")
    else:
        print(f"[TXT] File '{txt_path}' belum ada, akan dibuat secara otomatis saat menyimpan hasil.")
    return terproses


def simpan_ke_txt(nometer, status_exist, txt_path="cek_nik.txt"):
    """Menyimpan status Nomor Meter ke file 'cek_nik.txt' dengan format 'nometer|status_exist'"""
    try:
        with open(txt_path, "a", encoding="utf-8") as f:
            f.write(f"{nometer}|{status_exist}\n")
        print(f"[TXT] Berhasil menyimpan '{nometer}|{status_exist}' ke '{txt_path}'.")
        return True
    except Exception as e:
        print(f"[ERROR TXT] Gagal menyimpan ke '{txt_path}': {e}")
        return False


def proses_cek_nometer(excel_path="data_cek.xlsx", txt_path="cek_nik.txt"):
    """Fungsi utama membaca data_cek.xlsx kolom A dan melakukan pengecekan Nomor Meter di Fasih"""
    if not os.path.exists(excel_path):
        print(f"[ERROR] File Excel '{excel_path}' tidak ditemukan!")
        return

    # 1. Baca Nomor Meter yang sudah diproses sebelumnya dari cek_nik.txt
    sudah_dibaca = baca_nometer_terproses(txt_path)

    # 2. Baca data dari data_cek.xlsx (Kolom A)
    print(f"[EXCEL] Membaca data Nomor Meter dari kolom A pada '{excel_path}'...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
    except Exception as e:
        print(f"[ERROR] Gagal membaca file Excel '{excel_path}': {e}")
        return

    max_row = sheet.max_row
    print(f"[EXCEL] Total baris di Sheet: {max_row}")

    for row in range(2, max_row + 1):
        meter_raw = sheet.cell(row=row, column=1).value  # Kolom A
        if meter_raw is None:
            continue

        # Clean string nometer
        if isinstance(meter_raw, float):
            nometer = str(int(meter_raw)).strip()
        else:
            nometer = str(meter_raw).strip()

        if not nometer or nometer.lower() in ["none", "nan"]:
            continue

        # 3. Cek apakah Nomor Meter sudah dibaca di cek_nik.txt
        if nometer in sudah_dibaca:
            print(f"[SKIP] Baris {row} | NOMETER {nometer} sudah tercatat di '{txt_path}'. Dilewati.")
            continue

        print(f"\n--------------------------------------------------")
        print(f"Memproses Baris {row} | NOMETER: {nometer}")
        print(f"--------------------------------------------------")

        try:
            # ==========================================
            # NAVIGASI DASHBOARD (Referensi bot_emulator_meter.py)
            # ==========================================
            is_on_form = d(textContains="101b. Nomor meter").exists() or d(text="Cek Nomor Meter").exists()

            if not is_on_form:
                # Transisi ke Daftar Assignment (Self-Healing)
                region_row = d(resourceId="id.go.bpsfasih:id/updateListingLayout")
                if region_row.exists():
                    print("[DASHBOARD] Berada di halaman SLS Wilayah. Mengklik wilayah untuk masuk...")
                    region_row.click()
                    time.sleep(SLEEP_LONG)

                # Klik FAB & Tambah Assignment
                ya_btn_found = False
                for attempt in range(1, 4):
                    print(f"[DASHBOARD] Mengklik FAB (Percobaan {attempt}/3)...")
                    fab = d(resourceId="id.go.bpsfasih:id/expendable_fab")
                    if fab.wait(exists=True, timeout=5):
                        fab.click()
                        time.sleep(SLEEP_SHORT)
                    else:
                        print(f"[WARNING] Tombol FAB tidak ditemukan pada percobaan {attempt}.")
                        time.sleep(1.0)
                        continue

                    print(f"[DASHBOARD] Mengklik 'Tambah Assignment' (Percobaan {attempt}/3)...")
                    tambah_btn = d(resourceId="id.go.bpsfasih:id/fab_addAssignment")
                    if not tambah_btn.exists():
                        tambah_btn = d(text="Tambah Assignment")
                    if tambah_btn.exists(timeout=5):
                        tambah_btn.click()
                        time.sleep(SLEEP_MEDIUM)
                    else:
                        print(f"[WARNING] Tombol 'Tambah Assignment' tidak ditemukan pada percobaan {attempt}.")
                        time.sleep(1.0)
                        continue

                    print("[DASHBOARD] Menyetujui konfirmasi dialog 'YA'...")
                    ya_btn = d(resourceId="id.go.bpsfasih:id/rButton_bottomDialog")
                    if not ya_btn.exists(): ya_btn = d(text="YA")
                    if not ya_btn.exists(): ya_btn = d(text="Ya")
                    if ya_btn.exists(timeout=5):
                        ya_btn.click()
                        print("Menunggu form termuat (5 detik)...")
                        time.sleep(5.0)
                        ya_btn_found = True
                        break
                    else:
                        print(f"[WARNING] Tombol konfirmasi 'YA' tidak muncul (Percobaan {attempt}/3). Mengulangi...")
                        time.sleep(1.0)

                if not ya_btn_found:
                    raise Exception("Tombol konfirmasi 'YA' tidak ditemukan setelah 3 kali percobaan (FAB -> Tambah Assignment).")

                # Scan & Ambil Waktu
                found_scan_waktu = False
                for scan_attempt in range(1, 11):
                    print(f"[BLOK I] Scanning teks 'Ambil Waktu' (Percobaan {scan_attempt}/10)...")
                    if d(text="Ambil Waktu").wait(exists=True, timeout=1.0) or d(textContains="Ambil Waktu").exists():
                        found_scan_waktu = True
                        print(f"[BLOK I] Teks 'Ambil Waktu' ditemukan pada percobaan {scan_attempt}.")
                        break

                if not found_scan_waktu:
                    print(f"[SKIP] Baris {row} | NOMETER {nometer} dilewati karena teks 'Ambil Waktu' tidak ditemukan.")
                    continue

                print("[BLOK I] Mengetuk tombol 'Ambil Waktu'...")
                waktu_btn = d(text="Ambil Waktu")
                if waktu_btn.exists():
                    waktu_sukses = False
                    for attempt_waktu in range(1, 4):
                        print(f"[BLOK I] Mengetuk tombol 'Ambil Waktu' (Percobaan {attempt_waktu}/3)...")
                        waktu_btn.click()
                        time.sleep(SLEEP_SHORT)
                        
                        dialog_waktu = d(textContains="Apakah Anda yakin ingin mengambil waktu saat ini")
                        if not dialog_waktu.exists():
                            dialog_waktu = d(textContains="mengambil waktu saat ini")
                        
                        if dialog_waktu.wait(exists=True, timeout=2.0) or d(text="ya").exists() or d(text="Ya").exists() or d(text="YA").exists():
                            print("[BLOK I] Teks konfirmasi waktu terdeteksi. Mengetuk tombol 'ya'...")
                            ya_waktu = d(text="ya")
                            if not ya_waktu.exists(): ya_waktu = d(text="Ya")
                            if not ya_waktu.exists(): ya_waktu = d(text="YA")
                            if ya_waktu.exists():
                                ya_waktu.click()
                                time.sleep(SLEEP_MEDIUM)
                            waktu_sukses = True
                            break
                        else:
                            print(f"[WARNING] Dialog konfirmasi waktu tidak muncul. Mengulangi...")
                            time.sleep(1.0)
                    if not waktu_sukses:
                        print("[WARNING] Dialog konfirmasi waktu tidak muncul setelah 3 kali percobaan.")
                else:
                    print("[INFO] Waktu wawancara sudah terisi. Melewati langkah Ambil Waktu.")

                # Scroll cepat ke bawah sampai ke posisi 101b. Nomor meter
                print("[BLOK I] Men-scroll cepat ke bawah sampai posisi Nomor Meter...")
                swipe_ke_bawah_cepat(times=5)
                time.sleep(SLEEP_SHORT)
            else:
                print("[INFO] Sudah berada di form pengisian Nomor Meter. Melewati langkah awal.")

            # Input Nomor Meter
            print("[BLOK I] Mencari kolom input Nomor Meter...")
            id_label = d(textContains="101b. Nomor meter")
            if not id_label.exists():
                try:
                    d(scrollable=True).scroll.to(textContains="101b. Nomor meter")
                except Exception:
                    pass

            if id_label.wait(exists=True, timeout=5):
                id_input = id_label.down(className="android.widget.EditText")
                if id_input.exists():
                    print(f"[BLOK I] Menulis Nomor Meter via set_text: {nometer}...")
                    id_input.set_text(str(nometer))
                    time.sleep(SLEEP_SHORT)
                else:
                    raise Exception("Kolom input EditText Nomor Meter tidak ditemukan.")
            else:
                raise Exception("Label '101b. Nomor meter' tidak ditemukan.")

            # Kirim ENTER
            d.shell("input keyevent KEYCODE_ENTER")
            time.sleep(SLEEP_SHORT)

            # Ketuk tombol "Cek Nomor Meter"
            print("[BLOK I] Mengetuk tombol 'Cek Nomor Meter'...")
            cek_id_btn = d(text="Cek Nomor Meter")
            if not cek_id_btn.exists():
                try:
                    d(scrollable=True).scroll.to(text="Cek Nomor Meter")
                except Exception:
                    pass

            if cek_id_btn.exists(timeout=5):
                t_start = time.time()
                cek_id_btn.click()
                print(f"[{datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Mengetuk tombol 'Cek Nomor Meter'...")
                
                # Menunggu progress bar loading selesai
                print(f"[{datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Menunggu progress bar loading selesai...")
                time.sleep(0.1)
                try:
                    d(resourceId="id.go.bpsfasih:id/card_progress").wait_gone(timeout=30)
                except Exception:
                    pass
                t_loading = time.time()
                print(f"[{datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Progress bar loading selesai (+{t_loading - t_start:.2f}s dari awal)")
                
                # Cek apakah limit API check-nometerpln tercapai
                limit_msg = "Permintaan API check-nometerpln sudah terlampaui (limit)."
                if d(textContains=limit_msg).exists():
                    print(f"[{datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]}] PROSES BERHENTI KARENA CEK NO METER MELEWATI LIMIT API!")
                    input("Tekan ENTER untuk keluar...")
                    sys.exit(1)
                
                # Cek status hasil
                target_status = "DITEMUKAN DAN BELUM TERCATAT PADA SISTEM FASIH"
                already_recorded = "DITEMUKAN DAN SUDAH TERCATAT PADA SISTEM FASIH"
                
                time.sleep(0.1)
                status_found = d(textContains=target_status).exists()
                
                is_already_registered = (
                    d(textContains=already_recorded).exists() or
                    d(textContains="Nomor Meter sudah terdaftar di FASIH").exists() or
                    d(textContains="Nomor meter sudah terdaftar di FASIH").exists() or
                    d(textContains="sudah terdaftar di FASIH").exists()
                )

                is_not_found = (
                    d(textContains="Nomor meter tidak ditemukan").exists() or
                    d(textContains="Nomor Meter tidak ditemukan").exists() or
                    d(textContains="nomor meter tidak ditemukan").exists() or
                    d(textContains="tidak ditemukan").exists() or
                    d(text="STATUS").exists()
                )

                status_exist = "TIDAK DITEMUKAN / UNKNOWN"
                if status_found:
                    status_exist = "DITEMUKAN DAN BELUM TERCATAT PADA SISTEM FASIH"
                    print(f"[STATUS] NOMETER {nometer} -> {status_exist}")
                elif is_already_registered:
                    status_exist = "SUDAH TERCATAT PADA SISTEM FASIH"
                    print(f"[STATUS] NOMETER {nometer} -> {status_exist}")
                elif is_not_found:
                    status_exist = "Nomor meter tidak ditemukan"
                    print(f"[STATUS] NOMETER {nometer} -> {status_exist}")
                else:
                    print(f"[WARNING] Status respon khusus tidak terdeteksi secara eksplisit untuk NOMETER {nometer}.")
                    status_exist = "BELUM TERDAFTAR FASIH"

                # Simpan ke cek_nik.txt & tambahkan ke memori
                simpan_ke_txt(nometer, status_exist, txt_path)
                sudah_dibaca.add(nometer)

            else:
                raise Exception("Tombol 'Cek Nomor Meter' tidak ditemukan di layar.")

        except Exception as err:
            print(f"[ERROR] Gagal memproses baris {row} (NOMETER {nometer}): {err}")
            err_msg = f"Error : {err}"
            simpan_ke_txt(nometer, err_msg, txt_path)
            sudah_dibaca.add(nometer)

def main():
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

    print("==================================================")
    print("   BOT CEK STATUS NOMOR METER FASIH (cek_nometer_1.py) ")
    print("==================================================")

    if not hubungkan_emulator():
        return

    proses_cek_nometer()

    print("\n==================================================")
    print("  PROSES CEK STATUS NOMOR METER SELESAI DENGAN SUKSES ")
    print("==================================================")

if __name__ == "__main__":
    main()
