import os
import time
import openpyxl
import subprocess
import sys
import math

import urllib.request
import urllib.parse
import re
import json

# Auto-install Playwright jika belum terinstall
try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("[INFO] Library 'playwright' belum terinstall. Menginstall sekarang...")
    subprocess.run([sys.executable, "-m", "pip", "install", "playwright"], check=True)
    subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"], check=True)
    from playwright.sync_api import sync_playwright

# Auto-install Pillow jika belum terinstall
try:
    from PIL import Image
except ImportError:
    print("[INFO] Library 'Pillow' belum terinstall. Menginstall sekarang...")
    subprocess.run([sys.executable, "-m", "pip", "install", "Pillow"], check=True)
    from PIL import Image

def get_nearest_pano_details(lat, lon):
    """Mencari detail panorama Street View terdekat secara progresif dengan radius bertingkat"""
    # Coba radius 50m, 100m, 200m, 500m, lalu 1000m (dibatasi agar tidak mengambil panorama terpencil/rusak)
    for radius in [50, 100, 200, 500, 1000]:
        url = "https://maps.googleapis.com/maps/api/js/GeoPhotoService.SingleImageSearch"
        pb = f'!1m5!1sapiv3!5sUS!11m2!1m1!1b0!2m4!1m2!3d{lat}!4d{lon}!2d{radius}!3m10!2m2!1sen!2sGB!9m1!1e2!11m4!1m3!1e2!2b1!3e2!4m10!1e1!1e2!1e3!1e4!1e8!1e6!5m1!1e2!6m1!1e2'
        
        query = urllib.parse.urlencode({'pb': pb, 'callback': 'callbackfunc'})
        full_url = f"{url}?{query}"
        
        try:
            req = urllib.request.Request(full_url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=10) as response:
                data = response.read().decode('utf-8')
                match = re.search(r'callbackfunc\((.*)\)', data, re.DOTALL)
                if match:
                    json_str = match.group(1)
                    coords = re.findall(r'\[null,null,(-?\d+\.\d+),(-?\d+\.\d+)\]', json_str)
                    pano_ids = re.findall(r'\[2,"([A-Za-z0-9_\-]+)"\]', json_str)
                    
                    plat = None
                    plng = None
                    pano_id = None
                    
                    if coords:
                        plat = float(coords[0][0])
                        plng = float(coords[0][1])
                    
                    for pid in pano_ids:
                        if len(pid) == 22:
                            pano_id = pid
                            break
                            
                    if plat is not None and plng is not None:
                        print(f"  - Panorama ditemukan di radius {radius}m.")
                        return plat, plng, pano_id
        except Exception as e:
            print(f"  - [WARNING] Gagal memanggil API panorama pada radius {radius}m: {e}")
            
    return None

from konfigurasi import (
    EXCEL_FILE_2 as EXCEL_FILE,
    FOTO_DIR,
    LOG_HITAM_FILE,
    OVERWRITE_EXISTING,
    ZOOM_FOV,
    BEARING_OFFSET,
    PITCH_VALUE,
)

def calculate_bearing(lat1, lon1, lat2, lon2):
    """Menghitung sudut kompas (bearing) dari titik 1 ke titik 2"""
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lon = math.radians(lon2 - lon1)
    
    y = math.sin(delta_lon) * math.cos(lat2_rad)
    x = math.cos(lat1_rad) * math.sin(lat2_rad) - \
        math.sin(lat1_rad) * math.cos(lat2_rad) * math.cos(delta_lon)
        
    bearing = math.atan2(y, x)
    return (math.degrees(bearing) + 360) % 360

def check_canvas_black(page):
    """Mengecek apakah canvas WebGL Google Maps Street View bernilai hitam (blank)"""
    try:
        return page.evaluate("""() => {
            const canvases = Array.from(document.querySelectorAll('canvas'));
            if (canvases.length === 0) return true;
            
            let canvas = canvases[0];
            let maxArea = 0;
            for (let c of canvases) {
                const rect = c.getBoundingClientRect();
                const area = rect.width * rect.height;
                if (area > maxArea) {
                    maxArea = area;
                    canvas = c;
                }
            }
            
            try {
                const testC = document.createElement('canvas');
                testC.width = 10;
                testC.height = 10;
                const ctx = testC.getContext('2d');
                ctx.drawImage(canvas, 0, 0, 10, 10);
                const imgData = ctx.getImageData(0, 0, 10, 10).data;
                
                let allBlack = true;
                for (let i = 0; i < imgData.length; i += 4) {
                    if (imgData[i] > 10 || imgData[i+1] > 10 || imgData[i+2] > 10) {
                        allBlack = false;
                        break;
                    }
                }
                return allBlack;
            } catch(e) {
                return false;
            }
        }""")
    except Exception:
        return False

def catat_idpel_hitam(idpel):
    """Mencatat IDPEL yang panoramanya hitam ke file idpel_gambar_hitam.txt tanpa duplikasi"""
    filename = LOG_HITAM_FILE
    existing = set()
    if os.path.exists(filename):
        try:
            with open(filename, "r", encoding="utf-8") as f:
                existing = set(line.strip() for line in f if line.strip())
        except Exception:
            pass
    
    if str(idpel) not in existing:
        try:
            with open(filename, "a", encoding="utf-8") as f:
                f.write(f"{idpel}\n")
            print(f"  - [LOG TXT] IDPEL {idpel} berhasil dicatat ke '{filename}'")
        except Exception as e:
            print(f"  - [WARNING] Gagal mencatat IDPEL ke TXT: {e}")

def get_idpel_hitam_set():
    """Membaca daftar IDPEL dari file idpel_gambar_hitam.txt sebagai set"""
    filename = LOG_HITAM_FILE
    if os.path.exists(filename):
        try:
            with open(filename, "r", encoding="utf-8") as f:
                return set(line.strip() for line in f if line.strip())
        except Exception:
            pass
    return set()

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] File Excel '{EXCEL_FILE}' tidak ditemukan!")
        return

    # Pengecekan apakah file Excel sedang dibuka di program lain (seperti Microsoft Excel)
    while True:
        try:
            with open(EXCEL_FILE, "r+"):
                pass
            break
        except (PermissionError, IOError):
            print(f"\n==================================================")
            print(f"[PERINGATAN] File '{EXCEL_FILE}' sedang DIBUKA di Excel/program lain!")
            print(f"Mohon CLOSE / TUTUP terlebih dahulu file '{EXCEL_FILE}'.")
            print(f"==================================================")
            input("Tekan ENTER di sini setelah Anda menutup file Excel tersebut untuk melanjutkan...")

    # Buat direktori foto jika belum ada
    if not os.path.exists(FOTO_DIR):
        os.makedirs(FOTO_DIR)
        print(f"[INFO] Membuat direktori foto di '{FOTO_DIR}'")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    print(f"[START] Memulai proses screenshot Street View...")
    
    with sync_playwright() as p:
        # Jalankan Chromium headless dengan SwiftShader & bypass bot detection agar rendering WebGL selalu aktif
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--ignore-gpu-blocklist",
                "--use-gl=angle",
                "--use-angle=swiftshader",
                "--disable-blink-features=AutomationControlled"
            ]
        )
        # Gunakan context dengan viewport & User-Agent normal untuk mem-bypass deteksi headless Google Maps
        context = browser.new_context(
            viewport={"width": 1280, "height": 720},
            locale="id-ID", # Set bahasa Indonesia agar tombol persetujuan cookie konsisten
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = context.new_page()

        # Force WebGL to preserve drawing buffer so screenshots are never black
        page.add_init_script("""
            const originalGetContext = HTMLCanvasElement.prototype.getContext;
            HTMLCanvasElement.prototype.getContext = function(type, attributes) {
                if (type === 'webgl' || type === 'experimental-webgl' || type === 'webgl2') {
                    const newAttributes = { ...(attributes || {}), preserveDrawingBuffer: true };
                    return originalGetContext.call(this, type, newAttributes);
                }
                return originalGetContext.apply(this, arguments);
            };
        """)

        # Intercept dan bersihkan nama jalan serta blokir data POI/Bisnis/Restoran
        def handle_all_requests(route):
            req_url = route.request.url
            
            # 1. Bersihkan nama jalan di photometa
            if "photometa" in req_url:
                try:
                    response = route.fetch()
                    body = response.text()
                    # Ganti semua pola ["Nama Jalan","id"] dengan ["","id"] untuk menyembunyikan label jalan
                    sanitized_body = re.sub(r'\["([^"]+)","([a-z]{2}(-[a-zA-Z]{2})?)"\]', r'["","\2"]', body)
                    route.fulfill(response=response, body=sanitized_body)
                except Exception:
                    route.continue_()
            
            # 2. Blokir request data POI/Bisnis (tanpa memblokir /vt/ yang berisiko membuat panorama hitam)
            elif any(x in req_url for x in ["/preview/", "/passiveassist", "/poi", "/place", "/entity"]):
                route.abort()
                
            else:
                route.continue_()

        # Tangkap semua request (**/*) untuk diproses di dalam fungsi python di atas
        page.route("**/*", handle_all_requests)


        total_sukses = 0
        total_skip = 0
        total_gagal = 0

        for row in range(2, sheet.max_row + 1):
            idpel = sheet.cell(row=row, column=1).value  # Column A
            lat = sheet.cell(row=row, column=5).value    # Column E (KOORDINAT_X)
            lng = sheet.cell(row=row, column=6).value    # Column F (KOORDINAT_Y)

            if not idpel:
                break

            # Bersihkan idpel dari tipe float jika dibaca sebagai angka berkoma
            if isinstance(idpel, float):
                idpel = int(idpel)
            idpel = str(idpel).strip()

            # Lewati jika IDPEL ini sudah tercatat di idpel_gambar_hitam.txt
            if idpel in get_idpel_hitam_set():
                print(f"[SKIP] IDPEL {idpel} | Tercatat sebagai gambar hitam di '{LOG_HITAM_FILE}'.")
                total_skip += 1
                continue

            # Cek jika file foto sudah ada (baik .jpg maupun .png)
            dest_path_jpg = os.path.join(FOTO_DIR, f"{idpel}.jpg")
            dest_path_png = os.path.join(FOTO_DIR, f"{idpel}.png")
            foto_exists = os.path.exists(dest_path_jpg) or os.path.exists(dest_path_png)

            if foto_exists:
                if OVERWRITE_EXISTING:
                    # Hapus file lama jika ada untuk digantikan dengan versi baru
                    if os.path.exists(dest_path_jpg):
                        try:
                            os.remove(dest_path_jpg)
                        except Exception:
                            pass
                    if os.path.exists(dest_path_png):
                        try:
                            os.remove(dest_path_png)
                        except Exception:
                            pass
                else:
                    print(f"[SKIP] IDPEL {idpel} | File foto sudah ada.")
                    total_skip += 1
                    continue

            if not lat or not lng:
                print(f"[SKIP] IDPEL {idpel} | Koordinat kosong.")
                total_skip += 1
                continue

            # Parsing koordinat float
            try:
                lat = float(lat)
                lng = float(lng)
            except ValueError:
                print(f"[SKIP] IDPEL {idpel} | Format koordinat salah.")
                total_skip += 1
                continue

            print(f"\n[ROW {row}] Memproses IDPEL {idpel} | Koordinat Target: {lat}, {lng}...")

            # Cari detail panorama terdekat (termasuk pano_id)
            pano_details = get_nearest_pano_details(lat, lng)

            if pano_details and pano_details[0] is not None:
                pano_lat, pano_lng, pano_id = pano_details
                # Hitung bearing dari koordinat pano ke target koordinat kita
                bearing = calculate_bearing(pano_lat, pano_lng, lat, lng)
                # Terapkan offset bearing kompas untuk menghindari posisi grid watermark Google
                bearing = (bearing + BEARING_OFFSET) % 360
                
                print(f"  - Pano terdekat ditemukan di: {pano_lat}, {pano_lng}")
                print(f"  - Sudut kamera asli: {bearing:.2f}°")
                if pano_id:
                    # Format lengkap dengan pano_id untuk memaksa Google Maps menampilkan Street View asli dengan config dinamis
                    url = f"https://www.google.com/maps/@{pano_lat},{pano_lng},3a,{ZOOM_FOV}y,{bearing:.2f}h,{PITCH_VALUE}t/data=!3m6!1e1!3m4!1s{pano_id}!2e0!7i16384!8i8192"
                else:
                    url = f"https://www.google.com/maps/@{pano_lat},{pano_lng},18z,3a,{ZOOM_FOV}y,{bearing:.2f}h,{PITCH_VALUE}t"
            else:
                print("  - Tidak menemukan panorama terdekat (maksimal 1000m). Melewati baris ini (tidak diambil screenshot).")
                total_gagal += 1
                continue
            
            try:
                page.goto(url)
                
                # Tunggu dan klik persetujuan cookie Google jika muncul
                try:
                    consent_btn = page.locator('button:has-text("Setuju"), button:has-text("Saya setuju"), button:has-text("Accept all"), button:has-text("Agree"), button:has-text("Terima semua")')
                    if consent_btn.count() > 0:
                        consent_btn.first.click(timeout=3000)
                        print("  - Persetujuan cookie disetujui.")
                        time.sleep(1)
                except Exception:
                    pass

                # Tunggu canvas Street View selesai dimuat
                page.wait_for_selector("canvas", timeout=15000)
                
                # Tunggu secara dinamis hingga canvas memiliki gambar (tidak hitam)
                print("  - Menunggu rendering gambar panorama selesai...", end="", flush=True)
                start_wait = time.time()
                is_ready = False
                for _ in range(30):  # Maksimal 15 detik (30 * 0.5s)
                    if not check_canvas_black(page):
                        is_ready = True
                        print(f" [RENDERING SELESAI dalam {time.time() - start_wait:.1f}s]")
                        break
                    time.sleep(0.5)
                
                # Jika masih hitam, coba berjalan melangkah di Street View (ArrowUp)
                if not is_ready:
                    print("\n  - [TRY WALK] Panorama hitam, mencoba berjalan/berpindah posisi di Street View...", end="", flush=True)
                    try:
                        page.focus("canvas")
                        for step in range(1, 4):  # Coba 3 langkah maju
                            page.keyboard.press("ArrowUp")
                            time.sleep(1.5)
                            if not check_canvas_black(page):
                                is_ready = True
                                print(f" [BERHASIL TERBUKA di langkah {step}]")
                                break
                    except Exception:
                        pass

                if not is_ready:
                    print(" [TIMEOUT / PANORAMA HITAM PERMANEN]")
                    catat_idpel_hitam(idpel)
                    total_gagal += 1
                    continue  # Melewati pengambilan screenshot jika panorama gagal ter-render

                # Berikan jeda 2 detik tambahan agar gambar selesai memuat resolusi tajam
                time.sleep(2)

                # Sembunyikan seluruh elemen navigasi, pencarian, dan overlay lainnya secara otomatis
                try:
                    js_hide_overlays = """
                    (function() {
                        const canvases = Array.from(document.querySelectorAll('canvas'));
                        if (canvases.length === 0) return;
                        
                        // Cari canvas dengan luas area terbesar (Street View utama)
                        let canvas = canvases[0];
                        let maxArea = 0;
                        for (let c of canvases) {
                            const rect = c.getBoundingClientRect();
                            const area = rect.width * rect.height;
                            if (area > maxArea) {
                                maxArea = area;
                                canvas = c;
                            }
                        }
                        
                        const pathSet = new Set();
                        let curr = canvas;
                        while (curr) {
                            pathSet.add(curr);
                            curr = curr.parentElement;
                        }
                        
                        const ignoreTags = new Set(['HTML', 'BODY', 'HEAD', 'STYLE', 'LINK', 'SCRIPT', 'META', 'TITLE']);
                        const allElements = document.getElementsByTagName('*');
                        for (let el of allElements) {
                            if (!ignoreTags.has(el.tagName)) {
                                if (!pathSet.has(el)) {
                                    el.style.setProperty('visibility', 'hidden', 'important');
                                } else {
                                    el.style.setProperty('visibility', 'visible', 'important');
                                }
                            }
                        }

                        // Sembunyikan khusus ikon POI, bisnis, restoran, penginapan, dan marker overlay di atas canvas
                        const poiSelectors = [
                            '[jsaction*="poi"]',
                            '[jsaction*="place"]',
                            '[jsaction*="card"]',
                            '[role="button"]',
                            'button[aria-label]',
                            '[data-category]',
                            '.gmnoscreen',
                            'div[style*="cursor: pointer"]',
                            'div[style*="cursor:pointer"]'
                        ];
                        const poiElements = document.querySelectorAll(poiSelectors.join(','));
                        for (let el of poiElements) {
                            if (!pathSet.has(el)) {
                                el.style.setProperty('visibility', 'hidden', 'important');
                                el.style.setProperty('opacity', '0', 'important');
                            }
                        }
                    })();
                    """
                    page.evaluate(js_hide_overlays)
                    time.sleep(1) # Jeda agar transisi penyembunyian selesai stabil
                except Exception as e:
                    print(f"  - [WARNING] Gagal menyembunyikan overlay: {e}")

                # Ambil screenshot dan simpan ke file JPG
                page.screenshot(path=dest_path_jpg, quality=90, type="jpeg")
                
                # Crop menjadi rasio 1:1 (Square) di bagian tengah gambar
                try:
                    img = Image.open(dest_path_jpg)
                    w, h = img.size
                    min_side = min(w, h)
                    left = (w - min_side) // 2
                    top = (h - min_side) // 2
                    right = left + min_side
                    bottom = top + min_side
                    
                    img_cropped = img.crop((left, top, right, bottom))
                    img_cropped.save(dest_path_jpg, "JPEG", quality=95)
                    print(f"  - [SUKSES] Screenshot di-crop 1:1 ({min_side}x{min_side}) & disimpan ke: {dest_path_jpg}")
                    total_sukses += 1
                except Exception as e:
                    print(f"  - [WARNING] Gagal melakukan crop gambar: {e}")
                    total_gagal += 1

            except Exception as e:
                print(f"  - [GAGAL] Terjadi kesalahan: {e}")
                total_gagal += 1

        browser.close()
    
    print("\n==================================================")
    print("PROSES SCREENSHOT STREET VIEW SELESAI!")
    print(f"Total Baris Sukses  : {total_sukses}")
    print(f"Total Baris Di-skip : {total_skip}")
    print(f"Total Baris Gagal   : {total_gagal}")
    print("==================================================")

if __name__ == "__main__":
    main()
